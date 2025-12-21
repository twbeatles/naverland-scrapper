#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
네이버 부동산 크롤러 Pro Plus v11.0 (전면 리팩토링)

v11.0 신규 기능:
- HiDPI 디스플레이 지원
- Toast 알림 시스템
- 예외 처리 및 타입 힌트 강화
- 코드 품질 대폭 개선
- UI/UX 반응형 레이아웃 최적화
- 스레드 안전성 강화
- 메모리 관리 개선

v10.5 기능:
- 현대적인 UI 테마 (Glassmorphism, 그라데이션)
- 다크/라이트 모드 개선
- 버그 수정 (매물 URL, bare except)
- 자동 로그 정리

v7.3 이전 기능:
- 가격 변동 추적 및 신규 매물 배지
- 고급 필터 및 엑셀 템플릿
- URL 일괄 등록
"""

import sys, os, re, json, csv, time, random, shutil, logging, sqlite3, webbrowser
from queue import Queue, Empty as QueueEmpty, Full as QueueFull
from pathlib import Path
from datetime import datetime
from threading import Lock
from typing import Optional, List, Dict, Any, Tuple
from logging.handlers import RotatingFileHandler
from json import JSONDecodeError
from urllib.error import URLError, HTTPError
from urllib.request import urlopen, Request
from socket import timeout as SocketTimeout

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QLineEdit, QComboBox, QSpinBox, QDoubleSpinBox,
    QCheckBox, QTableWidget, QTableWidgetItem, QTextBrowser, QProgressBar,
    QTabWidget, QGroupBox, QSplitter, QScrollArea, QFrame, QListWidget,
    QListWidgetItem, QHeaderView, QMessageBox, QFileDialog, QInputDialog, 
    QTimeEdit, QStatusBar, QMenu, QSystemTrayIcon, QStyle, QApplication,
    QDialog, QDialogButtonBox, QSlider, QAbstractItemView, QToolTip, QSizePolicy
)
from PyQt6.QtCore import Qt, QTimer, QTime, QThread, pyqtSignal, QUrl, QPoint
from PyQt6.QtGui import QAction, QColor, QShortcut, QKeySequence, QFont, QDesktopServices, QCursor

# 선택적 라이브러리
try:
    from bs4 import BeautifulSoup
    BS4_AVAILABLE = True
except ImportError:
    BS4_AVAILABLE = False

try:
    import undetected_chromedriver as uc
    UC_AVAILABLE = True
except ImportError:
    UC_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from plyer import notification
    NOTIFICATION_AVAILABLE = True
except ImportError:
    NOTIFICATION_AVAILABLE = False

try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    import matplotlib.dates as mdates
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

# ============ CONFIG ============
APP_VERSION = "v11.0"
APP_TITLE = f"🏠 네이버 부동산 크롤러 Pro Plus {APP_VERSION} (Refactored Edition)"

# 경로 설정 - 실행 파일 위치 기준으로 안정적으로 설정
def get_base_dir():
    """실행 파일의 디렉토리를 안정적으로 반환"""
    if getattr(sys, 'frozen', False):
        # PyInstaller로 패키징된 경우
        return Path(sys.executable).parent
    else:
        # 일반 Python 스크립트 실행
        return Path(__file__).resolve().parent

BASE_DIR = get_base_dir()
DATA_DIR = BASE_DIR / "data"
LOG_DIR = BASE_DIR / "logs"
DB_PATH = DATA_DIR / "complexes.db"
SETTINGS_PATH = DATA_DIR / "settings.json"
PRESETS_PATH = DATA_DIR / "presets.json"
HISTORY_PATH = DATA_DIR / "search_history.json"

# 디렉토리 생성 (에러 로깅 포함)
def ensure_directories():
    """필수 디렉토리 생성"""
    for d in [DATA_DIR, LOG_DIR]:
        try:
            d.mkdir(parents=True, exist_ok=True)
            print(f"[INFO] 디렉토리 확인: {d}")
        except Exception as e:
            print(f"[ERROR] 디렉토리 생성 실패: {d} - {e}")

ensure_directories()
print(f"[INFO] DB 경로: {DB_PATH}")

def cleanup_old_logs(days=30):
    """지정 일수 이상 오래된 로그 파일 정리"""
    try:
        from datetime import timedelta
        cutoff_date = datetime.now() - timedelta(days=days)
        removed_count = 0
        for log_file in LOG_DIR.glob("crawler_*.log*"):
            try:
                file_date_str = log_file.stem.replace("crawler_", "").split(".")[0]
                file_date = datetime.strptime(file_date_str, "%Y%m%d")
                if file_date < cutoff_date:
                    log_file.unlink()
                    removed_count += 1
            except (ValueError, OSError) as e:
                continue  # 파싱 실패 시 무시
        if removed_count > 0:
            print(f"[INFO] 오래된 로그 파일 {removed_count}개 삭제")
    except Exception as e:
        print(f"[WARN] 로그 정리 중 오류: {e}")

# 시작 시 오래된 로그 파일 정리
cleanup_old_logs()

CRAWL_SPEED_PRESETS = {
    "빠름": {"min": 1, "max": 2, "desc": "빠른 수집 (차단 위험)"},
    "보통": {"min": 3, "max": 5, "desc": "권장 속도"},
    "느림": {"min": 5, "max": 8, "desc": "안전한 수집"},
    "매우 느림": {"min": 8, "max": 12, "desc": "가장 안전"}
}

SHORTCUTS = {
    "start_crawl": "Ctrl+R", "stop_crawl": "Ctrl+Shift+R", 
    "save_excel": "Ctrl+S", "save_csv": "Ctrl+Shift+S",
    "refresh": "F5", "search": "Ctrl+F", "settings": "Ctrl+,",
    "quit": "Ctrl+Q", "minimize_tray": "Ctrl+M", "toggle_theme": "Ctrl+T"
}

# 거래유형 색상
TRADE_COLORS = {
    "매매": {"bg": "#FFEBEE", "fg": "#C62828", "dark_bg": "#4A1C1C", "dark_fg": "#FF8A80"},
    "전세": {"bg": "#E8F5E9", "fg": "#2E7D32", "dark_bg": "#1C4A1C", "dark_fg": "#69F0AE"},
    "월세": {"bg": "#E3F2FD", "fg": "#1565C0", "dark_bg": "#1C2A4A", "dark_fg": "#82B1FF"}
}

def get_complex_url(cid): return f"https://new.land.naver.com/complexes/{cid}?ms=37.5,127,16&a=APT&e=RETAIL"
def get_article_url(cid, aid): return f"https://new.land.naver.com/complexes/{cid}?articleId={aid}"

# ============ LOGGER ============
def setup_logger(name="realestate_crawler"):
    logger = logging.getLogger(name)
    if logger.handlers: return logger
    logger.setLevel(logging.DEBUG)
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S'))
    logger.addHandler(ch)
    fh = RotatingFileHandler(LOG_DIR / f"crawler_{datetime.now().strftime('%Y%m%d')}.log", maxBytes=10*1024*1024, backupCount=5, encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(name)s:%(lineno)d - %(message)s'))
    logger.addHandler(fh)
    return logger

def get_logger(name=None): return logging.getLogger(f"realestate_crawler.{name}" if name else "realestate_crawler")

# ============ HELPERS ============
class PriceConverter:
    @staticmethod
    def to_int(price_str):
        if not price_str: return 0
        price_str = str(price_str).replace(",", "").replace(" ", "").strip()
        total = 0
        if "억" in price_str:
            parts = price_str.split("억")
            try: total += int(float(parts[0])) * 10000
            except (ValueError, TypeError): pass
            if len(parts) > 1 and parts[1]:
                remain = parts[1].replace("만", "").strip()
                if remain:
                    try: total += int(float(remain))
                    except (ValueError, TypeError): pass
        elif "만" in price_str:
            try: total = int(float(price_str.replace("만", "").strip()))
            except (ValueError, TypeError): pass
        else:
            try: total = int(float(price_str))
            except (ValueError, TypeError): pass
        return total
    
    @staticmethod
    def to_string(price_int):
        if price_int >= 10000:
            uk, man = price_int // 10000, price_int % 10000
            return f"{uk}억 {man:,}만" if man else f"{uk}억"
        elif price_int > 0:
            return f"{price_int:,}만"
        return "0"

class AreaConverter:
    PYEONG_RATIO = 0.3025
    @classmethod
    def sqm_to_pyeong(cls, sqm): return round(sqm * cls.PYEONG_RATIO, 1)
    @classmethod
    def pyeong_to_sqm(cls, pyeong): return round(pyeong / cls.PYEONG_RATIO, 2)

class DateTimeHelper:
    @staticmethod
    def now_string(fmt="%Y-%m-%d %H:%M:%S"): return datetime.now().strftime(fmt)
    @staticmethod
    def file_timestamp(): return datetime.now().strftime("%Y%m%d_%H%M%S")

class NaverURLParser:
    """네이버 부동산 URL에서 단지 정보 추출 (v7.3)"""
    
    # URL 패턴들
    PATTERNS = [
        # 신규 URL 형식: /complex/123456
        r'land\.naver\.com/complex/(\d+)',
        # 구형 URL: complexNo=123456
        r'complexNo=(\d+)',
        # 매물 상세: articleId와 함께
        r'complexNo=(\d+).*articleId=\d+',
        # 단지 정보 API
        r'/api/.*complex[=/](\d+)',
        # 모바일 URL
        r'm\.land\.naver\.com.*complex[=/](\d+)',
    ]
    
    @classmethod
    def extract_complex_id(cls, url):
        """URL에서 단지 ID 추출"""
        for pattern in cls.PATTERNS:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        return None
    
    @classmethod
    def extract_from_text(cls, text):
        """텍스트에서 모든 단지 URL/ID 추출"""
        results = []
        # URL에서 추출
        urls = re.findall(r'https?://[^\s<>"\']+', text)
        for url in urls:
            cid = cls.extract_complex_id(url)
            if cid and cid not in [r[1] for r in results]:
                results.append(("URL에서 추출", cid))
        
        # 단독 숫자 ID (5자리 이상)
        ids = re.findall(r'\b(\d{5,10})\b', text)
        for cid in ids:
            if cid not in [r[1] for r in results]:
                results.append(("ID 직접 입력", cid))
        
        return results
    
    @classmethod
    def fetch_complex_name(cls, complex_id):
        """단지 ID로 단지명 조회 (네이버 API)"""
        try:
            import urllib.request
            url = f"https://new.land.naver.com/api/complexes/{complex_id}?sameAddressGroup=false"
            req = urllib.request.Request(url)
            req.add_header('User-Agent', 'Mozilla/5.0')
            with urllib.request.urlopen(req, timeout=5) as response:
                data = json.loads(response.read().decode('utf-8'))
                return data.get('complexDetail', {}).get('complexName', f'단지_{complex_id}')
        except (URLError, HTTPError, SocketTimeout, JSONDecodeError) as e:
            get_logger('NaverURLParser').debug(f"단지명 조회 실패 ({complex_id}): {e}")
            return f'단지_{complex_id}'

class ExcelTemplate:
    """엑셀 내보내기 템플릿 (v7.3)"""
    
    DEFAULT_COLUMNS = [
        ("단지명", True),
        ("거래유형", True),
        ("매매가", True),
        ("보증금", True),
        ("월세", True),
        ("면적(㎡)", True),
        ("면적(평)", True),
        ("층/방향", True),
        ("타입/특징", True),
        ("매물ID", False),
        ("단지ID", False),
        ("수집시각", True),
        ("신규여부", False),
        ("가격변동", False),
    ]
    
    @classmethod
    def get_default_template(cls):
        return {name: enabled for name, enabled in cls.DEFAULT_COLUMNS}
    
    @classmethod
    def get_column_order(cls):
        return [name for name, _ in cls.DEFAULT_COLUMNS]

# ============ SETTINGS ============
DEFAULT_SETTINGS = {
    "theme": "dark", "crawl_speed": "보통", "minimize_to_tray": True,
    "show_notifications": True, "confirm_before_close": True,
    "play_sound_on_complete": True, "default_sort_column": "가격",
    "default_sort_order": "asc", "max_search_history": 20,
    "window_geometry": None, "splitter_sizes": None,
    # v7.3 신규 설정
    "excel_template": None,  # 엑셀 컬럼 템플릿
    "show_new_badge": True,  # 신규 매물 배지 표시
    "show_price_change": True,  # 가격 변동 표시
    "price_change_threshold": 0,  # 가격 변동 알림 기준 (만원, 0=모두)
}

class SettingsManager:
    _instance = None
    _lock = Lock()
    def __new__(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
                    cls._instance._initialized = False
        return cls._instance
    def __init__(self):
        if self._initialized: return
        self._initialized = True
        self._settings = DEFAULT_SETTINGS.copy()
        self._load()
    def _load(self):
        if SETTINGS_PATH.exists():
            try:
                with open(SETTINGS_PATH, 'r', encoding='utf-8') as f:
                    self._settings.update(json.load(f))
            except (OSError, JSONDecodeError) as e:
                get_logger('SettingsManager').warning(f"설정 로드 실패: {e}")
    def _save(self):
        try:
            with open(SETTINGS_PATH, 'w', encoding='utf-8') as f:
                json.dump(self._settings, f, ensure_ascii=False, indent=2)
        except OSError as e:
            get_logger('SettingsManager').warning(f"설정 저장 실패: {e}")
    def get(self, key, default=None): return self._settings.get(key, default)
    def set(self, key, value): self._settings[key] = value; self._save()
    def update(self, data): self._settings.update(data); self._save()

class FilterPresetManager:
    def __init__(self):
        self._presets = {}
        self._load()
    def _load(self):
        if PRESETS_PATH.exists():
            try:
                with open(PRESETS_PATH, 'r', encoding='utf-8') as f:
                    self._presets = json.load(f)
            except (OSError, JSONDecodeError) as e:
                get_logger('FilterPresetManager').warning(f"프리셋 로드 실패: {e}")
    def _save(self):
        try:
            with open(PRESETS_PATH, 'w', encoding='utf-8') as f:
                json.dump(self._presets, f, ensure_ascii=False, indent=2)
        except OSError as e:
            get_logger('FilterPresetManager').warning(f"프리셋 저장 실패: {e}")
    def add(self, name, config): self._presets[name] = config; self._save(); return True
    def get(self, name): return self._presets.get(name)
    def delete(self, name):
        if name in self._presets: del self._presets[name]; self._save(); return True
        return False
    def get_all_names(self): return list(self._presets.keys())

class SearchHistoryManager:
    """최근 검색 기록 관리"""
    def __init__(self, max_items=20):
        self.max_items = max_items
        self._history = []
        self._load()
    
    def _load(self):
        if HISTORY_PATH.exists():
            try:
                with open(HISTORY_PATH, 'r', encoding='utf-8') as f:
                    self._history = json.load(f)
            except (OSError, JSONDecodeError) as e:
                get_logger('SearchHistoryManager').warning(f"검색 기록 로드 실패: {e}")
    
    def _save(self):
        try:
            with open(HISTORY_PATH, 'w', encoding='utf-8') as f:
                json.dump(self._history[:self.max_items], f, ensure_ascii=False, indent=2)
        except OSError as e:
            get_logger('SearchHistoryManager').warning(f"검색 기록 저장 실패: {e}")
    
    def add(self, search_info):
        """검색 기록 추가"""
        search_info['timestamp'] = DateTimeHelper.now_string()
        # 중복 제거
        self._history = [h for h in self._history if h.get('complexes') != search_info.get('complexes')]
        self._history.insert(0, search_info)
        self._history = self._history[:self.max_items]
        self._save()
    
    def get_recent(self, count=10):
        return self._history[:count]
    
    def clear(self):
        self._history = []
        self._save()

settings = SettingsManager()

# ============ DATABASE ============
class ConnectionPool:
    def __init__(self, db_path, pool_size=5):
        self.db_path = Path(db_path)
        self.pool_size = pool_size
        self._pool = Queue(maxsize=pool_size)
        self._lock = Lock()
        print(f"[DB] ConnectionPool 초기화: {self.db_path}")
        self._initialize_pool()
    
    def _initialize_pool(self):
        for i in range(self.pool_size):
            try:
                conn = self._create_connection()
                self._pool.put(conn)
            except Exception as e:
                print(f"[DB ERROR] 연결 생성 실패 ({i+1}/{self.pool_size}): {e}")
    
    def _create_connection(self):
        # 부모 디렉토리 확인/생성
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        
        conn = sqlite3.connect(str(self.db_path), check_same_thread=False, timeout=30)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=30000")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.row_factory = sqlite3.Row
        return conn
    
    def get_connection(self):
        try:
            return self._pool.get(timeout=10)
        except Exception as e:
            print(f"[DB WARN] 풀에서 연결 가져오기 실패, 새 연결 생성: {e}")
            return self._create_connection()
    
    def return_connection(self, conn):
        if conn is None:
            return
        try:
            self._pool.put_nowait(conn)
        except QueueFull:
            try:
                conn.close()
            except Exception as e:
                get_logger('ConnectionPool').debug(f"연결 종료 중 오류: {e}")
    
    def close_all(self):
        """모든 연결 안전하게 종료"""
        print("[DB] ConnectionPool 종료 시작...")
        closed_count = 0
        error_count = 0
        
        # 최대 시도 횟수 제한
        max_attempts = self.pool_size + 5
        attempts = 0
        
        while attempts < max_attempts:
            attempts += 1
            try:
                conn = self._pool.get_nowait()
                try:
                    conn.close()
                    closed_count += 1
                except Exception as e:
                    print(f"[DB WARN] 연결 종료 실패: {e}")
                    error_count += 1
            except QueueEmpty:
                # 큐가 비었음
                break
        
        print(f"[DB] ConnectionPool 종료 완료: {closed_count}개 종료, {error_count}개 오류")

class ComplexDatabase:
    def __init__(self, db_path=None):
        self.db_path = Path(db_path) if db_path else DB_PATH
        print(f"[DB] ComplexDatabase 초기화: {self.db_path}")
        self._pool = ConnectionPool(self.db_path)
        self._init_tables()
    
    def _init_tables(self):
        conn = self._pool.get_connection()
        try:
            c = conn.cursor()
            c.execute('''CREATE TABLE IF NOT EXISTS complexes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                complex_id TEXT NOT NULL UNIQUE,
                memo TEXT DEFAULT "",
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')
            c.execute('''CREATE TABLE IF NOT EXISTS groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                description TEXT DEFAULT "",
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')
            c.execute('''CREATE TABLE IF NOT EXISTS group_complexes (
                group_id INTEGER,
                complex_id INTEGER,
                PRIMARY KEY (group_id, complex_id),
                FOREIGN KEY (group_id) REFERENCES groups(id),
                FOREIGN KEY (complex_id) REFERENCES complexes(id)
            )''')
            c.execute('''CREATE TABLE IF NOT EXISTS crawl_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                complex_name TEXT,
                complex_id TEXT,
                trade_types TEXT,
                item_count INTEGER,
                crawled_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')
            c.execute('''CREATE TABLE IF NOT EXISTS price_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                complex_id TEXT,
                trade_type TEXT,
                pyeong REAL,
                min_price INTEGER,
                max_price INTEGER,
                avg_price INTEGER,
                item_count INTEGER,
                snapshot_date DATE DEFAULT CURRENT_DATE
            )''')
            c.execute('''CREATE TABLE IF NOT EXISTS alert_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                complex_id TEXT,
                complex_name TEXT,
                trade_type TEXT,
                area_min REAL DEFAULT 0,
                area_max REAL DEFAULT 999,
                price_min INTEGER DEFAULT 0,
                price_max INTEGER DEFAULT 999999999,
                enabled INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')
            # v7.3 신규: 매물 히스토리 (신규 매물/가격 변동 추적)
            c.execute('''CREATE TABLE IF NOT EXISTS article_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                article_id TEXT NOT NULL,
                complex_id TEXT NOT NULL,
                complex_name TEXT,
                trade_type TEXT,
                price INTEGER,
                price_text TEXT,
                area_pyeong REAL,
                floor_info TEXT,
                feature TEXT,
                first_seen DATE DEFAULT CURRENT_DATE,
                last_seen DATE DEFAULT CURRENT_DATE,
                last_price INTEGER,
                price_change INTEGER DEFAULT 0,
                UNIQUE(article_id, complex_id)
            )''')
            # 인덱스 추가
            c.execute('CREATE INDEX IF NOT EXISTS idx_article_complex ON article_history(complex_id)')
            c.execute('CREATE INDEX IF NOT EXISTS idx_article_id ON article_history(article_id)')
            conn.commit()
            print("[DB] 테이블 초기화 완료")
        except Exception as e:
            print(f"[DB ERROR] 테이블 초기화 실패: {e}")
            import traceback
            traceback.print_exc()
        finally:
            self._pool.return_connection(conn)
    
    def add_complex(self, name, complex_id, memo=""):
        """단지 추가 - 디버깅 강화"""
        conn = self._pool.get_connection()
        try:
            c = conn.cursor()
            # 이미 존재하는지 확인
            c.execute("SELECT id FROM complexes WHERE complex_id = ?", (complex_id,))
            existing = c.fetchone()
            if existing:
                print(f"[DB] 단지 이미 존재: {name} ({complex_id})")
                return True  # 이미 존재하면 성공으로 처리
            
            c.execute("INSERT INTO complexes (name, complex_id, memo) VALUES (?, ?, ?)", 
                     (name, complex_id, memo))
            conn.commit()
            print(f"[DB] 단지 추가 성공: {name} ({complex_id})")
            return True
        except sqlite3.IntegrityError as e:
            print(f"[DB] 단지 중복 (정상): {name} ({complex_id})")
            return True
        except Exception as e:
            print(f"[DB ERROR] 단지 추가 실패: {name} ({complex_id}) - {e}")
            import traceback
            traceback.print_exc()
            return False
        finally:
            self._pool.return_connection(conn)
    
    def get_all_complexes(self):
        """모든 단지 조회 - 디버깅 강화"""
        conn = self._pool.get_connection()
        try:
            result = conn.cursor().execute(
                "SELECT id, name, complex_id, memo FROM complexes ORDER BY name"
            ).fetchall()
            print(f"[DB] 단지 조회: {len(result)}개")
            return result
        except Exception as e:
            print(f"[DB ERROR] 단지 조회 실패: {e}")
            import traceback
            traceback.print_exc()
            return []
        finally:
            self._pool.return_connection(conn)
    
    def delete_complex(self, db_id):
        conn = self._pool.get_connection()
        try:
            conn.cursor().execute("DELETE FROM complexes WHERE id = ?", (db_id,))
            conn.commit()
            print(f"[DB] 단지 삭제: ID={db_id}")
            return True
        except Exception as e:
            print(f"[DB ERROR] 단지 삭제 실패: {e}")
            return False
        finally:
            self._pool.return_connection(conn)
    
    def delete_complexes_bulk(self, db_ids):
        conn = self._pool.get_connection()
        try:
            c = conn.cursor()
            placeholders = ','.join('?' * len(db_ids))
            c.execute(f"DELETE FROM complexes WHERE id IN ({placeholders})", db_ids)
            conn.commit()
            print(f"[DB] 단지 일괄 삭제: {c.rowcount}개")
            return c.rowcount
        except Exception as e:
            print(f"[DB ERROR] 단지 일괄 삭제 실패: {e}")
            return 0
        finally:
            self._pool.return_connection(conn)
    
    def update_complex_memo(self, db_id, memo):
        conn = self._pool.get_connection()
        try:
            conn.cursor().execute("UPDATE complexes SET memo = ? WHERE id = ?", (memo, db_id))
            conn.commit()
            return True
        except Exception as e:
            print(f"[DB ERROR] 메모 업데이트 실패: {e}")
            return False
        finally:
            self._pool.return_connection(conn)
    
    def create_group(self, name, desc=""):
        conn = self._pool.get_connection()
        try:
            conn.cursor().execute("INSERT INTO groups (name, description) VALUES (?, ?)", (name, desc))
            conn.commit()
            print(f"[DB] 그룹 생성: {name}")
            return True
        except Exception as e:
            print(f"[DB ERROR] 그룹 생성 실패: {e}")
            return False
        finally:
            self._pool.return_connection(conn)
    
    def get_all_groups(self):
        conn = self._pool.get_connection()
        try:
            result = conn.cursor().execute("SELECT id, name, description FROM groups ORDER BY name").fetchall()
            print(f"[DB] 그룹 조회: {len(result)}개")
            return result
        except Exception as e:
            print(f"[DB ERROR] 그룹 조회 실패: {e}")
            return []
        finally:
            self._pool.return_connection(conn)
    
    def delete_group(self, group_id):
        conn = self._pool.get_connection()
        try:
            c = conn.cursor()
            c.execute("DELETE FROM group_complexes WHERE group_id = ?", (group_id,))
            c.execute("DELETE FROM groups WHERE id = ?", (group_id,))
            conn.commit()
            print(f"[DB] 그룹 삭제: ID={group_id}")
            return True
        except Exception as e:
            print(f"[DB ERROR] 그룹 삭제 실패: {e}")
            return False
        finally:
            self._pool.return_connection(conn)
    
    def add_complexes_to_group(self, group_id, complex_db_ids):
        conn = self._pool.get_connection()
        try:
            c = conn.cursor()
            count = 0
            for cid in complex_db_ids:
                try:
                    c.execute("INSERT OR IGNORE INTO group_complexes (group_id, complex_id) VALUES (?, ?)", (group_id, cid))
                    count += c.rowcount
                except Exception as e:
                    print(f"[DB WARN] 그룹에 단지 추가 실패: {cid} - {e}")
            conn.commit()
            print(f"[DB] 그룹에 단지 추가: {count}개")
            return count
        except Exception as e:
            print(f"[DB ERROR] 그룹에 단지 추가 실패: {e}")
            return 0
        finally:
            self._pool.return_connection(conn)
    
    def remove_complex_from_group(self, group_id, complex_db_id):
        conn = self._pool.get_connection()
        try:
            conn.cursor().execute("DELETE FROM group_complexes WHERE group_id = ? AND complex_id = ?", (group_id, complex_db_id))
            conn.commit()
            return True
        except Exception as e:
            print(f"[DB ERROR] 그룹에서 단지 제거 실패: {e}")
            return False
        finally:
            self._pool.return_connection(conn)
    
    def get_complexes_in_group(self, group_id):
        conn = self._pool.get_connection()
        try:
            result = conn.cursor().execute(
                'SELECT c.id, c.name, c.complex_id, c.memo FROM complexes c '
                'JOIN group_complexes gc ON c.id = gc.complex_id '
                'WHERE gc.group_id = ? ORDER BY c.name', (group_id,)
            ).fetchall()
            return result
        except Exception as e:
            print(f"[DB ERROR] 그룹 내 단지 조회 실패: {e}")
            return []
        finally:
            self._pool.return_connection(conn)
    
    def add_crawl_history(self, name, cid, types, count):
        conn = self._pool.get_connection()
        try:
            conn.cursor().execute(
                "INSERT INTO crawl_history (complex_name, complex_id, trade_types, item_count) VALUES (?, ?, ?, ?)",
                (name, cid, types, count)
            )
            conn.commit()
        except Exception as e:
            print(f"[DB ERROR] 크롤링 기록 저장 실패: {e}")
        finally:
            self._pool.return_connection(conn)
    
    def get_crawl_history(self, limit=100):
        conn = self._pool.get_connection()
        try:
            result = conn.cursor().execute(
                'SELECT complex_name, complex_id, trade_types, item_count, crawled_at '
                'FROM crawl_history ORDER BY crawled_at DESC LIMIT ?', (limit,)
            ).fetchall()
            return result
        except Exception as e:
            print(f"[DB ERROR] 크롤링 기록 조회 실패: {e}")
            return []
        finally:
            self._pool.return_connection(conn)
    
    def get_complex_price_history(self, complex_id, trade_type=None, pyeong=None):
        conn = self._pool.get_connection()
        try:
            sql = 'SELECT snapshot_date, trade_type, pyeong, min_price, max_price, avg_price FROM price_snapshots WHERE complex_id = ?'
            params = [complex_id]
            
            if trade_type and trade_type != "전체":
                sql += ' AND trade_type = ?'
                params.append(trade_type)
            
            if pyeong and pyeong != "전체":
                try:
                    p_val = float(pyeong.replace("평", ""))
                    sql += ' AND pyeong = ?'
                    params.append(p_val)
                except (ValueError, TypeError):
                    pass
                
            sql += ' ORDER BY snapshot_date DESC, pyeong'
            
            result = conn.cursor().execute(sql, params).fetchall()
            print(f"[DB] 가격 히스토리 조회: {len(result)}개 (조건: {trade_type}, {pyeong})")
            return result
        except Exception as e:
            print(f"[DB ERROR] 가격 히스토리 조회 실패: {e}")
            return []
        finally:
            self._pool.return_connection(conn)
    
    def add_price_snapshot(self, complex_id, trade_type, pyeong, min_price, max_price, avg_price, item_count):
        """가격 스냅샷 저장"""
        conn = self._pool.get_connection()
        try:
            conn.cursor().execute(
                'INSERT INTO price_snapshots (complex_id, trade_type, pyeong, min_price, max_price, avg_price, item_count) '
                'VALUES (?, ?, ?, ?, ?, ?, ?)',
                (complex_id, trade_type, pyeong, min_price, max_price, avg_price, item_count)
            )
            conn.commit()
            return True
        except Exception as e:
            print(f"[DB ERROR] 가격 스냅샷 저장 실패: {e}")
            return False
        finally:
            self._pool.return_connection(conn)
    
    def add_alert_setting(self, cid, name, ttype, amin, amax, pmin, pmax):
        conn = self._pool.get_connection()
        try:
            conn.cursor().execute(
                'INSERT INTO alert_settings (complex_id, complex_name, trade_type, area_min, area_max, price_min, price_max) '
                'VALUES (?, ?, ?, ?, ?, ?, ?)', (cid, name, ttype, amin, amax, pmin, pmax)
            )
            conn.commit()
            print(f"[DB] 알림 설정 추가: {name}")
            return True
        except Exception as e:
            print(f"[DB ERROR] 알림 설정 추가 실패: {e}")
            return False
        finally:
            self._pool.return_connection(conn)
    
    def get_all_alert_settings(self):
        conn = self._pool.get_connection()
        try:
            return conn.cursor().execute(
                'SELECT id, complex_id, complex_name, trade_type, area_min, area_max, price_min, price_max, enabled '
                'FROM alert_settings ORDER BY created_at DESC'
            ).fetchall()
        except Exception as e:
            print(f"[DB ERROR] 알림 설정 조회 실패: {e}")
            return []
        finally:
            self._pool.return_connection(conn)
    
    def toggle_alert_setting(self, aid, enabled):
        conn = self._pool.get_connection()
        try:
            conn.cursor().execute("UPDATE alert_settings SET enabled = ? WHERE id = ?", (1 if enabled else 0, aid))
            conn.commit()
        except Exception as e:
            print(f"[DB ERROR] 알림 설정 토글 실패: {e}")
        finally:
            self._pool.return_connection(conn)
    
    def delete_alert_setting(self, aid):
        conn = self._pool.get_connection()
        try:
            conn.cursor().execute("DELETE FROM alert_settings WHERE id = ?", (aid,))
            conn.commit()
        except Exception as e:
            print(f"[DB ERROR] 알림 설정 삭제 실패: {e}")
        finally:
            self._pool.return_connection(conn)
    
    def check_alerts(self, cid, ttype, area, price):
        conn = self._pool.get_connection()
        try:
            return conn.cursor().execute(
                'SELECT id, complex_name FROM alert_settings '
                'WHERE complex_id = ? AND trade_type = ? AND enabled = 1 '
                'AND area_min <= ? AND area_max >= ? AND price_min <= ? AND price_max >= ?',
                (cid, ttype, area, area, price, price)
            ).fetchall()
        except Exception as e:
            print(f"[DB ERROR] 알림 체크 실패: {e}")
            return []
        finally:
            self._pool.return_connection(conn)
    
    # ========== v7.3 신규: 매물 히스토리 메서드 ==========
    
    def check_article_history(self, article_id, complex_id, current_price):
        """
        매물 히스토리 확인 및 업데이트
        Returns: (is_new, price_change, previous_price)
        - is_new: 신규 매물 여부
        - price_change: 가격 변동 (양수: 상승, 음수: 하락, 0: 동일)
        - previous_price: 이전 가격
        """
        conn = self._pool.get_connection()
        try:
            c = conn.cursor()
            c.execute(
                'SELECT price, first_seen FROM article_history WHERE article_id = ? AND complex_id = ?',
                (article_id, complex_id)
            )
            row = c.fetchone()
            
            today = datetime.now().strftime('%Y-%m-%d')
            
            if row is None:
                # 신규 매물
                return True, 0, 0
            else:
                previous_price = row[0] if row[0] else 0
                first_seen = row[1]
                price_change = current_price - previous_price if previous_price else 0
                
                # 같은 날 첫 발견이면 신규로 표시
                is_new = (first_seen == today)
                
                return is_new, price_change, previous_price
                
        except Exception as e:
            print(f"[DB ERROR] 매물 히스토리 확인 실패: {e}")
            return False, 0, 0
        finally:
            self._pool.return_connection(conn)
    
    def update_article_history(self, article_id, complex_id, complex_name, trade_type, 
                               price, price_text, area_pyeong, floor_info, feature):
        """매물 히스토리 업데이트 (UPSERT)"""
        conn = self._pool.get_connection()
        try:
            c = conn.cursor()
            today = datetime.now().strftime('%Y-%m-%d')
            
            # 기존 데이터 확인
            c.execute(
                'SELECT id, price FROM article_history WHERE article_id = ? AND complex_id = ?',
                (article_id, complex_id)
            )
            existing = c.fetchone()
            
            if existing:
                # 업데이트 - 가격 변동 계산
                old_price = existing[1] if existing[1] else 0
                price_change = price - old_price if old_price else 0
                
                c.execute('''
                    UPDATE article_history SET
                        last_seen = ?,
                        last_price = price,
                        price = ?,
                        price_text = ?,
                        price_change = ?,
                        floor_info = ?,
                        feature = ?
                    WHERE article_id = ? AND complex_id = ?
                ''', (today, price, price_text, price_change, floor_info, feature, article_id, complex_id))
            else:
                # 신규 삽입
                c.execute('''
                    INSERT INTO article_history 
                    (article_id, complex_id, complex_name, trade_type, price, price_text, 
                     area_pyeong, floor_info, feature, first_seen, last_seen)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (article_id, complex_id, complex_name, trade_type, price, price_text,
                      area_pyeong, floor_info, feature, today, today))
            
            conn.commit()
            return True
        except Exception as e:
            print(f"[DB ERROR] 매물 히스토리 업데이트 실패: {e}")
            return False
        finally:
            self._pool.return_connection(conn)
    
    def get_article_history_stats(self, complex_id=None):
        """매물 히스토리 통계"""
        conn = self._pool.get_connection()
        try:
            today = datetime.now().strftime('%Y-%m-%d')
            
            if complex_id:
                # 특정 단지 통계
                result = conn.cursor().execute('''
                    SELECT 
                        COUNT(*) as total,
                        SUM(CASE WHEN first_seen = ? THEN 1 ELSE 0 END) as new_today,
                        SUM(CASE WHEN price_change > 0 THEN 1 ELSE 0 END) as price_up,
                        SUM(CASE WHEN price_change < 0 THEN 1 ELSE 0 END) as price_down
                    FROM article_history WHERE complex_id = ?
                ''', (today, complex_id)).fetchone()
            else:
                # 전체 통계
                result = conn.cursor().execute('''
                    SELECT 
                        COUNT(*) as total,
                        SUM(CASE WHEN first_seen = ? THEN 1 ELSE 0 END) as new_today,
                        SUM(CASE WHEN price_change > 0 THEN 1 ELSE 0 END) as price_up,
                        SUM(CASE WHEN price_change < 0 THEN 1 ELSE 0 END) as price_down
                    FROM article_history
                ''', (today,)).fetchone()
            
            return {
                'total': result[0] or 0,
                'new_today': result[1] or 0,
                'price_up': result[2] or 0,
                'price_down': result[3] or 0
            }
        except Exception as e:
            print(f"[DB ERROR] 매물 통계 조회 실패: {e}")
            return {'total': 0, 'new_today': 0, 'price_up': 0, 'price_down': 0}
        finally:
            self._pool.return_connection(conn)
    
    def cleanup_old_articles(self, days=30):
        """오래된 매물 히스토리 정리"""
        conn = self._pool.get_connection()
        try:
            c = conn.cursor()
            c.execute('''
                DELETE FROM article_history 
                WHERE julianday('now') - julianday(last_seen) > ?
            ''', (days,))
            deleted = c.rowcount
            conn.commit()
            print(f"[DB] 오래된 매물 {deleted}개 정리 (>{days}일)")
            return deleted
        except Exception as e:
            print(f"[DB ERROR] 매물 정리 실패: {e}")
            return 0
        finally:
            self._pool.return_connection(conn)
    
    def backup_database(self, path):
        try:
            shutil.copy2(self.db_path, path)
            print(f"[DB] 백업 완료: {path}")
            return True
        except Exception as e:
            print(f"[DB ERROR] 백업 실패: {e}")
            return False
    
    def restore_database(self, path):
        """DB 복원 - 안전한 복원 로직"""
        print(f"[DB] 복원 시작: {path}")
        
        # 1. 원본 파일 존재 확인
        if not Path(path).exists():
            print(f"[DB ERROR] 복원 파일이 존재하지 않음: {path}")
            return False
        
        try:
            # 2. 기존 연결 풀 안전하게 종료
            print("[DB] 기존 연결 풀 종료 중...")
            if self._pool:
                try:
                    self._pool.close_all()
                except Exception as e:
                    print(f"[DB WARN] 연결 풀 종료 중 오류 (무시): {e}")
            
            # 3. 잠시 대기 (파일 핸들 해제를 위해)
            import time
            time.sleep(0.5)
            
            # 4. 기존 DB 파일 백업 (안전을 위해)
            backup_path = self.db_path.with_suffix('.db.backup')
            if self.db_path.exists():
                try:
                    shutil.copy2(self.db_path, backup_path)
                    print(f"[DB] 기존 DB 백업: {backup_path}")
                except Exception as e:
                    print(f"[DB WARN] 기존 DB 백업 실패 (무시): {e}")
            
            # 5. 새 파일 복사
            print(f"[DB] 파일 복사: {path} -> {self.db_path}")
            shutil.copy2(path, self.db_path)
            
            # 6. 새 연결 풀 생성
            print("[DB] 새 연결 풀 생성 중...")
            self._pool = ConnectionPool(self.db_path)
            
            # 7. 연결 테스트
            conn = self._pool.get_connection()
            test_result = conn.cursor().execute("SELECT COUNT(*) FROM complexes").fetchone()
            self._pool.return_connection(conn)
            print(f"[DB] 복원 완료! 단지 수: {test_result[0]}개")
            
            # 8. 백업 파일 삭제
            if backup_path.exists():
                try:
                    backup_path.unlink()
                except OSError as e:
                    get_logger('ComplexDatabase').debug(f"백업 파일 삭제 실패 (무시): {e}")
            
            return True
            
        except Exception as e:
            print(f"[DB ERROR] 복원 실패: {e}")
            import traceback
            traceback.print_exc()
            
            # 복원 실패 시 기존 백업에서 복구 시도
            backup_path = self.db_path.with_suffix('.db.backup')
            if backup_path.exists():
                try:
                    print("[DB] 백업에서 복구 시도...")
                    shutil.copy2(backup_path, self.db_path)
                    self._pool = ConnectionPool(self.db_path)
                    print("[DB] 백업에서 복구 완료")
                except Exception as e2:
                    print(f"[DB ERROR] 백업 복구도 실패: {e2}")
            
            return False

# ============ CRAWLER ============
class CrawlerThread(QThread):
    log_signal = pyqtSignal(str, int)
    progress_signal = pyqtSignal(int, str, int)  # percent, current_name, remaining_seconds
    item_signal = pyqtSignal(dict)
    stats_signal = pyqtSignal(dict)
    complex_finished_signal = pyqtSignal(str, str, str, int)
    finished_signal = pyqtSignal(list)
    error_signal = pyqtSignal(str)
    alert_triggered_signal = pyqtSignal(str, str, str, float, int)
    
    def __init__(self, targets, trade_types, area_filter, price_filter, db, speed="보통"):
        super().__init__()
        self.targets = targets
        self.trade_types = trade_types
        self.area_filter = area_filter
        self.price_filter = price_filter
        self.db = db
        self.speed = speed
        self._running = True
        self.collected_data = []
        self.stats = {"total_found": 0, "filtered_out": 0, "by_trade_type": {"매매": 0, "전세": 0, "월세": 0}}
        self.start_time = None
        self.items_per_second = 0
    
    def stop(self): self._running = False
    def log(self, msg, level=20): self.log_signal.emit(msg, level)
    
    def run(self):
        if not UC_AVAILABLE or not BS4_AVAILABLE:
            self.error_signal.emit("필수 라이브러리 미설치\npip install undetected-chromedriver beautifulsoup4")
            return
        driver = None
        self.start_time = time.time()
        try:
            self.log("🚀 크롤링 시작...")
            self.log("🔧 Chrome 드라이버 초기화 중...")
            
            options = uc.ChromeOptions()
            options.add_argument("--headless=new")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-gpu")
            options.add_argument("--disable-software-rasterizer")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
            options.add_argument("--log-level=3")
            
            try:
                driver = uc.Chrome(options=options, version_main=None)
                self.log("✅ Chrome 드라이버 초기화 성공")
            except Exception as e:
                self.log(f"⚠️ Headless 실패, 일반 모드 시도... ({e})", 30)
                options2 = uc.ChromeOptions()
                options2.add_argument("--no-sandbox")
                options2.add_argument("--disable-dev-shm-usage")
                options2.add_argument("--disable-gpu")
                options2.add_argument("--window-size=1920,1080")
                options2.add_argument("--start-minimized")
                driver = uc.Chrome(options=options2, version_main=None)
                self.log("✅ Chrome 드라이버 초기화 성공 (일반 모드)")
            
            driver.set_page_load_timeout(30)
            driver.implicitly_wait(5)
            
            total = len(self.targets) * len(self.trade_types)
            current = 0
            
            for name, cid in self.targets:
                if not self._running: break
                complex_count = 0
                for ttype in self.trade_types:
                    if not self._running: break
                    current += 1
                    
                    # 예상 남은 시간 계산
                    elapsed = time.time() - self.start_time
                    avg_time = elapsed / current if current > 0 else 5
                    remaining = int(avg_time * (total - current))
                    
                    self.progress_signal.emit(int(current / total * 100), f"{name} ({ttype})", remaining)
                    self.log(f"\n📍 [{current}/{total}] {name} - {ttype}")
                    
                    try:
                        count = self._crawl(driver, name, cid, ttype)
                        complex_count += count
                        self.stats["by_trade_type"][ttype] = self.stats["by_trade_type"].get(ttype, 0) + count
                        self.log(f"   ✅ {count}건 수집")
                    except Exception as e:
                        self.log(f"   ❌ 오류: {e}", 40)
                        import traceback
                        self.log(f"   상세: {traceback.format_exc()}", 40)
                    
                    speed_cfg = CRAWL_SPEED_PRESETS.get(self.speed, CRAWL_SPEED_PRESETS["보통"])
                    time.sleep(random.uniform(speed_cfg["min"], speed_cfg["max"]))
                
                self.complex_finished_signal.emit(name, cid, ",".join(self.trade_types), complex_count)
            
            self.log(f"\n{'='*50}\n✅ 완료! 총 {len(self.collected_data)}건")
        except Exception as e:
            self.log(f"❌ 치명적 오류: {e}", 40)
            import traceback
            self.log(f"상세:\n{traceback.format_exc()}", 40)
            self.error_signal.emit(str(e))
        finally:
            if driver:
                try:
                    driver.quit()
                    self.log("✅ Chrome 드라이버 종료 완료")
                except Exception as e:
                    self.log(f"⚠️ Chrome 드라이버 종료 중 오류: {e}", 30)
            self.finished_signal.emit(self.collected_data)
    
    def _crawl(self, driver, name, cid, ttype):
        trade_param = {"매매": "A1", "전세": "B1", "월세": "B2"}.get(ttype, "A1")
        url = f"https://new.land.naver.com/complexes/{cid}?ms=37.5,127,16&a=APT&e=RETAIL&tradeTypes={trade_param}"
        
        self.log(f"   🔗 URL 접속 중...")
        driver.get(url)
        time.sleep(4)
        
        try:
            article_tab = driver.find_element("css selector", "a[href*='articleList'], .tab_item[data-tab='article']")
            article_tab.click()
            time.sleep(2)
        except Exception as e:
            # 탭 클릭 실패는 정상적인 상황일 수 있음 (탭이 없는 경우)
            self.log(f"   ℹ️ 매물 탭 찾기 실패 (정상): {type(e).__name__}", 10)
        
        self._scroll(driver)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        return self._parse(soup, name, cid, ttype)
    
    def _scroll(self, driver):
        try:
            container = None
            for sel in [".article_list", ".item_list", ".complex_list", "[class*='article']"]:
                try:
                    container = driver.find_element("css selector", sel)
                    break
                except Exception:
                    continue
            
            if not container:
                last_h = 0
                for _ in range(10):
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
                    time.sleep(1)
                    new_h = driver.execute_script("return document.body.scrollHeight")
                    if new_h == last_h: break
                    last_h = new_h
                return
            
            last_h, attempts = 0, 0
            while attempts < 30 and self._running:
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", container)
                time.sleep(1.2)
                new_h = driver.execute_script("return arguments[0].scrollHeight", container)
                if new_h == last_h:
                    attempts += 1
                    if attempts >= 3: break
                else: attempts = 0
                last_h = new_h
        except Exception as e:
            self.log(f"   ⚠️ 스크롤 오류: {e}", 30)
    
    def _parse(self, soup, name, cid, ttype):
        items = []
        article_items = []
        
        for sel in [".item_article", ".item_inner", ".article_item", "[class*='ArticleItem']", ".complex_item", "li[data-article-id]", ".list_item"]:
            found = soup.select(sel)
            if found:
                article_items = found
                self.log(f"   📋 선택자 '{sel}': {len(found)}개 발견")
                break
        
        if not article_items:
            self.log("   ⚠️ 표준 선택자 실패, 대체 방식 시도...")
            article_items = soup.find_all(['div', 'li'], class_=lambda x: x and ('item' in x.lower() or 'article' in x.lower()))
        
        self.log(f"   🔍 파싱 대상: {len(article_items)}개")
        
        matched_count, skipped_type = 0, 0
        
        for item in article_items:
            if not self._running: break
            try:
                data = self._parse_item(item, name, cid, ttype)
                if data and data.get("면적(㎡)", 0) > 0:
                    detected_type = data.get("거래유형", "")
                    if detected_type == ttype:
                        if self._check_filters(data, ttype):
                            self.collected_data.append(data)
                            self.item_signal.emit(data)
                            items.append(data)
                            self.stats["total_found"] += 1
                            matched_count += 1
                        else:
                            self.stats["filtered_out"] += 1
                        self.stats_signal.emit(self.stats)
                    else:
                        skipped_type += 1
            except Exception as e:
                self.log(f"   ⚠️ 항목 파싱 중 오류: {e}", 30)
        
        if skipped_type > 0:
            self.log(f"   ℹ️ 다른 거래유형 {skipped_type}건 제외 (요청: {ttype})")
        
        return matched_count
    
    def _parse_item(self, item, name, cid, ttype):
        full_text = item.get_text(separator=" ", strip=True)
        detected_type = ttype
        
        for sel in [".type", ".trade_type", "[class*='type']", ".item_type", ".article_type"]:
            elem = item.select_one(sel)
            if elem:
                type_text = elem.get_text(strip=True)
                if "매매" in type_text: detected_type = "매매"
                elif "전세" in type_text: detected_type = "전세"
                elif "월세" in type_text: detected_type = "월세"
                break
        
        price_text = ""
        for sel in [".item_price strong", ".price_line", ".article_price", "[class*='price']", ".selling_price", ".trade_price", "strong[class*='Price']", ".price"]:
            elem = item.select_one(sel)
            if elem:
                price_text = elem.get_text(strip=True)
                if price_text and ("억" in price_text or "만" in price_text or price_text.replace(",", "").replace("/", "").isdigit()):
                    break
        
        if not price_text:
            price_match = re.search(r'(\d+억\s*\d*,?\d*만?|\d+,?\d*만)', full_text)
            if price_match: price_text = price_match.group(1)
        
        if re.search(r'\d+[억만]?\s*/\s*\d+', price_text): detected_type = "월세"
        elif "전세" in full_text[:50]: detected_type = "전세"
        elif "매매" in full_text[:50]: detected_type = "매매"
        
        area_text, sqm, pyeong = "", 0, 0
        for sel in [".item_area", ".info_area", ".article_area", "[class*='area']"]:
            elem = item.select_one(sel)
            if elem: area_text = elem.get_text(strip=True); break
        if not area_text: area_text = full_text
        
        sqm_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:㎡|m²)', area_text)
        if sqm_match:
            sqm = float(sqm_match.group(1))
            pyeong = AreaConverter.sqm_to_pyeong(sqm)
        else:
            pyeong_match = re.search(r'(\d+(?:\.\d+)?)\s*평', area_text)
            if pyeong_match:
                pyeong = float(pyeong_match.group(1))
                sqm = round(pyeong / 0.3025, 2)
        
        supply_match = re.search(r'(\d+(?:\.\d+)?)[㎡m²]?\s*/\s*(\d+(?:\.\d+)?)', area_text)
        if supply_match:
            sqm = float(supply_match.group(2))
            pyeong = AreaConverter.sqm_to_pyeong(sqm)
        
        # 층/방향 추출 - 다양한 선택자와 정규식 시도
        floor_text = ""
        floor_selectors = [
            ".item_floor", ".info_floor", ".floor", "[class*='floor']",
            ".article_floor", ".item_info .floor", "span.floor",
            ".info_article_floor", ".cell_floor", ".data_floor",
            "td.floor", ".item_cell.floor", "[class*='Floor']"
        ]
        for sel in floor_selectors:
            elem = item.select_one(sel)
            if elem:
                floor_text = elem.get_text(strip=True)
                if floor_text:
                    break
        
        # 정규식 fallback - 다양한 층 패턴
        if not floor_text:
            # "고층", "중층", "저층" 패턴
            level_match = re.search(r'(고층|중층|저층)', full_text)
            # "N층" 패턴
            floor_match = re.search(r'(\d+)\s*층', full_text)
            # "N/N층" 패턴 (현재층/총층)
            floor_total_match = re.search(r'(\d+)\s*/\s*(\d+)\s*층', full_text)
            
            if floor_total_match:
                floor_text = f"{floor_total_match.group(1)}/{floor_total_match.group(2)}층"
            elif floor_match:
                floor_text = f"{floor_match.group(1)}층"
            elif level_match:
                floor_text = level_match.group(1)
        
        # 방향 추출
        direction = ""
        direction_selectors = [
            ".item_direction", ".direction", "[class*='direction']",
            ".info_direction", ".cell_direction", "[class*='Direction']"
        ]
        for sel in direction_selectors:
            elem = item.select_one(sel)
            if elem:
                direction = elem.get_text(strip=True)
                if direction:
                    break
        
        # 정규식 fallback - 방향 패턴
        if not direction:
            dir_match = re.search(r'(동향|서향|남향|북향|남동향|남서향|북동향|북서향|동남향|동북향|서남향|서북향)', full_text)
            if dir_match:
                direction = dir_match.group(1)
        
        # 층/방향 결합
        if floor_text and direction:
            floor_text = f"{floor_text} {direction}"
        elif direction and not floor_text:
            floor_text = direction
        
        # 특징/설명 추출 - 개선된 로직 (v7.3)
        feature_text = ""
        
        # 필터링할 광고/무의미 키워드 (이 키워드만 있는 경우 무시)
        ad_keywords = [
            "부동산뱅크", "직방", "다방", "피터팬", "네이버부동산", "KB부동산",
            "부동산114", "호갱노노", "매물번호", "중개사무소", "공인중개사",
            "제공", "출처", "문의", "연락", "전화", "상담", "클릭", "바로가기",
            "더보기", "자세히", "확인하세요", "드립니다", "해드립니다"
        ]
        
        # 의미있는 특징 키워드
        meaningful_keywords = [
            # 매물 상태
            "급매", "급전", "급처분", "네고가능", "협의가능", "가격조정", "실매물",
            # 인테리어/상태
            "올수리", "풀수리", "리모델링", "인테리어", "풀옵션", "빌트인", "새것", "깨끗",
            "신축", "준신축", "수리완료", "도배완료", "장판교체", "싱크대교체",
            # 입주 관련
            "즉시입주", "입주가능", "공실", "실입주", "바로입주", "협의입주",
            # 위치/환경
            "역세권", "초역세권", "더블역세권", "학군", "학교앞", "공원앞", "공원뷰",
            "한강뷰", "산뷰", "오션뷰", "시티뷰", "조망좋음", "조망권", "남향",
            # 구조/시설
            "베란다확장", "확장형", "복층", "테라스", "정원", "마당", "옥상",
            "주차가능", "주차2대", "분리형", "투룸", "쓰리룸", "방3개", "방2개",
            "화장실2", "욕실2개", "드레스룸", "팬트리", "다용도실",
            # 층수
            "탑층", "로얄층", "고층", "중층", "저층", "1층", "꼭대기",
            # 거래 조건
            "전세안고", "전세끼고", "주인직거래", "세입자있음", "세놓은",
            # 기타 특징
            "펜트하우스", "복도식", "계단식", "엘리베이터", "경비실", "관리비저렴"
        ]
        
        feature_selectors = [
            ".item_desc", ".feature", ".info_sub", "[class*='desc']",
            ".article_desc", ".item_feature", ".description",
            ".info_article_feature", ".cell_feature", ".data_feature",
            ".item_info_desc", ".tag_list", ".item_tag", "[class*='tag']",
            ".item_detail", ".detail_info", ".sub_info"
        ]
        
        # 1차: 선택자로 추출 시도
        for sel in feature_selectors:
            elem = item.select_one(sel)
            if elem:
                text = elem.get_text(separator=" ", strip=True)
                if text and len(text) > 2:
                    # 광고 키워드만 있는지 체크
                    is_ad_only = any(ad in text for ad in ad_keywords) and \
                                 not any(kw in text for kw in meaningful_keywords)
                    if not is_ad_only:
                        # 광고 키워드 제거
                        cleaned = text
                        for ad in ad_keywords:
                            cleaned = cleaned.replace(ad, "").strip()
                        if cleaned and len(cleaned) > 2:
                            feature_text = cleaned[:100]
                            break
        
        # 2차: 텍스트에서 의미있는 키워드 추출
        if not feature_text or len(feature_text) < 3:
            found_features = []
            for kw in meaningful_keywords:
                if kw in full_text:
                    found_features.append(kw)
                    if len(found_features) >= 6:  # 최대 6개
                        break
            if found_features:
                feature_text = ", ".join(found_features)
        
        # 3차: 방/화장실 개수 정보 추출
        if not feature_text:
            room_info = []
            room_match = re.search(r'(\d)\s*룸|방\s*(\d)|(\d)\s*베드', full_text)
            bath_match = re.search(r'(\d)\s*욕|화장실\s*(\d)|(\d)\s*배스', full_text)
            if room_match:
                num = room_match.group(1) or room_match.group(2) or room_match.group(3)
                room_info.append(f"방{num}개")
            if bath_match:
                num = bath_match.group(1) or bath_match.group(2) or bath_match.group(3)
                room_info.append(f"화장실{num}개")
            if room_info:
                feature_text = ", ".join(room_info)
        
        article_id = ""
        link = item.select_one("a[href*='articleId']")
        if link:
            href = link.get('href', '')
            id_match = re.search(r'articleId=(\d+)', href)
            if id_match: article_id = id_match.group(1)
        else:
            article_id = item.get('data-article-id', '') or item.get('data-id', '')
        
        매매가, 보증금, 월세 = "", "", ""
        if detected_type == "매매":
            매매가 = price_text.replace("매매", "").strip()
        elif detected_type == "전세":
            보증금 = price_text.replace("전세", "").strip()
        else:
            price_clean = price_text.replace("월세", "").strip()
            if "/" in price_clean:
                parts = price_clean.split("/")
                보증금 = parts[0].strip()
                월세 = parts[1].strip() if len(parts) > 1 else ""
            else:
                보증금 = price_clean
        
        return {
            "단지명": name, "단지ID": cid, "거래유형": detected_type,
            "매매가": 매매가, "보증금": 보증금, "월세": 월세,
            "면적(㎡)": sqm, "면적(평)": pyeong, "층/방향": floor_text,
            "타입/특징": feature_text, "매물ID": article_id,
            "수집시각": DateTimeHelper.now_string()
        }
    
    def _check_filters(self, data, ttype):
        if self.area_filter.get("enabled"):
            sqm = data.get("면적(㎡)", 0)
            if sqm < self.area_filter.get("min", 0) or sqm > self.area_filter.get("max", 999):
                return False
        if self.price_filter.get("enabled"):
            price_range = self.price_filter.get(ttype, {})
            min_p, max_p = price_range.get("min", 0), price_range.get("max", 999999)
            if ttype == "매매": price = PriceConverter.to_int(data.get("매매가", "0"))
            else: price = PriceConverter.to_int(data.get("보증금", "0"))
            if price < min_p or price > max_p: return False
        return True

# ============ EXPORTER ============
class DataExporter:
    # v7.3: 확장된 컬럼 (신규, 가격변동 포함)
    COLUMNS = [
        "단지명", "거래유형", "매매가", "보증금", "월세", 
        "면적(㎡)", "면적(평)", "층/방향", "타입/특징", 
        "매물ID", "단지ID", "수집시각", "신규여부", "가격변동"
    ]
    
    def __init__(self, data): 
        self.data = data
    
    def to_excel(self, path, template=None):
        """엑셀로 내보내기 - 템플릿 지원 (v7.3)"""
        if not OPENPYXL_AVAILABLE: return None
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "매물 데이터"
            
            # 템플릿에서 컬럼 결정
            if template and 'order' in template and 'columns' in template:
                columns = [c for c in template['order'] if template['columns'].get(c, False)]
            else:
                columns = ["단지명", "거래유형", "매매가", "보증금", "월세", 
                          "면적(㎡)", "면적(평)", "층/방향", "타입/특징", "수집시각"]
            
            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # 헤더 작성
            for col, h in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # 거래유형별 색상
            trade_colors = {
                "매매": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
                "전세": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
                "월세": PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            }
            
            # v7.3: 신규/가격변동 색상
            new_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            price_up_font = Font(color="FF0000")
            price_down_font = Font(color="008000")
            
            # 데이터 작성
            for ri, item in enumerate(self.data, 2):
                tt = item.get("거래유형", "")
                fill = trade_colors.get(tt)
                
                for ci, cn in enumerate(columns, 1):
                    # 특수 컬럼 처리
                    if cn == "신규여부":
                        value = "🆕 신규" if item.get('is_new', False) else ""
                    elif cn == "가격변동":
                        pc = item.get('price_change', 0)
                        if pc > 0:
                            value = f"+{PriceConverter.to_string(pc)}"
                        elif pc < 0:
                            value = PriceConverter.to_string(pc)
                        else:
                            value = ""
                    else:
                        value = item.get(cn, "")
                    
                    cell = ws.cell(row=ri, column=ci, value=value)
                    
                    # 스타일 적용
                    if fill:
                        cell.fill = fill
                    
                    # 신규 매물 강조
                    if item.get('is_new', False) and cn == "단지명":
                        cell.fill = new_fill
                    
                    # 가격 변동 색상
                    if cn == "가격변동":
                        pc = item.get('price_change', 0)
                        if pc > 0:
                            cell.font = price_up_font
                        elif pc < 0:
                            cell.font = price_down_font
            
            # 컬럼 너비 설정
            for col in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            ws.freeze_panes = "A2"
            wb.save(path)
            return path
        except Exception as e:
            print(f"[EXPORT ERROR] Excel 저장 실패: {e}")
            return None
    
    def to_csv(self, path, template=None):
        """CSV로 내보내기 - 템플릿 지원"""
        try:
            # 템플릿에서 컬럼 결정
            if template and 'order' in template and 'columns' in template:
                columns = [c for c in template['order'] if template['columns'].get(c, False)]
            else:
                columns = ["단지명", "거래유형", "매매가", "보증금", "월세", 
                          "면적(㎡)", "면적(평)", "층/방향", "타입/특징", "수집시각"]
            
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')
                writer.writeheader()
                
                for item in self.data:
                    # 특수 컬럼 추가
                    row = dict(item)
                    row['신규여부'] = "신규" if item.get('is_new', False) else ""
                    pc = item.get('price_change', 0)
                    row['가격변동'] = PriceConverter.to_string(pc) if pc else ""
                    writer.writerow(row)
            return path
        except Exception as e:
            print(f"[EXPORT ERROR] CSV 저장 실패: {e}")
            return None
    
    def to_json(self, path):
        """JSON으로 내보내기"""
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump({
                    "exported_at": DateTimeHelper.now_string(), 
                    "total_count": len(self.data),
                    "new_count": sum(1 for d in self.data if d.get('is_new', False)),
                    "price_change_count": sum(1 for d in self.data if d.get('price_change', 0) != 0),
                    "data": self.data
                }, f, ensure_ascii=False, indent=2)
            return path
        except Exception as e:
            print(f"[EXPORT ERROR] JSON 저장 실패: {e}")
            return None

# ============ STYLES ============
def get_dark_stylesheet():
    return """
/* === v10.5 Enhanced Dark Theme === */

/* Main Window & Base */
QMainWindow, QWidget { 
    background-color: #1a1a2e; 
    color: #eaeaea; 
    font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;
    font-size: 13px;
}

/* GroupBox - Modern Card Style */
QGroupBox { 
    border: 1px solid rgba(255, 255, 255, 0.1); 
    border-radius: 12px; 
    margin-top: 1.2em; 
    padding: 15px;
    padding-top: 25px;
    font-weight: bold;
    font-size: 14px;
    background-color: rgba(40, 40, 60, 0.95);
}
QGroupBox::title { 
    subcontrol-origin: margin; 
    left: 15px; 
    padding: 0 10px;
    color: #64b5f6;
}

/* Input Fields - Glassmorphism */
QLineEdit, QSpinBox, QTimeEdit, QDoubleSpinBox, QComboBox { 
    background-color: rgba(255, 255, 255, 0.08); 
    border: 1px solid rgba(255, 255, 255, 0.15); 
    border-radius: 8px; 
    padding: 8px 12px; 
    color: #fff; 
    min-height: 32px;
    selection-background-color: #4a9eff;
}
QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, QComboBox:focus { 
    border: 2px solid #4a9eff;
    background-color: rgba(255, 255, 255, 0.12);
}
QLineEdit:hover, QSpinBox:hover, QDoubleSpinBox:hover, QComboBox:hover { 
    border-color: rgba(255, 255, 255, 0.3);
}
QComboBox::drop-down {
    border: none;
    padding-right: 8px;
}
QComboBox QAbstractItemView {
    background-color: #2a2a40;
    border: 1px solid rgba(255, 255, 255, 0.15);
    border-radius: 8px;
    selection-background-color: #4a6ea8;
}

/* Table Widget - Enhanced */
QTableWidget { 
    background-color: rgba(30, 30, 50, 0.95); 
    gridline-color: rgba(100, 100, 140, 0.4); 
    border: 1px solid rgba(255, 255, 255, 0.1); 
    border-radius: 10px; 
    color: #eee; 
    alternate-background-color: rgba(60, 60, 100, 0.35);
}
QTableWidget::item {
    border-bottom: 1px solid rgba(100, 100, 140, 0.3);
    padding: 4px 8px;
}
QTableWidget::item:selected { 
    background-color: rgba(74, 110, 168, 0.8); 
}
QTableWidget::item:hover { 
    background-color: rgba(255, 255, 255, 0.12); 
}
QHeaderView::section { 
    background-color: rgba(50, 50, 80, 0.95); 
    padding: 10px; 
    border: none;
    border-bottom: 1px solid rgba(100, 100, 140, 0.5);
    color: #fff; 
    font-weight: bold;
    font-size: 12px;
}
QHeaderView::section:first {
    border-top-left-radius: 8px;
}
QHeaderView::section:last {
    border-top-right-radius: 8px;
}

/* Text Browser / Log */
QTextBrowser { 
    background-color: rgba(15, 15, 25, 0.98); 
    color: #4ade80; 
    border: 1px solid rgba(255, 255, 255, 0.1); 
    border-radius: 10px; 
    font-family: 'Consolas', 'Courier New', monospace; 
    font-size: 12px;
    padding: 10px;
    line-height: 1.4;
}

/* Buttons - Enhanced with Gradients */
QPushButton { 
    background-color: rgba(60, 60, 90, 0.95);
    border: 1px solid rgba(255, 255, 255, 0.15); 
    border-radius: 8px; 
    padding: 10px 20px; 
    color: white; 
    font-weight: bold; 
    min-height: 36px;
}
QPushButton:hover { 
    background-color: rgba(80, 80, 110, 0.95);
    border-color: rgba(255, 255, 255, 0.3);
}
QPushButton:pressed { 
    background-color: rgba(50, 50, 75, 0.95);
}
QPushButton:disabled { 
    background-color: rgba(40, 40, 60, 0.5); 
    color: #666; 
    border-color: rgba(255, 255, 255, 0.05);
}

/* Start Button - Green Gradient */
QPushButton#startButton { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #34d399, stop:1 #22c55e);
    border: none;
    color: #fff;
    font-size: 15px;
}
QPushButton#startButton:hover { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #4ade80, stop:1 #22c55e);
}
QPushButton#startButton:disabled {
    background: rgba(34, 197, 94, 0.3);
    color: rgba(255, 255, 255, 0.5);
}

/* Stop Button - Red Gradient */
QPushButton#stopButton { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f87171, stop:1 #ef4444);
    border: none;
    color: #fff;
}
QPushButton#stopButton:hover { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #fca5a5, stop:1 #ef4444);
}

/* Save Button - Blue Gradient */
QPushButton#saveButton { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #60a5fa, stop:1 #3b82f6);
    border: none;
    color: #fff;
}
QPushButton#saveButton:hover { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #93c5fd, stop:1 #3b82f6);
}

/* Link Button - Purple */
QPushButton#linkButton { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #a78bfa, stop:1 #8b5cf6);
    border: none;
    min-width: 65px; 
    padding: 6px 12px; 
    min-height: 28px;
    font-size: 12px;
}
QPushButton#linkButton:hover { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #c4b5fd, stop:1 #8b5cf6);
}

/* Progress Bar - Animated Gradient */
QProgressBar { 
    border: none;
    border-radius: 10px; 
    text-align: center; 
    background-color: rgba(40, 40, 60, 0.8); 
    color: #fff; 
    min-height: 28px;
    font-weight: bold;
}
QProgressBar::chunk { 
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3b82f6, stop:0.5 #8b5cf6, stop:1 #22c55e);
    border-radius: 8px;
}

/* Checkbox & Radio */
QCheckBox, QRadioButton { 
    color: #eee; 
    spacing: 10px;
    font-size: 13px;
}
QCheckBox::indicator { 
    width: 20px; 
    height: 20px; 
    border-radius: 5px; 
    border: 2px solid rgba(255, 255, 255, 0.3); 
    background-color: rgba(255, 255, 255, 0.08);
}
QCheckBox::indicator:hover {
    border-color: #4a9eff;
}
QCheckBox::indicator:checked { 
    background-color: #3b82f6; 
    border-color: #60a5fa;
    image: url(data:image/svg+xml;base64,PHN2ZyB2aWV3Qm94PSIwIDAgMjQgMjQiIGZpbGw9IndoaXRlIiB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciPjxwYXRoIGQ9Ik05IDE2LjE3TDQuODMgMTJsLTEuNDIgMS40MUw5IDE5IDIxIDdsLTEuNDEtMS40MUw5IDE2LjE3eiIvPjwvc3ZnPg==);
}

/* Tab Widget - Modern */
QTabWidget::pane { 
    border: 1px solid rgba(255, 255, 255, 0.1); 
    border-radius: 10px; 
    background-color: rgba(40, 40, 60, 0.95);
    margin-top: -1px;
}
QTabBar::tab { 
    background-color: rgba(50, 50, 75, 0.8);
    color: #aaa; 
    padding: 12px 24px; 
    margin-right: 3px; 
    border-top-left-radius: 10px; 
    border-top-right-radius: 10px;
    font-size: 13px;
}
QTabBar::tab:selected { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #4a6ea8, stop:1 #3b5998);
    color: #fff;
    font-weight: bold;
}
QTabBar::tab:hover:!selected { 
    background-color: rgba(70, 70, 100, 0.9);
    color: #ddd;
}

/* Scroll Bars - Minimal */
QScrollBar:vertical { 
    background-color: transparent; 
    width: 10px; 
    border-radius: 5px;
    margin: 2px;
}
QScrollBar::handle:vertical { 
    background-color: rgba(255, 255, 255, 0.2); 
    border-radius: 5px; 
    min-height: 40px;
}
QScrollBar::handle:vertical:hover { 
    background-color: rgba(255, 255, 255, 0.35);
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { 
    height: 0px;
}
QScrollBar:horizontal {
    background-color: transparent;
    height: 10px;
    border-radius: 5px;
}
QScrollBar::handle:horizontal {
    background-color: rgba(255, 255, 255, 0.2);
    border-radius: 5px;
    min-width: 40px;
}

/* Menu - Modern Dropdown */
QMenu { 
    background-color: rgba(40, 40, 60, 0.98);
    border: 1px solid rgba(255, 255, 255, 0.15); 
    border-radius: 10px;
    color: #fff; 
    padding: 8px;
}
QMenu::item { 
    padding: 10px 30px;
    border-radius: 6px;
}
QMenu::item:selected { 
    background-color: rgba(74, 110, 168, 0.8);
}
QMenu::separator {
    height: 1px;
    background-color: rgba(255, 255, 255, 0.1);
    margin: 5px 10px;
}

/* Slider - Modern */
QSlider::groove:horizontal { 
    height: 6px; 
    background: rgba(255, 255, 255, 0.1);
    border-radius: 3px;
}
QSlider::handle:horizontal { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #60a5fa, stop:1 #3b82f6);
    width: 18px; 
    height: 18px;
    margin: -6px 0; 
    border-radius: 9px;
}
QSlider::handle:horizontal:hover { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #93c5fd, stop:1 #60a5fa);
}
QSlider::sub-page:horizontal {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3b82f6, stop:1 #60a5fa);
    border-radius: 3px;
}

/* Tooltip - Modern */
QToolTip { 
    background-color: rgba(50, 50, 75, 0.98);
    color: #fff; 
    border: 1px solid rgba(255, 255, 255, 0.2);
    border-radius: 8px; 
    padding: 8px 12px;
    font-size: 12px;
}

/* List Widget */
QListWidget {
    background-color: rgba(30, 30, 50, 0.95);
    border: 1px solid rgba(255, 255, 255, 0.1);
    border-radius: 10px;
    alternate-background-color: rgba(255, 255, 255, 0.03);
}
QListWidget::item {
    padding: 8px;
    border-radius: 6px;
}
QListWidget::item:selected {
    background-color: rgba(74, 110, 168, 0.8);
}
QListWidget::item:hover:!selected {
    background-color: rgba(255, 255, 255, 0.08);
}

/* Status Bar */
QStatusBar {
    background-color: rgba(30, 30, 50, 0.95);
    color: #aaa;
    border-top: 1px solid rgba(255, 255, 255, 0.1);
}

/* Menu Bar */
QMenuBar {
    background-color: rgba(25, 25, 40, 0.98);
    color: #eee;
    padding: 5px;
}
QMenuBar::item {
    padding: 8px 15px;
    border-radius: 6px;
}
QMenuBar::item:selected {
    background-color: rgba(74, 110, 168, 0.8);
}
"""


def get_light_stylesheet():
    return """
/* === v10.5 Enhanced Light Theme === */

/* Main Window & Base */
QMainWindow, QWidget { 
    background-color: #f8fafc; 
    color: #1e293b; 
    font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;
    font-size: 13px;
}

/* GroupBox - Modern Card Style */
QGroupBox { 
    border: 1px solid rgba(0, 0, 0, 0.08);
    border-radius: 12px; 
    margin-top: 1.2em; 
    padding: 15px;
    padding-top: 25px;
    font-weight: bold;
    font-size: 14px;
    background-color: #ffffff;
    box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
}
QGroupBox::title { 
    subcontrol-origin: margin; 
    left: 15px; 
    padding: 0 10px;
    color: #3b82f6;
}

/* Input Fields */
QLineEdit, QSpinBox, QTimeEdit, QDoubleSpinBox, QComboBox { 
    background-color: #ffffff;
    border: 1px solid #e2e8f0; 
    border-radius: 8px; 
    padding: 8px 12px; 
    color: #1e293b; 
    min-height: 32px;
    selection-background-color: #3b82f6;
    selection-color: #fff;
}
QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, QComboBox:focus { 
    border: 2px solid #3b82f6;
    background-color: #fff;
}
QLineEdit:hover, QSpinBox:hover, QDoubleSpinBox:hover, QComboBox:hover { 
    border-color: #94a3b8;
}
QComboBox::drop-down {
    border: none;
    padding-right: 8px;
}
QComboBox QAbstractItemView {
    background-color: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    selection-background-color: #3b82f6;
    selection-color: #fff;
}

/* Table Widget - Clean */
QTableWidget { 
    background-color: #ffffff;
    gridline-color: #d1d5db;
    border: 1px solid #e2e8f0; 
    border-radius: 10px; 
    color: #1e293b;
    alternate-background-color: #f1f5f9;
}
QTableWidget::item {
    border-bottom: 1px solid #e5e7eb;
    padding: 4px 8px;
}
QTableWidget::item:selected { 
    background-color: #3b82f6;
    color: #fff;
}
QTableWidget::item:hover { 
    background-color: #e2e8f0;
}
QHeaderView::section { 
    background-color: #f8fafc;
    padding: 10px; 
    border: none;
    border-bottom: 1px solid #d1d5db;
    color: #475569; 
    font-weight: bold;
    font-size: 12px;
}
QHeaderView::section:first {
    border-top-left-radius: 8px;
}
QHeaderView::section:last {
    border-top-right-radius: 8px;
}

/* Text Browser / Log */
QTextBrowser { 
    background-color: #ffffff;
    color: #15803d;
    border: 1px solid #e2e8f0;
    border-radius: 10px; 
    font-family: 'Consolas', 'Courier New', monospace; 
    font-size: 12px;
    padding: 10px;
}

/* Buttons */
QPushButton { 
    background-color: #f1f5f9;
    border: 1px solid #e2e8f0; 
    border-radius: 8px; 
    padding: 10px 20px; 
    color: #475569; 
    font-weight: bold; 
    min-height: 36px;
}
QPushButton:hover { 
    background-color: #e2e8f0;
    border-color: #cbd5e1;
}
QPushButton:pressed { 
    background-color: #cbd5e1;
}
QPushButton:disabled { 
    background-color: #f8fafc; 
    color: #94a3b8; 
    border-color: #f1f5f9;
}

/* Start Button - Green Gradient */
QPushButton#startButton { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #22c55e, stop:1 #16a34a);
    border: none;
    color: #fff;
    font-size: 15px;
}
QPushButton#startButton:hover { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #4ade80, stop:1 #22c55e);
}
QPushButton#startButton:disabled {
    background: #86efac;
    color: rgba(255, 255, 255, 0.7);
}

/* Stop Button - Red Gradient */
QPushButton#stopButton { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ef4444, stop:1 #dc2626);
    border: none;
    color: #fff;
}
QPushButton#stopButton:hover { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f87171, stop:1 #ef4444);
}

/* Save Button - Blue Gradient */
QPushButton#saveButton { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3b82f6, stop:1 #2563eb);
    border: none;
    color: #fff;
}
QPushButton#saveButton:hover { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #60a5fa, stop:1 #3b82f6);
}

/* Link Button - Purple */
QPushButton#linkButton { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #8b5cf6, stop:1 #7c3aed);
    border: none;
    color: #fff;
    min-width: 65px; 
    padding: 6px 12px; 
    min-height: 28px;
    font-size: 12px;
}
QPushButton#linkButton:hover { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #a78bfa, stop:1 #8b5cf6);
}

/* Progress Bar */
QProgressBar { 
    border: none;
    border-radius: 10px; 
    text-align: center; 
    background-color: #e2e8f0;
    color: #475569;
    min-height: 28px;
    font-weight: bold;
}
QProgressBar::chunk { 
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3b82f6, stop:0.5 #8b5cf6, stop:1 #22c55e);
    border-radius: 8px;
}

/* Checkbox & Radio */
QCheckBox, QRadioButton { 
    color: #1e293b;
    spacing: 10px;
    font-size: 13px;
}
QCheckBox::indicator { 
    width: 20px; 
    height: 20px; 
    border-radius: 5px; 
    border: 2px solid #cbd5e1;
    background-color: #fff;
}
QCheckBox::indicator:hover {
    border-color: #3b82f6;
}
QCheckBox::indicator:checked { 
    background-color: #3b82f6; 
    border-color: #3b82f6;
}

/* Tab Widget */
QTabWidget::pane { 
    border: 1px solid #e2e8f0; 
    border-radius: 10px; 
    background-color: #ffffff;
    margin-top: -1px;
}
QTabBar::tab { 
    background-color: #f1f5f9;
    color: #64748b; 
    padding: 12px 24px; 
    margin-right: 3px; 
    border-top-left-radius: 10px; 
    border-top-right-radius: 10px;
    font-size: 13px;
}
QTabBar::tab:selected { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3b82f6, stop:1 #2563eb);
    color: #fff;
    font-weight: bold;
}
QTabBar::tab:hover:!selected { 
    background-color: #e2e8f0;
    color: #475569;
}

/* Scroll Bars - Minimal */
QScrollBar:vertical { 
    background-color: transparent; 
    width: 10px; 
    border-radius: 5px;
    margin: 2px;
}
QScrollBar::handle:vertical { 
    background-color: #cbd5e1;
    border-radius: 5px; 
    min-height: 40px;
}
QScrollBar::handle:vertical:hover { 
    background-color: #94a3b8;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { 
    height: 0px;
}
QScrollBar:horizontal {
    background-color: transparent;
    height: 10px;
    border-radius: 5px;
}
QScrollBar::handle:horizontal {
    background-color: #cbd5e1;
    border-radius: 5px;
    min-width: 40px;
}

/* Menu */
QMenu { 
    background-color: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    color: #1e293b; 
    padding: 8px;
}
QMenu::item { 
    padding: 10px 30px;
    border-radius: 6px;
}
QMenu::item:selected { 
    background-color: #3b82f6;
    color: #fff;
}
QMenu::separator {
    height: 1px;
    background-color: #e2e8f0;
    margin: 5px 10px;
}

/* Slider */
QSlider::groove:horizontal { 
    height: 6px; 
    background: #e2e8f0;
    border-radius: 3px;
}
QSlider::handle:horizontal { 
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3b82f6, stop:1 #2563eb);
    width: 18px; 
    height: 18px;
    margin: -6px 0; 
    border-radius: 9px;
}
QSlider::sub-page:horizontal {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3b82f6, stop:1 #60a5fa);
    border-radius: 3px;
}

/* Tooltip */
QToolTip { 
    background-color: #1e293b;
    color: #fff; 
    border: none;
    border-radius: 8px; 
    padding: 8px 12px;
    font-size: 12px;
}

/* List Widget */
QListWidget {
    background-color: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    alternate-background-color: #f8fafc;
}
QListWidget::item {
    padding: 8px;
    border-radius: 6px;
}
QListWidget::item:selected {
    background-color: #3b82f6;
    color: #fff;
}
QListWidget::item:hover:!selected {
    background-color: #f1f5f9;
}

/* Status Bar */
QStatusBar {
    background-color: #f8fafc;
    color: #64748b;
    border-top: 1px solid #e2e8f0;
}

/* Menu Bar */
QMenuBar {
    background-color: #ffffff;
    color: #1e293b;
    padding: 5px;
    border-bottom: 1px solid #e2e8f0;
}
QMenuBar::item {
    padding: 8px 15px;
    border-radius: 6px;
}
QMenuBar::item:selected {
    background-color: #3b82f6;
    color: #fff;
}
"""

def get_stylesheet(theme="dark"):
    return get_light_stylesheet() if theme == "light" else get_dark_stylesheet()

# ============ WIDGETS ============

class ToastWidget(QWidget):
    """v11.0: 비침습적 Toast 알림 위젯"""
    def __init__(self, message: str, toast_type: str = "info", parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Tool | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating)
        
        # 타입별 색상 설정
        colors = {
            "success": {"bg": "rgba(34, 197, 94, 0.95)", "icon": "✅"},
            "error": {"bg": "rgba(239, 68, 68, 0.95)", "icon": "❌"},
            "warning": {"bg": "rgba(245, 158, 11, 0.95)", "icon": "⚠️"},
            "info": {"bg": "rgba(59, 130, 246, 0.95)", "icon": "ℹ️"},
        }
        
        color_info = colors.get(toast_type, colors["info"])
        
        # 레이아웃 설정
        layout = QHBoxLayout(self)
        layout.setContentsMargins(20, 15, 20, 15)
        layout.setSpacing(12)
        
        # 아이콘
        icon_label = QLabel(color_info["icon"])
        icon_label.setStyleSheet("font-size: 20px;")
        layout.addWidget(icon_label)
        
        # 메시지
        msg_label = QLabel(message)
        msg_label.setStyleSheet(
            "color: white; font-size: 13px; font-weight: bold; padding: 0;"
        )
        msg_label.setWordWrap(True)
        layout.addWidget(msg_label, 1)
        
        # 스타일
        self.setStyleSheet(f"""
            ToastWidget {{
                background-color: {color_info["bg"]};
                border-radius: 12px;
                border: 1px solid rgba(255, 255, 255, 0.2);
            }}
        """)
        
        # 크기 조정
        self.setMinimumWidth(300)
        self.setMaximumWidth(500)
        self.adjustSize()
        
        # 애니메이션 타이머
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.fade_out)
        self.timer.setSingleShot(True)
        
    def show_toast(self, duration: int = 3000):
        """Toast 표시"""
        self.show()
        self.raise_()
        self.timer.start(duration)
    
    def fade_out(self):
        """페이드 아웃 애니메이션"""
        self.close()
        if self.parent():
            try:
                self.parent().toast_widgets.remove(self)
                self.parent()._reposition_toasts()
            except (AttributeError, ValueError):
                pass

class SortableTableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        t1 = self.text().replace(",", "")
        t2 = other.text().replace(",", "")
        v1, v2 = self._extract_number(t1), self._extract_number(t2)
        if v1 is not None and v2 is not None: return v1 < v2
        return t1 < t2
    
    def _extract_number(self, text):
        try:
            text = re.sub(r'\(\d+건\)', '', text).strip()
            if "평" in text: return float(text.replace("평", "").strip())
            if "/" in text: text = text.split("/")[0]
            if "억" in text or "만" in text: return float(PriceConverter.to_int(text))
            return None
        except (ValueError, TypeError, AttributeError):
            return None

class SearchBar(QWidget):
    search_changed = pyqtSignal(str)
    def __init__(self, placeholder="검색...", parent=None):
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(QLabel("🔍"))
        self.input = QLineEdit()
        self.input.setPlaceholderText(placeholder)
        self.input.setClearButtonEnabled(True)
        self.input.textChanged.connect(lambda t: self.search_changed.emit(t))
        layout.addWidget(self.input)
    def text(self): return self.input.text()
    def clear(self): self.input.clear()
    def setFocus(self): self.input.setFocus()

class SpeedSlider(QWidget):
    speed_changed = pyqtSignal(str)
    SPEEDS = ["빠름", "보통", "느림", "매우 느림"]
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        header = QHBoxLayout()
        header.addWidget(QLabel("⚡ 속도:"))
        self.label = QLabel("보통")
        self.label.setStyleSheet("font-weight: bold; color: #4a9eff;")
        header.addWidget(self.label)
        self.desc_label = QLabel("(권장 속도)")
        self.desc_label.setStyleSheet("color: #888; font-size: 11px;")
        header.addWidget(self.desc_label)
        header.addStretch()
        layout.addLayout(header)
        self.slider = QSlider(Qt.Orientation.Horizontal)
        self.slider.setRange(0, 3)
        self.slider.setValue(1)
        self.slider.setTickPosition(QSlider.TickPosition.TicksBelow)
        self.slider.valueChanged.connect(self._on_change)
        self.slider.setToolTip("크롤링 속도를 조절합니다. 느릴수록 차단 위험이 낮습니다.")
        layout.addWidget(self.slider)
    def _on_change(self, val):
        speed = self.SPEEDS[val]
        self.label.setText(speed)
        desc = CRAWL_SPEED_PRESETS.get(speed, {}).get("desc", "")
        self.desc_label.setText(f"({desc})")
        self.speed_changed.emit(speed)
    def current_speed(self): return self.SPEEDS[self.slider.value()]
    def set_speed(self, speed):
        if speed in self.SPEEDS: self.slider.setValue(self.SPEEDS.index(speed))

class SummaryCard(QFrame):
    """결과 요약 카드 위젿 (v7.3 확장)"""
    def __init__(self, parent=None, theme="dark"):
        super().__init__(parent)
        self.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        self._theme = theme
        self._apply_theme()
        layout = QHBoxLayout(self)
        layout.setSpacing(15)
        
        # 총 수집
        self.total_widget = self._create_stat_widget("📊 총 수집", "0건", "#3498db")
        layout.addWidget(self.total_widget)
        
        # 매매
        self.trade_widget = self._create_stat_widget("🏠 매매", "0건", "#e74c3c")
        layout.addWidget(self.trade_widget)
        
        # 전세
        self.jeonse_widget = self._create_stat_widget("📋 전세", "0건", "#2ecc71")
        layout.addWidget(self.jeonse_widget)
        
        # 월세
        self.monthly_widget = self._create_stat_widget("💰 월세", "0건", "#9b59b6")
        layout.addWidget(self.monthly_widget)
        
        # 구분선
        self.sep = QFrame()
        self.sep.setFrameShape(QFrame.Shape.VLine)
        self._update_separator_style()
        layout.addWidget(self.sep)
        
        # v7.3 신규: 신규 매물
        self.new_widget = self._create_stat_widget("🆕 신규", "0건", "#f39c12")
        layout.addWidget(self.new_widget)
        
        # v7.3 신규: 가격 상승
        self.price_up_widget = self._create_stat_widget("📈 상승", "0건", "#e74c3c")
        layout.addWidget(self.price_up_widget)
        
        # v7.3 신규: 가격 하락
        self.price_down_widget = self._create_stat_widget("📉 하락", "0건", "#27ae60")
        layout.addWidget(self.price_down_widget)
        
        # 필터 제외
        self.filtered_widget = self._create_stat_widget("🚫 제외", "0건", "#95a5a6")
        layout.addWidget(self.filtered_widget)
        
        layout.addStretch()
    
    def _apply_theme(self):
        """테마에 따른 스타일 적용"""
        if self._theme == "dark":
            self.setStyleSheet("""
                SummaryCard { 
                    background-color: rgba(40, 40, 60, 0.95); 
                    border-radius: 10px; 
                    border: 1px solid rgba(255, 255, 255, 0.1);
                    padding: 10px;
                }
            """)
        else:
            self.setStyleSheet("""
                SummaryCard { 
                    background-color: #ffffff; 
                    border-radius: 10px; 
                    border: 1px solid #e2e8f0;
                    padding: 10px;
                }
            """)
    
    def _update_separator_style(self):
        """구분선 스타일 업데이트"""
        if self._theme == "dark":
            self.sep.setStyleSheet("color: rgba(255, 255, 255, 0.2);")
        else:
            self.sep.setStyleSheet("color: #e2e8f0;")
    
    def set_theme(self, theme):
        """테마 변경 시 호출"""
        self._theme = theme
        self._apply_theme()
        self._update_separator_style()
        self._update_title_colors()
    
    def _update_title_colors(self):
        """타이틀 레이블 색상 업데이트"""
        title_color = "#aaa" if self._theme == "dark" else "#64748b"
        for widget in [self.total_widget, self.trade_widget, self.jeonse_widget, 
                       self.monthly_widget, self.new_widget, self.price_up_widget,
                       self.price_down_widget, self.filtered_widget]:
            labels = widget.findChildren(QLabel)
            for label in labels:
                if label.objectName() != "value":
                    label.setStyleSheet(f"color: {title_color}; font-size: 11px;")
    
    def _create_stat_widget(self, title, value, color):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(8, 5, 8, 5)
        layout.setSpacing(2)
        
        title_label = QLabel(title)
        title_color = "#aaa" if self._theme == "dark" else "#64748b"
        title_label.setStyleSheet(f"color: {title_color}; font-size: 11px;")
        layout.addWidget(title_label)
        
        value_label = QLabel(value)
        value_label.setObjectName("value")
        value_label.setStyleSheet(f"color: {color}; font-size: 16px; font-weight: bold;")
        layout.addWidget(value_label)
        
        return widget
    
    def update_stats(self, total=0, trade=0, jeonse=0, monthly=0, filtered=0, 
                     new_count=0, price_up=0, price_down=0):
        self.total_widget.findChild(QLabel, "value").setText(f"{total}건")
        self.trade_widget.findChild(QLabel, "value").setText(f"{trade}건")
        self.jeonse_widget.findChild(QLabel, "value").setText(f"{jeonse}건")
        self.monthly_widget.findChild(QLabel, "value").setText(f"{monthly}건")
        self.filtered_widget.findChild(QLabel, "value").setText(f"{filtered}건")
        self.new_widget.findChild(QLabel, "value").setText(f"{new_count}건")
        self.price_up_widget.findChild(QLabel, "value").setText(f"{price_up}건")
        self.price_down_widget.findChild(QLabel, "value").setText(f"{price_down}건")
    
    def reset(self):
        self.update_stats(0, 0, 0, 0, 0, 0, 0, 0)

class ProgressWidget(QWidget):
    """진행 상태 위젯 - 예상 시간 표시"""
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)
        
        # 상태 표시줄
        status_layout = QHBoxLayout()
        self.status_label = QLabel("대기 중...")
        self.status_label.setStyleSheet("font-weight: bold;")
        status_layout.addWidget(self.status_label)
        
        self.time_label = QLabel("")
        self.time_label.setStyleSheet("color: #888;")
        status_layout.addWidget(self.time_label)
        status_layout.addStretch()
        layout.addLayout(status_layout)
        
        # 프로그레스바
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setMinimumHeight(25)
        layout.addWidget(self.progress_bar)
    
    def update_progress(self, percent, current_name, remaining_seconds):
        self.progress_bar.setValue(percent)
        self.status_label.setText(f"🔄 {current_name}")
        
        if remaining_seconds > 0:
            mins, secs = divmod(remaining_seconds, 60)
            if mins > 0:
                self.time_label.setText(f"예상 남은 시간: {mins}분 {secs}초")
            else:
                self.time_label.setText(f"예상 남은 시간: {secs}초")
        else:
            self.time_label.setText("")
    
    def reset(self):
        self.progress_bar.setValue(0)
        self.status_label.setText("대기 중...")
        self.time_label.setText("")
    
    def complete(self):
        self.progress_bar.setValue(100)
        self.status_label.setText("✅ 완료!")
        self.time_label.setText("")

class ColoredTableWidgetItem(QTableWidgetItem):
    def __init__(self, text, trade_type=None, is_dark=True):
        super().__init__(text)
        if trade_type in TRADE_COLORS:
            colors = TRADE_COLORS[trade_type]
            bg = colors["dark_bg"] if is_dark else colors["bg"]
            fg = colors["dark_fg"] if is_dark else colors["fg"]
            self.setBackground(QColor(bg))
            self.setForeground(QColor(fg))

class LinkButton(QPushButton):
    """클릭 가능한 링크 버튼"""
    def __init__(self, url, parent=None):
        super().__init__("🔗 보기", parent)
        self.url = url
        self.setObjectName("linkButton")
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.setToolTip(f"클릭하여 열기:\n{url[:50]}...")
        # 버튼 크기 고정
        self.setFixedHeight(26)
        self.setMaximumWidth(70)
        self.setMinimumWidth(60)
        self.setStyleSheet("""
            QPushButton {
                font-size: 11px;
                padding: 2px 6px;
                min-height: 22px;
                max-height: 24px;
            }
        """)
        self.clicked.connect(self._open_url)
    
    def _open_url(self):
        if self.url:
            webbrowser.open(self.url)

class MultiSelectDialog(QDialog):
    def __init__(self, title, items, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumSize(450, 550)
        self._selected = []
        self._setup_ui(items)
    
    def _setup_ui(self, items):
        layout = QVBoxLayout(self)
        
        # 상단 버튼
        btn_layout = QHBoxLayout()
        btn_all = QPushButton("✅ 전체 선택")
        btn_all.clicked.connect(self._select_all)
        btn_none = QPushButton("⬜ 전체 해제")
        btn_none.clicked.connect(self._deselect_all)
        btn_layout.addWidget(btn_all)
        btn_layout.addWidget(btn_none)
        layout.addLayout(btn_layout)
        
        # 리스트
        self.list_widget = QListWidget()
        self.list_widget.setAlternatingRowColors(True)
        self.list_widget.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        for item_data in items:
            text, data = (item_data if isinstance(item_data, tuple) else (str(item_data), item_data))
            item = QListWidgetItem()
            checkbox = QCheckBox(text)
            checkbox.setProperty("data", data)
            self.list_widget.addItem(item)
            self.list_widget.setItemWidget(item, checkbox)
        layout.addWidget(self.list_widget)
        
        # 카운트
        self.count_label = QLabel("선택: 0개")
        self.count_label.setStyleSheet("font-weight: bold; color: #4a9eff;")
        layout.addWidget(self.count_label)
        
        # 버튼
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        for i in range(self.list_widget.count()):
            cb = self.list_widget.itemWidget(self.list_widget.item(i))
            if isinstance(cb, QCheckBox):
                cb.stateChanged.connect(self._update_count)
    
    def _update_count(self):
        count = sum(1 for i in range(self.list_widget.count())
                   if isinstance(w := self.list_widget.itemWidget(self.list_widget.item(i)), QCheckBox) and w.isChecked())
        self.count_label.setText(f"선택: {count}개")
    
    def _select_all(self):
        for i in range(self.list_widget.count()):
            if isinstance(cb := self.list_widget.itemWidget(self.list_widget.item(i)), QCheckBox):
                cb.setChecked(True)
    
    def _deselect_all(self):
        for i in range(self.list_widget.count()):
            if isinstance(cb := self.list_widget.itemWidget(self.list_widget.item(i)), QCheckBox):
                cb.setChecked(False)
    
    def _on_accept(self):
        self._selected = []
        for i in range(self.list_widget.count()):
            if isinstance(cb := self.list_widget.itemWidget(self.list_widget.item(i)), QCheckBox) and cb.isChecked():
                self._selected.append(cb.property("data"))
        self.accept()
    
    def selected_items(self): return self._selected

# ============ DIALOGS ============
class PresetDialog(QDialog):
    def __init__(self, parent=None, preset_manager=None):
        super().__init__(parent)
        self.preset_manager = preset_manager
        self.selected_preset = None
        self._setup_ui()
    
    def _setup_ui(self):
        self.setWindowTitle("📁 필터 프리셋")
        self.setMinimumSize(400, 350)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("저장된 프리셋:"))
        self.list = QListWidget()
        self.list.setAlternatingRowColors(True)
        self.list.itemDoubleClicked.connect(self._load)
        layout.addWidget(self.list)
        btn_layout = QHBoxLayout()
        btn_load = QPushButton("📂 불러오기")
        btn_load.clicked.connect(self._load)
        btn_del = QPushButton("🗑️ 삭제")
        btn_del.clicked.connect(self._delete)
        btn_layout.addWidget(btn_load)
        btn_layout.addWidget(btn_del)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        self._refresh()
    
    def _refresh(self):
        self.list.clear()
        if self.preset_manager:
            for name in self.preset_manager.get_all_names():
                self.list.addItem(name)
    
    def _load(self):
        if item := self.list.currentItem():
            self.selected_preset = item.text()
            self.accept()
    
    def _delete(self):
        if (item := self.list.currentItem()) and self.preset_manager:
            self.preset_manager.delete(item.text())
            self._refresh()

class AlertSettingDialog(QDialog):
    def __init__(self, parent=None, db=None):
        super().__init__(parent)
        self.db = db
        self._setup_ui()
    
    def _setup_ui(self):
        self.setWindowTitle("🔔 알림 설정")
        self.setMinimumSize(650, 550)
        layout = QVBoxLayout(self)
        
        # 추가 폼
        add_g = QGroupBox("새 알림 추가")
        add_l = QGridLayout()
        add_l.addWidget(QLabel("단지:"), 0, 0)
        self.combo_complex = QComboBox()
        for _, name, cid, _ in (self.db.get_all_complexes() if self.db else []):
            self.combo_complex.addItem(f"{name} ({cid})", (cid, name))
        add_l.addWidget(self.combo_complex, 0, 1, 1, 3)
        
        add_l.addWidget(QLabel("유형:"), 1, 0)
        self.combo_type = QComboBox()
        self.combo_type.addItems(["매매", "전세", "월세"])
        add_l.addWidget(self.combo_type, 1, 1)
        
        add_l.addWidget(QLabel("면적(평):"), 2, 0)
        self.spin_area_min = QDoubleSpinBox()
        self.spin_area_min.setRange(0, 200)
        add_l.addWidget(self.spin_area_min, 2, 1)
        add_l.addWidget(QLabel("~"), 2, 2)
        self.spin_area_max = QDoubleSpinBox()
        self.spin_area_max.setRange(0, 200)
        self.spin_area_max.setValue(100)
        add_l.addWidget(self.spin_area_max, 2, 3)
        
        add_l.addWidget(QLabel("가격(만원):"), 3, 0)
        self.spin_price_min = QSpinBox()
        self.spin_price_min.setRange(0, 999999)
        self.spin_price_min.setSingleStep(1000)
        add_l.addWidget(self.spin_price_min, 3, 1)
        add_l.addWidget(QLabel("~"), 3, 2)
        self.spin_price_max = QSpinBox()
        self.spin_price_max.setRange(0, 999999)
        self.spin_price_max.setValue(100000)
        self.spin_price_max.setSingleStep(1000)
        add_l.addWidget(self.spin_price_max, 3, 3)
        
        btn_add = QPushButton("➕ 추가")
        btn_add.clicked.connect(self._add)
        add_l.addWidget(btn_add, 4, 0, 1, 4)
        add_g.setLayout(add_l)
        layout.addWidget(add_g)
        
        layout.addWidget(QLabel("설정된 알림:"))
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["단지", "유형", "면적", "가격", "활성", "삭제"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)
        
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)
        self._load()
    
    def _add(self):
        if self.combo_complex.count() == 0: return
        data = self.combo_complex.currentData()
        if not data: return
        cid, name = data
        if self.db.add_alert_setting(cid, name, self.combo_type.currentText(),
            self.spin_area_min.value(), self.spin_area_max.value(),
            self.spin_price_min.value(), self.spin_price_max.value()):
            self._load()
    
    def _load(self):
        self.table.setRowCount(0)
        if not self.db: return
        for aid, cid, name, tt, amin, amax, pmin, pmax, enabled in self.db.get_all_alert_settings():
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(name or cid))
            self.table.setItem(row, 1, QTableWidgetItem(tt))
            self.table.setItem(row, 2, QTableWidgetItem(f"{amin}~{amax}평"))
            self.table.setItem(row, 3, QTableWidgetItem(f"{pmin:,}~{pmax:,}만"))
            check = QCheckBox()
            check.setChecked(enabled == 1)
            check.stateChanged.connect(lambda s, a=aid: self.db.toggle_alert_setting(a, s == Qt.CheckState.Checked.value))
            self.table.setCellWidget(row, 4, check)
            btn = QPushButton("🗑️")
            btn.clicked.connect(lambda _, a=aid: self._delete(a))
            self.table.setCellWidget(row, 5, btn)
    
    def _delete(self, aid):
        self.db.delete_alert_setting(aid)
        self._load()

# ============ v7.3 신규 다이얼로그 ============

class AdvancedFilterDialog(QDialog):
    """고급 결과 필터 다이얼로그 (v7.3)"""
    filter_applied = pyqtSignal(dict)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🔍 고급 필터")
        self.setMinimumWidth(450)
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        
        # 가격 필터
        price_group = QGroupBox("💰 가격 필터")
        pg = QGridLayout(price_group)
        pg.addWidget(QLabel("최소 가격:"), 0, 0)
        self.price_min = QSpinBox()
        self.price_min.setRange(0, 9999999)
        self.price_min.setSuffix(" 만원")
        self.price_min.setSpecialValueText("제한 없음")
        pg.addWidget(self.price_min, 0, 1)
        pg.addWidget(QLabel("최대 가격:"), 0, 2)
        self.price_max = QSpinBox()
        self.price_max.setRange(0, 9999999)
        self.price_max.setValue(9999999)
        self.price_max.setSuffix(" 만원")
        self.price_max.setSpecialValueText("제한 없음")
        pg.addWidget(self.price_max, 0, 3)
        layout.addWidget(price_group)
        
        # 면적 필터
        area_group = QGroupBox("📐 면적 필터")
        ag = QGridLayout(area_group)
        ag.addWidget(QLabel("최소 면적:"), 0, 0)
        self.area_min = QDoubleSpinBox()
        self.area_min.setRange(0, 500)
        self.area_min.setSuffix(" 평")
        self.area_min.setSpecialValueText("제한 없음")
        ag.addWidget(self.area_min, 0, 1)
        ag.addWidget(QLabel("최대 면적:"), 0, 2)
        self.area_max = QDoubleSpinBox()
        self.area_max.setRange(0, 500)
        self.area_max.setValue(500)
        self.area_max.setSuffix(" 평")
        ag.addWidget(self.area_max, 0, 3)
        layout.addWidget(area_group)
        
        # 층수 필터
        floor_group = QGroupBox("🏢 층수 필터")
        fg = QHBoxLayout(floor_group)
        self.floor_low = QCheckBox("저층")
        self.floor_mid = QCheckBox("중층")
        self.floor_high = QCheckBox("고층")
        self.floor_low.setChecked(True)
        self.floor_mid.setChecked(True)
        self.floor_high.setChecked(True)
        fg.addWidget(self.floor_low)
        fg.addWidget(self.floor_mid)
        fg.addWidget(self.floor_high)
        fg.addStretch()
        layout.addWidget(floor_group)
        
        # 특수 필터
        special_group = QGroupBox("⭐ 특수 필터")
        sg = QHBoxLayout(special_group)
        self.only_new = QCheckBox("🆕 신규 매물만")
        self.only_price_down = QCheckBox("📉 가격 하락만")
        self.only_price_change = QCheckBox("📊 가격 변동만")
        sg.addWidget(self.only_new)
        sg.addWidget(self.only_price_down)
        sg.addWidget(self.only_price_change)
        sg.addStretch()
        layout.addWidget(special_group)
        
        # 키워드 필터
        keyword_group = QGroupBox("🔤 키워드 필터")
        kg = QVBoxLayout(keyword_group)
        kg.addWidget(QLabel("포함 키워드 (쉼표로 구분):"))
        self.include_keywords = QLineEdit()
        self.include_keywords.setPlaceholderText("예: 급매, 역세권, 올수리")
        kg.addWidget(self.include_keywords)
        kg.addWidget(QLabel("제외 키워드 (쉼표로 구분):"))
        self.exclude_keywords = QLineEdit()
        self.exclude_keywords.setPlaceholderText("예: 반지하, 탑층")
        kg.addWidget(self.exclude_keywords)
        layout.addWidget(keyword_group)
        
        # 버튼
        btn_layout = QHBoxLayout()
        btn_reset = QPushButton("초기화")
        btn_reset.clicked.connect(self._reset)
        btn_apply = QPushButton("적용")
        btn_apply.clicked.connect(self._apply)
        btn_apply.setDefault(True)
        btn_layout.addWidget(btn_reset)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_apply)
        layout.addLayout(btn_layout)
    
    def _reset(self):
        self.price_min.setValue(0)
        self.price_max.setValue(9999999)
        self.area_min.setValue(0)
        self.area_max.setValue(500)
        self.floor_low.setChecked(True)
        self.floor_mid.setChecked(True)
        self.floor_high.setChecked(True)
        self.only_new.setChecked(False)
        self.only_price_down.setChecked(False)
        self.only_price_change.setChecked(False)
        self.include_keywords.clear()
        self.exclude_keywords.clear()
    
    def _apply(self):
        filters = {
            'price_min': self.price_min.value(),
            'price_max': self.price_max.value(),
            'area_min': self.area_min.value(),
            'area_max': self.area_max.value(),
            'floor_low': self.floor_low.isChecked(),
            'floor_mid': self.floor_mid.isChecked(),
            'floor_high': self.floor_high.isChecked(),
            'only_new': self.only_new.isChecked(),
            'only_price_down': self.only_price_down.isChecked(),
            'only_price_change': self.only_price_change.isChecked(),
            'include_keywords': [k.strip() for k in self.include_keywords.text().split(',') if k.strip()],
            'exclude_keywords': [k.strip() for k in self.exclude_keywords.text().split(',') if k.strip()],
        }
        self.filter_applied.emit(filters)
        self.accept()

class URLBatchDialog(QDialog):
    """URL 일괄 등록 다이얼로그 (v7.3)"""
    complexes_added = pyqtSignal(list)  # [(name, id), ...]
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🔗 URL 일괄 등록")
        self.setMinimumSize(600, 500)
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        
        # 안내
        info = QLabel(
            "네이버 부동산 URL 또는 단지 ID를 붙여넣으세요.\n"
            "여러 개를 한 번에 입력할 수 있습니다 (한 줄에 하나씩)."
        )
        info.setStyleSheet("color: #888; padding: 10px;")
        layout.addWidget(info)
        
        # 입력 영역
        self.input_text = QTextBrowser()
        self.input_text.setReadOnly(False)
        self.input_text.setPlaceholderText(
            "예시:\n"
            "https://new.land.naver.com/complexes/102378\n"
            "https://land.naver.com/complex?complexNo=123456\n"
            "123456\n"
            "789012"
        )
        self.input_text.setAcceptRichText(False)
        layout.addWidget(self.input_text, 2)
        
        # 파싱 버튼
        btn_parse = QPushButton("🔍 URL 분석")
        btn_parse.clicked.connect(self._parse_urls)
        layout.addWidget(btn_parse)
        
        # 결과 테이블
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(4)
        self.result_table.setHorizontalHeaderLabels(["✓", "단지 ID", "단지명", "상태"])
        self.result_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.result_table.setColumnWidth(0, 30)
        self.result_table.setColumnWidth(1, 100)
        self.result_table.setColumnWidth(3, 100)
        layout.addWidget(self.result_table, 3)
        
        # 진행 상태
        self.status_label = QLabel("")
        layout.addWidget(self.status_label)
        
        # 버튼
        btn_layout = QHBoxLayout()
        btn_select_all = QPushButton("전체 선택")
        btn_select_all.clicked.connect(self._select_all)
        btn_add = QPushButton("📥 선택 항목 추가")
        btn_add.clicked.connect(self._add_selected)
        btn_layout.addWidget(btn_select_all)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_add)
        layout.addLayout(btn_layout)
    
    def _parse_urls(self):
        text = self.input_text.toPlainText()
        if not text.strip():
            QMessageBox.warning(self, "입력 필요", "URL 또는 단지 ID를 입력하세요.")
            return
        
        self.result_table.setRowCount(0)
        results = NaverURLParser.extract_from_text(text)
        
        if not results:
            QMessageBox.warning(self, "파싱 실패", "유효한 URL이나 단지 ID를 찾지 못했습니다.")
            return
        
        self.status_label.setText(f"🔍 {len(results)}개 단지 발견, 이름 조회 중...")
        QApplication.processEvents()
        
        for source, cid in results:
            row = self.result_table.rowCount()
            self.result_table.insertRow(row)
            
            # 체크박스
            chk = QCheckBox()
            chk.setChecked(True)
            self.result_table.setCellWidget(row, 0, chk)
            
            # 단지 ID
            self.result_table.setItem(row, 1, QTableWidgetItem(cid))
            
            # 단지명 조회
            name = NaverURLParser.fetch_complex_name(cid)
            self.result_table.setItem(row, 2, QTableWidgetItem(name))
            
            # 상태
            status = "✅ 확인됨" if not name.startswith("단지_") else "⚠️ 이름 미확인"
            self.result_table.setItem(row, 3, QTableWidgetItem(status))
            
            QApplication.processEvents()
        
        self.status_label.setText(f"✅ {len(results)}개 단지 분석 완료")
    
    def _select_all(self):
        for row in range(self.result_table.rowCount()):
            chk = self.result_table.cellWidget(row, 0)
            if chk:
                chk.setChecked(True)
    
    def _add_selected(self):
        selected = []
        for row in range(self.result_table.rowCount()):
            chk = self.result_table.cellWidget(row, 0)
            if chk and chk.isChecked():
                cid = self.result_table.item(row, 1).text()
                name = self.result_table.item(row, 2).text()
                selected.append((name, cid))
        
        if selected:
            self.complexes_added.emit(selected)
            self.accept()
        else:
            QMessageBox.warning(self, "선택 필요", "추가할 단지를 선택하세요.")

class ExcelTemplateDialog(QDialog):
    """엑셀 내보내기 템플릿 설정 (v7.3)"""
    template_saved = pyqtSignal(dict)
    
    def __init__(self, parent=None, current_template=None):
        super().__init__(parent)
        self.setWindowTitle("📊 엑셀 템플릿 설정")
        self.setMinimumSize(400, 500)
        self.current_template = current_template or ExcelTemplate.get_default_template()
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        
        info = QLabel("내보낼 컬럼을 선택하고 순서를 조정하세요:")
        layout.addWidget(info)
        
        # 컬럼 목록
        self.column_list = QListWidget()
        self.column_list.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        
        for col_name in ExcelTemplate.get_column_order():
            item = QListWidgetItem(col_name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked if self.current_template.get(col_name, True) else Qt.CheckState.Unchecked)
            self.column_list.addItem(item)
        
        layout.addWidget(self.column_list)
        
        # 순서 조정 버튼
        order_layout = QHBoxLayout()
        btn_up = QPushButton("⬆️ 위로")
        btn_up.clicked.connect(self._move_up)
        btn_down = QPushButton("⬇️ 아래로")
        btn_down.clicked.connect(self._move_down)
        order_layout.addWidget(btn_up)
        order_layout.addWidget(btn_down)
        order_layout.addStretch()
        layout.addLayout(order_layout)
        
        # 전체 선택/해제
        select_layout = QHBoxLayout()
        btn_all = QPushButton("전체 선택")
        btn_all.clicked.connect(lambda: self._set_all(True))
        btn_none = QPushButton("전체 해제")
        btn_none.clicked.connect(lambda: self._set_all(False))
        btn_reset = QPushButton("기본값")
        btn_reset.clicked.connect(self._reset)
        select_layout.addWidget(btn_all)
        select_layout.addWidget(btn_none)
        select_layout.addWidget(btn_reset)
        select_layout.addStretch()
        layout.addLayout(select_layout)
        
        # 저장 버튼
        btn_save = QPushButton("💾 저장")
        btn_save.clicked.connect(self._save)
        layout.addWidget(btn_save)
    
    def _move_up(self):
        row = self.column_list.currentRow()
        if row > 0:
            item = self.column_list.takeItem(row)
            self.column_list.insertItem(row - 1, item)
            self.column_list.setCurrentRow(row - 1)
    
    def _move_down(self):
        row = self.column_list.currentRow()
        if row < self.column_list.count() - 1:
            item = self.column_list.takeItem(row)
            self.column_list.insertItem(row + 1, item)
            self.column_list.setCurrentRow(row + 1)
    
    def _set_all(self, checked):
        for i in range(self.column_list.count()):
            self.column_list.item(i).setCheckState(
                Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
            )
    
    def _reset(self):
        self.column_list.clear()
        default = ExcelTemplate.get_default_template()
        for col_name in ExcelTemplate.get_column_order():
            item = QListWidgetItem(col_name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked if default.get(col_name, True) else Qt.CheckState.Unchecked)
            self.column_list.addItem(item)
    
    def _save(self):
        template = {}
        column_order = []
        for i in range(self.column_list.count()):
            item = self.column_list.item(i)
            name = item.text()
            enabled = item.checkState() == Qt.CheckState.Checked
            template[name] = enabled
            column_order.append(name)
        
        result = {'columns': template, 'order': column_order}
        self.template_saved.emit(result)
        self.accept()

class SettingsDialog(QDialog):
    settings_changed = pyqtSignal(dict)
    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()
        self._load()
    
    def _setup_ui(self):
        self.setWindowTitle("⚙️ 설정")
        self.setMinimumSize(450, 450)
        layout = QVBoxLayout(self)
        
        # 테마
        tg = QGroupBox("🎨 테마")
        tl = QHBoxLayout()
        self.combo_theme = QComboBox()
        self.combo_theme.addItems(["dark", "light"])
        tl.addWidget(QLabel("테마:"))
        tl.addWidget(self.combo_theme)
        tl.addStretch()
        tg.setLayout(tl)
        layout.addWidget(tg)
        
        # 시스템
        sg = QGroupBox("🖥️ 시스템")
        sl = QVBoxLayout()
        self.check_tray = QCheckBox("닫기 시 트레이로 최소화")
        self.check_notify = QCheckBox("데스크톱 알림 표시")
        self.check_confirm = QCheckBox("종료 전 확인")
        self.check_sound = QCheckBox("크롤링 완료 시 알림음 재생")
        sl.addWidget(self.check_tray)
        sl.addWidget(self.check_notify)
        sl.addWidget(self.check_confirm)
        sl.addWidget(self.check_sound)
        sg.setLayout(sl)
        layout.addWidget(sg)
        
        # 크롤링
        cg = QGroupBox("🔄 크롤링")
        cl = QHBoxLayout()
        self.combo_speed = QComboBox()
        self.combo_speed.addItems(list(CRAWL_SPEED_PRESETS.keys()))
        cl.addWidget(QLabel("기본 속도:"))
        cl.addWidget(self.combo_speed)
        cl.addStretch()
        cg.setLayout(cl)
        layout.addWidget(cg)
        
        # 정렬
        og = QGroupBox("📊 결과 정렬")
        ol = QHBoxLayout()
        self.combo_sort_col = QComboBox()
        self.combo_sort_col.addItems(["가격", "면적", "단지명", "거래유형"])
        ol.addWidget(QLabel("기준:"))
        ol.addWidget(self.combo_sort_col)
        self.combo_sort_order = QComboBox()
        self.combo_sort_order.addItems(["오름차순", "내림차순"])
        ol.addWidget(self.combo_sort_order)
        ol.addStretch()
        og.setLayout(ol)
        layout.addWidget(og)
        
        layout.addStretch()
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def _load(self):
        self.combo_theme.setCurrentText(settings.get("theme", "dark"))
        self.check_tray.setChecked(settings.get("minimize_to_tray", True))
        self.check_notify.setChecked(settings.get("show_notifications", True))
        self.check_confirm.setChecked(settings.get("confirm_before_close", True))
        self.check_sound.setChecked(settings.get("play_sound_on_complete", True))
        self.combo_speed.setCurrentText(settings.get("crawl_speed", "보통"))
        self.combo_sort_col.setCurrentText(settings.get("default_sort_column", "가격"))
        self.combo_sort_order.setCurrentText("오름차순" if settings.get("default_sort_order", "asc") == "asc" else "내림차순")
    
    def _save(self):
        new = {
            "theme": self.combo_theme.currentText(),
            "minimize_to_tray": self.check_tray.isChecked(),
            "show_notifications": self.check_notify.isChecked(),
            "confirm_before_close": self.check_confirm.isChecked(),
            "play_sound_on_complete": self.check_sound.isChecked(),
            "crawl_speed": self.combo_speed.currentText(),
            "default_sort_column": self.combo_sort_col.currentText(),
            "default_sort_order": "asc" if self.combo_sort_order.currentText() == "오름차순" else "desc"
        }
        settings.update(new)
        self.settings_changed.emit(new)
        self.accept()

class ShortcutsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⌨️ 단축키")
        self.setMinimumSize(400, 350)
        layout = QVBoxLayout(self)
        tbl = QTableWidget()
        tbl.setColumnCount(2)
        tbl.setHorizontalHeaderLabels(["기능", "단축키"])
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        tbl.setAlternatingRowColors(True)
        shortcuts = [
            ("🚀 크롤링 시작", SHORTCUTS["start_crawl"]),
            ("⏹️ 크롤링 중지", SHORTCUTS["stop_crawl"]),
            ("💾 Excel 저장", SHORTCUTS["save_excel"]),
            ("📄 CSV 저장", SHORTCUTS["save_csv"]),
            ("🔄 새로고침", SHORTCUTS["refresh"]),
            ("🔍 검색", SHORTCUTS["search"]),
            ("🎨 테마 변경", SHORTCUTS["toggle_theme"]),
            ("📥 트레이 최소화", SHORTCUTS["minimize_tray"]),
            ("❌ 종료", SHORTCUTS["quit"])
        ]
        tbl.setRowCount(len(shortcuts))
        for i, (d, k) in enumerate(shortcuts):
            tbl.setItem(i, 0, QTableWidgetItem(d))
            tbl.setItem(i, 1, QTableWidgetItem(k))
        layout.addWidget(tbl)
        btn = QPushButton("닫기")
        btn.clicked.connect(self.accept)
        layout.addWidget(btn)

class ChartWidget(QWidget):
    """v10.0: Analytics Chart using Matplotlib"""
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        if MATPLOTLIB_AVAILABLE:
            self.figure = Figure(figsize=(5, 3), dpi=100, facecolor='#2b2b2b')
            self.canvas = FigureCanvas(self.figure)
            self.ax = self.figure.add_subplot(111)
            self.ax.set_facecolor('#2b2b2b')
            self.ax.tick_params(colors='white')
            self.ax.xaxis.label.set_color('white')
            self.ax.yaxis.label.set_color('white')
            for spine in self.ax.spines.values():
                spine.set_color('#555555')
            layout.addWidget(self.canvas)
        else:
            layout.addWidget(QLabel("Matplotlib 라이브러리가 설치되지 않았습니다.\n(pip install matplotlib)"))

    def update_chart(self, data: List[Tuple[str, int]]):
        if not MATPLOTLIB_AVAILABLE or not data: return
        self.ax.clear()
        
        # Sort by date
        data.sort(key=lambda x: x[0])
        
        dates = [datetime.strptime(d[0], "%Y-%m-%d") for d in data]
        prices = [d[1] for d in data]
        
        self.ax.plot(dates, prices, marker='o', linestyle='-', color='#3498db', linewidth=2)
        self.ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
        self.ax.grid(True, linestyle='--', alpha=0.3)
        self.ax.set_title("Price Trend", color='white')
        self.canvas.draw()

class AboutDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("ℹ️ 정보")
        self.setMinimumSize(500, 500)
        layout = QVBoxLayout(self)
        browser = QTextBrowser()
        browser.setOpenExternalLinks(True)
        browser.setHtml(f"""
        <div style="text-align: center; padding: 20px;">
            <h1 style="color: #3b82f6; margin-bottom: 5px;">🏠 네이버 부동산 크롤러</h1>
            <h2 style="margin-top: 0;">Pro Plus {APP_VERSION}</h2>
            <p style="color: #64748b; font-size: 14px;">UI/UX 대폭 개선 버전</p>
        </div>
        
        <h3 style="color: #3b82f6; border-bottom: 2px solid #3b82f6; padding-bottom: 5px;">🆕 v10.5 업데이트</h3>
        <ul>
            <li>🎨 <b>현대적인 UI 테마</b> - Glassmorphism 효과, 그라데이션 버튼</li>
            <li>🌗 <b>개선된 다크/라이트 모드</b> - 더 나은 색상 대비 및 가독성</li>
            <li>🔧 <b>버그 수정</b> - 매물 URL 열기, 예외 처리 개선</li>
            <li>⚡ <b>성능 최적화</b> - 스레드 안전성 강화</li>
        </ul>
        
        <h3 style="color: #22c55e; border-bottom: 2px solid #22c55e; padding-bottom: 5px;">✨ 핵심 기능</h3>
        <ul>
            <li>📊 다중 단지 동시 크롤링</li>
            <li>🔍 가격/면적/층수 필터링</li>
            <li>📁 필터 프리셋 저장/불러오기</li>
            <li>🔔 조건 알림 시스템</li>
            <li>📈 시세 변동 추적 및 차트</li>
            <li>💾 Excel/CSV/JSON 내보내기</li>
            <li>🆕 신규 매물 및 가격 변동 표시</li>
            <li>⏱️ 예상 남은 시간 표시</li>
        </ul>
        
        <h3 style="color: #8b5cf6; border-bottom: 2px solid #8b5cf6; padding-bottom: 5px;">⌨️ 단축키</h3>
        <table style="width: 100%; border-collapse: collapse; margin-top: 10px;">
            <tr style="background-color: rgba(59, 130, 246, 0.1);">
                <td style="padding: 8px; border: 1px solid #e2e8f0;">Ctrl+R</td>
                <td style="padding: 8px; border: 1px solid #e2e8f0;">크롤링 시작</td>
            </tr>
            <tr>
                <td style="padding: 8px; border: 1px solid #e2e8f0;">Ctrl+S</td>
                <td style="padding: 8px; border: 1px solid #e2e8f0;">Excel 저장</td>
            </tr>
            <tr style="background-color: rgba(59, 130, 246, 0.1);">
                <td style="padding: 8px; border: 1px solid #e2e8f0;">Ctrl+T</td>
                <td style="padding: 8px; border: 1px solid #e2e8f0;">테마 변경</td>
            </tr>
            <tr>
                <td style="padding: 8px; border: 1px solid #e2e8f0;">F5</td>
                <td style="padding: 8px; border: 1px solid #e2e8f0;">새로고침</td>
            </tr>
        </table>
        
        <p style="color: #64748b; margin-top: 20px; text-align: center; font-size: 12px;">
            Made with ❤️ using Claude & Gemini AI
        </p>
        """)
        layout.addWidget(browser)
        btn = QPushButton("닫기")
        btn.clicked.connect(self.accept)
        layout.addWidget(btn)

class RecentSearchDialog(QDialog):
    """최근 검색 기록 다이얼로그"""
    def __init__(self, parent=None, history_manager=None):
        super().__init__(parent)
        self.history_manager = history_manager
        self.selected_search = None
        self._setup_ui()
    
    def _setup_ui(self):
        self.setWindowTitle("🕐 최근 검색 기록")
        self.setMinimumSize(500, 400)
        layout = QVBoxLayout(self)
        
        self.list = QListWidget()
        self.list.setAlternatingRowColors(True)
        self.list.itemDoubleClicked.connect(self._load)
        layout.addWidget(self.list)
        
        btn_layout = QHBoxLayout()
        btn_load = QPushButton("📂 불러오기")
        btn_load.clicked.connect(self._load)
        btn_clear = QPushButton("🗑️ 기록 지우기")
        btn_clear.clicked.connect(self._clear)
        btn_layout.addWidget(btn_load)
        btn_layout.addWidget(btn_clear)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        self._refresh()
    
    def _refresh(self):
        self.list.clear()
        if self.history_manager:
            for h in self.history_manager.get_recent():
                complexes = h.get('complexes', [])
                types = h.get('trade_types', [])
                timestamp = h.get('timestamp', '')
                text = f"[{timestamp}] {len(complexes)}개 단지 - {', '.join(types)}"
                item = QListWidgetItem(text)
                item.setData(Qt.ItemDataRole.UserRole, h)
                self.list.addItem(item)
    
    def _load(self):
        if item := self.list.currentItem():
            self.selected_search = item.data(Qt.ItemDataRole.UserRole)
            self.accept()
    
    def _clear(self):
        if self.history_manager:
            self.history_manager.clear()
            self._refresh()

# ============ MAIN WINDOW ============
class RealEstateApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.setMinimumSize(1400, 900)
        geo = settings.get("window_geometry")
        if geo: self.setGeometry(*geo)
        else: self.setGeometry(100, 100, 1500, 950)
        
        self.db = ComplexDatabase()
        self.preset_manager = FilterPresetManager()
        self.history_manager = SearchHistoryManager()
        self.crawler_thread = None
        self.collected_data = []
        self.grouped_rows = {}
        self.is_scheduled_run = False
        self.current_theme = settings.get("theme", "dark")
        self.tray_icon = None
        self.crawl_stats = {"매매": 0, "전세": 0, "월세": 0, "new": 0, "price_up": 0, "price_down": 0}
        self.advanced_filters = None  # v7.3: 고급 필터
        
        # v11.0: Toast 알림 시스템
        self.toast_widgets: List[ToastWidget] = []
        
        self.setStyleSheet(get_stylesheet(self.current_theme))
        self._init_ui()
        self._init_menu()
        self._init_shortcuts()
        self._init_tray()
        self._init_timers()
        self._load_initial_data()
        self.status_bar.showMessage(f"✨ {APP_TITLE} 준비 완료")
    
    def _init_ui(self):
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)
        layout.setContentsMargins(8, 8, 8, 8)
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)
        self._setup_crawler_tab()
        self._setup_db_tab()
        self._setup_groups_tab()
        self._setup_schedule_tab()
        self._setup_history_tab()
        self._setup_stats_tab()
        self._setup_guide_tab()
        self.status_bar = self.statusBar()
    
    def _setup_crawler_tab(self):
        tab = QWidget()
        layout = QHBoxLayout(tab)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left panel - v11.0: 동적 크기 조정
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setMinimumWidth(380)  # 최소 너비 감소
        # 최대 너비 제한 제거하여 해상도에 따라 유연하게 조정
        scroll.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
        scroll_content = QWidget()
        left = QVBoxLayout(scroll_content)
        left.setSpacing(10)
        
        # 1. 거래유형
        tg = QGroupBox("1️⃣ 거래 유형")
        tl = QHBoxLayout()
        self.check_trade = QCheckBox("매매")
        self.check_trade.setChecked(True)
        self.check_trade.setToolTip("아파트 매매 매물을 검색합니다")
        self.check_jeonse = QCheckBox("전세")
        self.check_jeonse.setChecked(True)
        self.check_jeonse.setToolTip("전세 매물을 검색합니다")
        self.check_monthly = QCheckBox("월세")
        self.check_monthly.setToolTip("월세 매물을 검색합니다")
        tl.addWidget(self.check_trade)
        tl.addWidget(self.check_jeonse)
        tl.addWidget(self.check_monthly)
        tl.addStretch()
        tg.setLayout(tl)
        left.addWidget(tg)
        
        # 2. 면적 필터
        ag = QGroupBox("2️⃣ 면적 필터")
        al = QVBoxLayout()
        self.check_area_filter = QCheckBox("면적 필터 사용")
        self.check_area_filter.stateChanged.connect(self._toggle_area_filter)
        al.addWidget(self.check_area_filter)
        area_input = QHBoxLayout()
        self.spin_area_min = QSpinBox()
        self.spin_area_min.setRange(0, 300)
        self.spin_area_min.setEnabled(False)
        self.spin_area_min.setToolTip("최소 면적 (㎡)")
        self.spin_area_max = QSpinBox()
        self.spin_area_max.setRange(0, 300)
        self.spin_area_max.setValue(200)
        self.spin_area_max.setEnabled(False)
        self.spin_area_max.setToolTip("최대 면적 (㎡)")
        area_input.addWidget(QLabel("최소:"))
        area_input.addWidget(self.spin_area_min)
        area_input.addWidget(QLabel("㎡  ~  최대:"))
        area_input.addWidget(self.spin_area_max)
        area_input.addWidget(QLabel("㎡"))
        al.addLayout(area_input)
        ag.setLayout(al)
        left.addWidget(ag)
        
        # 3. 가격 필터
        pg = QGroupBox("3️⃣ 가격 필터")
        pl = QVBoxLayout()
        self.check_price_filter = QCheckBox("가격 필터 사용")
        self.check_price_filter.stateChanged.connect(self._toggle_price_filter)
        pl.addWidget(self.check_price_filter)
        
        price_grid = QGridLayout()
        # 매매
        price_grid.addWidget(QLabel("매매:"), 0, 0)
        self.spin_trade_min = QSpinBox()
        self.spin_trade_min.setRange(0, 999999)
        self.spin_trade_min.setSingleStep(1000)
        self.spin_trade_min.setEnabled(False)
        self.spin_trade_min.setToolTip("매매 최소 가격 (만원)")
        price_grid.addWidget(self.spin_trade_min, 0, 1)
        price_grid.addWidget(QLabel("~"), 0, 2)
        self.spin_trade_max = QSpinBox()
        self.spin_trade_max.setRange(0, 999999)
        self.spin_trade_max.setValue(100000)
        self.spin_trade_max.setSingleStep(1000)
        self.spin_trade_max.setEnabled(False)
        self.spin_trade_max.setToolTip("매매 최대 가격 (만원)")
        price_grid.addWidget(self.spin_trade_max, 0, 3)
        price_grid.addWidget(QLabel("만원"), 0, 4)
        
        # 전세
        price_grid.addWidget(QLabel("전세:"), 1, 0)
        self.spin_jeonse_min = QSpinBox()
        self.spin_jeonse_min.setRange(0, 999999)
        self.spin_jeonse_min.setSingleStep(1000)
        self.spin_jeonse_min.setEnabled(False)
        price_grid.addWidget(self.spin_jeonse_min, 1, 1)
        price_grid.addWidget(QLabel("~"), 1, 2)
        self.spin_jeonse_max = QSpinBox()
        self.spin_jeonse_max.setRange(0, 999999)
        self.spin_jeonse_max.setValue(50000)
        self.spin_jeonse_max.setSingleStep(1000)
        self.spin_jeonse_max.setEnabled(False)
        price_grid.addWidget(self.spin_jeonse_max, 1, 3)
        price_grid.addWidget(QLabel("만원"), 1, 4)
        
        # 월세
        price_grid.addWidget(QLabel("월세:"), 2, 0)
        self.spin_monthly_min = QSpinBox()
        self.spin_monthly_min.setRange(0, 999999)
        self.spin_monthly_min.setSingleStep(100)
        self.spin_monthly_min.setEnabled(False)
        price_grid.addWidget(self.spin_monthly_min, 2, 1)
        price_grid.addWidget(QLabel("~"), 2, 2)
        self.spin_monthly_max = QSpinBox()
        self.spin_monthly_max.setRange(0, 999999)
        self.spin_monthly_max.setValue(5000)
        self.spin_monthly_max.setSingleStep(100)
        self.spin_monthly_max.setEnabled(False)
        price_grid.addWidget(self.spin_monthly_max, 2, 3)
        price_grid.addWidget(QLabel("만원"), 2, 4)
        
        pl.addLayout(price_grid)
        pg.setLayout(pl)
        left.addWidget(pg)
        
        # 4. 단지 목록
        cg = QGroupBox("4️⃣ 단지 목록")
        cl = QVBoxLayout()
        load_btn = QHBoxLayout()
        btn_db = QPushButton("💾 DB에서")
        btn_db.setToolTip("저장된 단지 DB에서 불러오기")
        btn_db.clicked.connect(self._show_db_load_dialog)
        btn_grp = QPushButton("📁 그룹에서")
        btn_grp.setToolTip("저장된 그룹에서 불러오기")
        btn_grp.clicked.connect(self._show_group_load_dialog)
        btn_history = QPushButton("🕐 최근검색")
        btn_history.setToolTip("최근 검색 기록에서 불러오기")
        btn_history.clicked.connect(self._show_history_dialog)
        load_btn.addWidget(btn_db)
        load_btn.addWidget(btn_grp)
        load_btn.addWidget(btn_history)
        cl.addLayout(load_btn)
        
        input_layout = QHBoxLayout()
        self.input_name = QLineEdit()
        self.input_name.setPlaceholderText("단지명")
        self.input_name.setToolTip("아파트 단지명을 입력하세요")
        self.input_id = QLineEdit()
        self.input_id.setPlaceholderText("단지 ID")
        self.input_id.setToolTip("네이버 부동산 URL에서 단지 ID를 확인할 수 있습니다")
        btn_add = QPushButton("➕")
        btn_add.setMaximumWidth(45)
        btn_add.setToolTip("단지 추가")
        btn_add.clicked.connect(self._add_complex)
        input_layout.addWidget(self.input_name, 2)
        input_layout.addWidget(self.input_id, 1)
        input_layout.addWidget(btn_add)
        cl.addLayout(input_layout)
        
        self.table_list = QTableWidget()
        self.table_list.setColumnCount(2)
        self.table_list.setHorizontalHeaderLabels(["단지명", "ID"])
        self.table_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table_list.setMinimumHeight(130)
        self.table_list.setAlternatingRowColors(True)
        self.table_list.setToolTip("더블클릭하면 네이버 부동산 페이지를 엽니다")
        self.table_list.doubleClicked.connect(self._open_complex_url)
        cl.addWidget(self.table_list)
        
        manage_btn = QHBoxLayout()
        btn_del = QPushButton("🗑️ 삭제")
        btn_del.clicked.connect(self._delete_complex)
        btn_clr = QPushButton("🧹 초기화")
        btn_clr.clicked.connect(self._clear_list)
        btn_sv = QPushButton("💾 DB저장")
        btn_sv.clicked.connect(self._save_to_db)
        # v7.3: URL 일괄 등록
        btn_url = QPushButton("🔗 URL등록")
        btn_url.setToolTip("URL 또는 단지ID 일괄 등록")
        btn_url.clicked.connect(self._show_url_batch_dialog)
        manage_btn.addWidget(btn_del)
        manage_btn.addWidget(btn_clr)
        manage_btn.addWidget(btn_sv)
        manage_btn.addWidget(btn_url)
        cl.addLayout(manage_btn)
        cg.setLayout(cl)
        left.addWidget(cg)
        
        # 5. 속도
        spg = QGroupBox("5️⃣ 크롤링 속도")
        spl = QVBoxLayout()
        self.speed_slider = SpeedSlider()
        self.speed_slider.set_speed(settings.get("crawl_speed", "보통"))
        spl.addWidget(self.speed_slider)
        spg.setLayout(spl)
        left.addWidget(spg)
        
        # 6. 실행
        eg = QGroupBox("6️⃣ 실행")
        el = QHBoxLayout()
        self.btn_start = QPushButton("▶️ 크롤링 시작")
        self.btn_start.setObjectName("startButton")
        self.btn_start.setMinimumHeight(45)
        self.btn_start.setToolTip(f"크롤링 시작 ({SHORTCUTS['start_crawl']})")
        self.btn_start.clicked.connect(self._start_crawling)
        self.btn_stop = QPushButton("⏹️ 중지")
        self.btn_stop.setObjectName("stopButton")
        self.btn_stop.setEnabled(False)
        self.btn_stop.setToolTip(f"크롤링 중지 ({SHORTCUTS['stop_crawl']})")
        self.btn_stop.clicked.connect(self._stop_crawling)
        self.btn_save = QPushButton("💾 저장")
        self.btn_save.setObjectName("saveButton")
        self.btn_save.setEnabled(False)
        self.btn_save.setToolTip("결과 저장 (Excel, CSV, JSON)")
        self.btn_save.clicked.connect(self._show_save_menu)
        el.addWidget(self.btn_start, 2)
        el.addWidget(self.btn_stop, 1)
        el.addWidget(self.btn_save, 1)
        eg.setLayout(el)
        left.addWidget(eg)
        
        left.addStretch()
        scroll.setWidget(scroll_content)
        splitter.addWidget(scroll)
        
        # Right panel
        right_w = QWidget()
        right = QVBoxLayout(right_w)
        right.setSpacing(8)
        
        # 요약 카드
        self.summary_card = SummaryCard(theme=self.current_theme)
        right.addWidget(self.summary_card)
        
        # 검색 및 정렬
        search_sort = QHBoxLayout()
        self.result_search = SearchBar("결과 검색...")
        self.result_search.search_changed.connect(self._filter_results)
        search_sort.addWidget(self.result_search, 3)
        
        # v7.3: 고급 필터 버튼
        btn_adv_filter = QPushButton("🔍 고급 필터")
        btn_adv_filter.setToolTip("가격, 면적, 층수 등 상세 필터")
        btn_adv_filter.clicked.connect(self._show_advanced_filter)
        search_sort.addWidget(btn_adv_filter)
        
        search_sort.addWidget(QLabel("정렬:"))
        self.combo_sort = QComboBox()
        self.combo_sort.addItems(["가격 ↑", "가격 ↓", "면적 ↑", "면적 ↓", "단지명 ↑", "단지명 ↓"])
        self.combo_sort.setToolTip("결과 정렬 기준")
        self.combo_sort.currentTextChanged.connect(self._sort_results)
        search_sort.addWidget(self.combo_sort, 1)
        right.addLayout(search_sort)
        
        # 결과 탭
        result_tabs = QTabWidget()
        result_tab = QWidget()
        rl = QVBoxLayout(result_tab)
        rl.setContentsMargins(0, 5, 0, 0)
        
        # v7.3: 확장된 컬럼 (신규, 변동 추가)
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(12)
        self.result_table.setHorizontalHeaderLabels([
            "단지명", "거래", "가격", "면적", "층/방향", "특징", 
            "🆕", "📊 변동", "시각", "링크", "URL", "가격(숫자)"
        ])
        self.result_table.setColumnHidden(10, True)  # URL 컬럼 숨김
        self.result_table.setColumnHidden(11, True)  # 가격 숫자 컬럼 숨김 (정렬용)
        self.result_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.result_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        self.result_table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeMode.Fixed)
        self.result_table.setColumnWidth(6, 40)  # 신규
        self.result_table.setColumnWidth(7, 80)  # 변동
        self.result_table.setColumnWidth(9, 80)  # 링크
        self.result_table.setSortingEnabled(True)
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setToolTip("더블클릭하면 해당 매물 페이지를 엽니다")
        self.result_table.doubleClicked.connect(self._open_article_url)
        rl.addWidget(self.result_table)
        result_tabs.addTab(result_tab, "📊 결과")
        
        log_tab = QWidget()
        ll = QVBoxLayout(log_tab)
        ll.setContentsMargins(0, 5, 0, 0)
        self.log_browser = QTextBrowser()
        self.log_browser.setMinimumHeight(150)
        ll.addWidget(self.log_browser)
        result_tabs.addTab(log_tab, "📝 로그")
        right.addWidget(result_tabs)
        
        # 진행 상태
        self.progress_widget = ProgressWidget()
        right.addWidget(self.progress_widget)
        
        splitter.addWidget(right_w)
        splitter.setSizes([450, 900])
        layout.addWidget(splitter)
        self.tabs.addTab(tab, "🏘️ 데이터 수집")
    
    def _setup_db_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        bl = QHBoxLayout()
        btn_rf = QPushButton("🔄 새로고침")
        btn_rf.clicked.connect(self._load_db_complexes)
        btn_dl = QPushButton("🗑️ 선택 삭제")
        btn_dl.clicked.connect(self._delete_db_complex)
        btn_dlm = QPushButton("🗑️ 다중 삭제")
        btn_dlm.clicked.connect(self._delete_db_complexes_multi)
        btn_memo = QPushButton("✏️ 메모 수정")
        btn_memo.clicked.connect(self._edit_memo)
        bl.addWidget(btn_rf)
        bl.addWidget(btn_dl)
        bl.addWidget(btn_dlm)
        bl.addWidget(btn_memo)
        bl.addStretch()
        layout.addLayout(bl)
        self.db_search = SearchBar("단지 검색...")
        self.db_search.search_changed.connect(self._filter_db_table)
        layout.addWidget(self.db_search)
        self.db_table = QTableWidget()
        self.db_table.setColumnCount(4)
        self.db_table.setHorizontalHeaderLabels(["ID", "단지명", "단지ID", "메모"])
        self.db_table.setColumnHidden(0, True)
        self.db_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.db_table.setAlternatingRowColors(True)
        self.db_table.doubleClicked.connect(self._open_db_complex_url)
        layout.addWidget(self.db_table)
        self.tabs.addTab(tab, "💾 단지 DB")
    
    def _setup_groups_tab(self):
        tab = QWidget()
        layout = QHBoxLayout(tab)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # 그룹 목록
        left_w = QWidget()
        left_l = QVBoxLayout(left_w)
        left_l.addWidget(QLabel("📁 그룹 목록"))
        gl = QHBoxLayout()
        btn_new = QPushButton("➕ 새 그룹")
        btn_new.clicked.connect(self._create_group)
        btn_del = QPushButton("🗑️ 삭제")
        btn_del.clicked.connect(self._delete_group)
        gl.addWidget(btn_new)
        gl.addWidget(btn_del)
        left_l.addLayout(gl)
        self.group_list = QListWidget()
        self.group_list.setAlternatingRowColors(True)
        self.group_list.itemClicked.connect(self._load_group_complexes)
        left_l.addWidget(self.group_list)
        splitter.addWidget(left_w)
        
        # 그룹 내 단지
        right_w = QWidget()
        right_l = QVBoxLayout(right_w)
        right_l.addWidget(QLabel("📋 그룹 내 단지"))
        rl = QHBoxLayout()
        btn_add = QPushButton("➕ 단지 추가")
        btn_add.clicked.connect(self._add_to_group)
        btn_add_multi = QPushButton("➕ 다중 추가")
        btn_add_multi.clicked.connect(self._add_to_group_multi)
        btn_rm = QPushButton("➖ 제거")
        btn_rm.clicked.connect(self._remove_from_group)
        rl.addWidget(btn_add)
        rl.addWidget(btn_add_multi)
        rl.addWidget(btn_rm)
        right_l.addLayout(rl)
        self.group_complex_table = QTableWidget()
        self.group_complex_table.setColumnCount(4)
        self.group_complex_table.setHorizontalHeaderLabels(["ID", "단지명", "단지ID", "메모"])
        self.group_complex_table.setColumnHidden(0, True)
        self.group_complex_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.group_complex_table.setAlternatingRowColors(True)
        right_l.addWidget(self.group_complex_table)
        splitter.addWidget(right_w)
        
        splitter.setSizes([300, 700])
        layout.addWidget(splitter)
        self.tabs.addTab(tab, "📁 그룹 관리")
    
    def _setup_schedule_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        sg = QGroupBox("⏰ 예약 크롤링")
        sl = QVBoxLayout()
        self.check_schedule = QCheckBox("예약 실행 활성화")
        sl.addWidget(self.check_schedule)
        tl = QHBoxLayout()
        tl.addWidget(QLabel("실행 시간:"))
        self.time_edit = QTimeEdit()
        self.time_edit.setTime(QTime(9, 0))
        tl.addWidget(self.time_edit)
        tl.addStretch()
        sl.addLayout(tl)
        gl = QHBoxLayout()
        gl.addWidget(QLabel("대상 그룹:"))
        self.schedule_group_combo = QComboBox()
        gl.addWidget(self.schedule_group_combo)
        gl.addStretch()
        sl.addLayout(gl)
        sg.setLayout(sl)
        layout.addWidget(sg)
        layout.addStretch()
        self.tabs.addTab(tab, "⏰ 예약 설정")
    
    def _setup_history_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        bl = QHBoxLayout()
        btn_rf = QPushButton("🔄 새로고침")
        btn_rf.clicked.connect(self._load_history)
        bl.addWidget(btn_rf)
        bl.addStretch()
        layout.addLayout(bl)
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(5)
        self.history_table.setHorizontalHeaderLabels(["단지명", "단지ID", "거래유형", "수집건수", "수집시각"])
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.history_table.setAlternatingRowColors(True)
        layout.addWidget(self.history_table)
        self.tabs.addTab(tab, "📜 히스토리")
    
    def _setup_stats_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        fl = QHBoxLayout()
        fl.addWidget(QLabel("단지:"))
        self.stats_complex_combo = QComboBox()
        fl.addWidget(self.stats_complex_combo)
        fl.addWidget(QLabel("유형:"))
        self.stats_type_combo = QComboBox()
        self.stats_type_combo.addItems(["전체", "매매", "전세", "월세"])
        fl.addWidget(self.stats_type_combo)
        
        fl.addWidget(QLabel("면적:"))
        self.stats_pyeong_combo = QComboBox()
        self.stats_pyeong_combo.addItem("전체")
        fl.addWidget(self.stats_pyeong_combo)
        
        btn_load = QPushButton("📊 조회")
        btn_load.clicked.connect(self._load_stats)
        fl.addWidget(btn_load)
        fl.addStretch()
        layout.addLayout(fl)
        self.stats_table = QTableWidget()
        self.stats_table.setColumnCount(6)
        self.stats_table.setHorizontalHeaderLabels(["날짜", "유형", "평형", "최저가", "최고가", "평균가"])
        self.stats_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.stats_table.setAlternatingRowColors(True)
        
        # v10.0: Chart Integration
        splitter = QSplitter(Qt.Orientation.Vertical)
        splitter.addWidget(self.stats_table)
        
        self.chart_widget = ChartWidget()
        splitter.addWidget(self.chart_widget)
        splitter.setSizes([300, 300])
        
        layout.addWidget(splitter)
        self.tabs.addTab(tab, "📈 통계/변동")
    
    def _setup_guide_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        browser = QTextBrowser()
        browser.setOpenExternalLinks(True)
        browser.setHtml("""
        <h2>📖 사용 가이드</h2>
        <h3>🔍 단지 ID 찾는 방법</h3>
        <ol>
            <li>네이버 부동산 (<a href="https://new.land.naver.com">new.land.naver.com</a>) 접속</li>
            <li>원하는 아파트 단지 검색</li>
            <li>URL에서 <code>/complexes/</code> 뒤의 숫자가 단지 ID입니다</li>
            <li>예: <code>https://new.land.naver.com/complexes/<b>12345</b></code> → ID: 12345</li>
        </ol>
        <h3>⌨️ 단축키</h3>
        <table border="1" cellpadding="8" style="border-collapse: collapse;">
            <tr><th>기능</th><th>단축키</th></tr>
            <tr><td>🚀 크롤링 시작</td><td>Ctrl+R</td></tr>
            <tr><td>⏹️ 크롤링 중지</td><td>Ctrl+Shift+R</td></tr>
            <tr><td>💾 Excel 저장</td><td>Ctrl+S</td></tr>
            <tr><td>📄 CSV 저장</td><td>Ctrl+Shift+S</td></tr>
            <tr><td>🔍 검색</td><td>Ctrl+F</td></tr>
            <tr><td>🎨 테마 변경</td><td>Ctrl+T</td></tr>
            <tr><td>📥 트레이 최소화</td><td>Ctrl+M</td></tr>
        </table>
        <h3>💡 팁</h3>
        <ul>
            <li>🖱️ 결과 테이블에서 <b>더블클릭</b>하면 해당 매물 페이지로 이동합니다</li>
            <li>📊 요약 카드에서 실시간 수집 현황을 확인할 수 있습니다</li>
            <li>⏱️ 예상 남은 시간을 참고하여 작업 시간을 예측하세요</li>
            <li>🔔 알림 설정을 통해 원하는 조건의 매물을 알림받을 수 있습니다</li>
        </ul>
        """)
        layout.addWidget(browser)
        self.tabs.addTab(tab, "📖 가이드")
    
    def _init_menu(self):
        menubar = self.menuBar()
        
        # 파일 메뉴
        file_menu = menubar.addMenu("📂 파일")
        file_menu.addAction("💾 DB 백업", self._backup_db)
        file_menu.addAction("📂 DB 복원", self._restore_db)
        file_menu.addSeparator()
        file_menu.addAction("⚙️ 설정", self._show_settings)
        file_menu.addSeparator()
        file_menu.addAction("❌ 종료", self._quit_app)
        
        # 필터 메뉴
        filter_menu = menubar.addMenu("🔍 필터")
        filter_menu.addAction("💾 현재 필터 저장", self._save_preset)
        filter_menu.addAction("📂 필터 불러오기", self._load_preset)
        
        # 알림 메뉴
        alert_menu = menubar.addMenu("🔔 알림")
        alert_menu.addAction("⚙️ 알림 설정", self._show_alert_settings)
        
        # 보기 메뉴
        view_menu = menubar.addMenu("👁️ 보기")
        view_menu.addAction("🎨 테마 전환", self._toggle_theme)
        
        # 도움말 메뉴
        help_menu = menubar.addMenu("❓ 도움말")
        help_menu.addAction("⌨️ 단축키", self._show_shortcuts)
        help_menu.addAction("ℹ️ 정보", self._show_about)
    
    def _init_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+R"), self, self._start_crawling)
        QShortcut(QKeySequence("Ctrl+Shift+R"), self, self._stop_crawling)
        QShortcut(QKeySequence("Ctrl+S"), self, self._save_excel)
        QShortcut(QKeySequence("Ctrl+Shift+S"), self, self._save_csv)
        QShortcut(QKeySequence("F5"), self, self._refresh_tab)
        QShortcut(QKeySequence("Ctrl+F"), self, self._focus_search)
        QShortcut(QKeySequence("Ctrl+T"), self, self._toggle_theme)
        QShortcut(QKeySequence("Ctrl+M"), self, self._minimize_to_tray)
        QShortcut(QKeySequence("Ctrl+Q"), self, self._quit_app)
    
    def _init_tray(self):
        if QSystemTrayIcon.isSystemTrayAvailable():
            self.tray_icon = QSystemTrayIcon(self)
            self.tray_icon.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
            tray_menu = QMenu()
            tray_menu.addAction("🔼 열기", self._show_from_tray)
            tray_menu.addAction("❌ 종료", self._quit_app)
            self.tray_icon.setContextMenu(tray_menu)
            self.tray_icon.activated.connect(self._tray_activated)
            self.tray_icon.show()
    
    def _init_timers(self):
        self.schedule_timer = QTimer(self)
        self.schedule_timer.timeout.connect(self._check_schedule)
        self.schedule_timer.start(60000)
    
    def _load_initial_data(self):
        self._load_db_complexes()
        self._load_all_groups()
        self._load_history()
        self._load_stats_complexes()
        self._load_schedule_groups()
        
        # Connect signals after loading
        self.stats_complex_combo.currentIndexChanged.connect(self._on_stats_complex_changed)
    
    # Event handlers
    def _toggle_area_filter(self, state):
        enabled = state == Qt.CheckState.Checked.value
        self.spin_area_min.setEnabled(enabled)
        self.spin_area_max.setEnabled(enabled)
    
    def _toggle_price_filter(self, state):
        enabled = state == Qt.CheckState.Checked.value
        for w in [self.spin_trade_min, self.spin_trade_max, self.spin_jeonse_min, 
                  self.spin_jeonse_max, self.spin_monthly_min, self.spin_monthly_max]:
            w.setEnabled(enabled)
    
    def _add_complex(self):
        name = self.input_name.text().strip()
        cid = self.input_id.text().strip()
        if name and cid:
            self._add_row(name, cid)
            self.input_name.clear()
            self.input_id.clear()
    
    def _add_row(self, name, cid):
        row = self.table_list.rowCount()
        self.table_list.insertRow(row)
        self.table_list.setItem(row, 0, QTableWidgetItem(name))
        self.table_list.setItem(row, 1, QTableWidgetItem(cid))
    
    def _delete_complex(self):
        row = self.table_list.currentRow()
        if row >= 0:
            self.table_list.removeRow(row)
    
    def _clear_list(self):
        self.table_list.setRowCount(0)
    
    def _save_to_db(self):
        """단지를 DB에 저장 - 디버깅 강화"""
        count = 0
        total = self.table_list.rowCount()
        print(f"[UI] DB 저장 시작: {total}개 단지")
        
        for r in range(total):
            name_item = self.table_list.item(r, 0)
            cid_item = self.table_list.item(r, 1)
            
            if not name_item or not cid_item:
                print(f"[UI WARN] 행 {r}: 데이터 없음")
                continue
            
            name = name_item.text().strip()
            cid = cid_item.text().strip()
            
            if not name or not cid:
                print(f"[UI WARN] 행 {r}: 빈 데이터")
                continue
            
            print(f"[UI] 저장 시도: {name} ({cid})")
            if self.db.add_complex(name, cid):
                count += 1
        
        print(f"[UI] DB 저장 완료: {count}/{total}개")
        QMessageBox.information(self, "저장 완료", f"{count}개 단지가 DB에 저장되었습니다.\n\nDB 경로: {self.db.db_path}")
        self._load_db_complexes()
        self._load_stats_complexes()  # 통계 탭도 갱신
    
    def _show_db_load_dialog(self):
        complexes = self.db.get_all_complexes()
        if not complexes:
            QMessageBox.information(self, "알림", "저장된 단지가 없습니다.")
            return
        items = [(f"{name} ({cid})", (name, cid)) for _, name, cid, _ in complexes]
        dlg = MultiSelectDialog("DB에서 불러오기", items, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            for name, cid in dlg.selected_items():
                self._add_row(name, cid)
    
    def _show_group_load_dialog(self):
        groups = self.db.get_all_groups()
        if not groups:
            QMessageBox.information(self, "알림", "저장된 그룹이 없습니다.")
            return
        items = [(name, gid) for gid, name, _ in groups]
        dlg = MultiSelectDialog("그룹에서 불러오기", items, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            for gid in dlg.selected_items():
                for _, name, cid, _ in self.db.get_complexes_in_group(gid):
                    self._add_row(name, cid)
    
    def _show_history_dialog(self):
        dlg = RecentSearchDialog(self, self.history_manager)
        if dlg.exec() == QDialog.DialogCode.Accepted and dlg.selected_search:
            search = dlg.selected_search
            self._clear_list()
            for name, cid in search.get('complexes', []):
                self._add_row(name, cid)
            # 거래유형 복원
            types = search.get('trade_types', [])
            self.check_trade.setChecked("매매" in types)
            self.check_jeonse.setChecked("전세" in types)
            self.check_monthly.setChecked("월세" in types)
    
    def _open_complex_url(self):
        row = self.table_list.currentRow()
        if row >= 0:
            cid = self.table_list.item(row, 1).text()
            webbrowser.open(get_complex_url(cid))
    
    def _open_db_complex_url(self):
        row = self.db_table.currentRow()
        if row >= 0:
            cid = self.db_table.item(row, 2).text()
            webbrowser.open(get_complex_url(cid))
    
    def _open_article_url(self):
        """결과 테이블에서 더블클릭 시 매물 URL 열기"""
        row = self.result_table.currentRow()
        if row >= 0:
            # URL은 인덱스 10에 저장됨 (숨겨진 컬럼)
            url_item = self.result_table.item(row, 10)
            if url_item and url_item.text():
                webbrowser.open(url_item.text())
    
    def _filter_results(self, text):
        for r in range(self.result_table.rowCount()):
            match = any(text.lower() in (self.result_table.item(r, c).text().lower() if self.result_table.item(r, c) else "") for c in range(7))
            self.result_table.setRowHidden(r, not match)
    
    def _filter_db_table(self, text):
        for r in range(self.db_table.rowCount()):
            match = any(text.lower() in (self.db_table.item(r, c).text().lower() if self.db_table.item(r, c) else "") for c in range(4))
            self.db_table.setRowHidden(r, not match)
    
    def _sort_results(self, sort_text):
        col_map = {"가격": 2, "면적": 3, "단지명": 0}
        for key, col in col_map.items():
            if key in sort_text:
                order = Qt.SortOrder.AscendingOrder if "↑" in sort_text else Qt.SortOrder.DescendingOrder
                self.result_table.sortItems(col, order)
                break
    
    def _start_crawling(self):
        # 이전 크롤러 스레드가 실행 중이면 안전하게 종료
        if self.crawler_thread and self.crawler_thread.isRunning():
            get_logger('RealEstateApp').warning("이전 크롤러가 실행 중, 종료 대기...")
            self.crawler_thread.stop()
            self.crawler_thread.wait(3000)  # 최대 3초 대기
        
        tgs = [(self.table_list.item(r, 0).text(), self.table_list.item(r, 1).text()) for r in range(self.table_list.rowCount())]
        if not tgs:
            QMessageBox.warning(self, "알림", "단지를 추가해주세요.")
            return
        tts = []
        if self.check_trade.isChecked(): tts.append("매매")
        if self.check_jeonse.isChecked(): tts.append("전세")
        if self.check_monthly.isChecked(): tts.append("월세")
        if not tts:
            QMessageBox.warning(self, "알림", "거래유형을 선택해주세요.")
            return
        
        # 검색 기록 저장
        self.history_manager.add({
            'complexes': tgs,
            'trade_types': tts
        })
        
        af = {"enabled": self.check_area_filter.isChecked(), "min": self.spin_area_min.value(), "max": self.spin_area_max.value()}
        pf = {"enabled": self.check_price_filter.isChecked(), 
              "매매": {"min": self.spin_trade_min.value(), "max": self.spin_trade_max.value()}, 
              "전세": {"min": self.spin_jeonse_min.value(), "max": self.spin_jeonse_max.value()}, 
              "월세": {"min": self.spin_monthly_min.value(), "max": self.spin_monthly_max.value()}}
        
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.btn_save.setEnabled(False)
        self.log_browser.clear()
        self.result_table.setRowCount(0)
        self.progress_widget.reset()
        self.summary_card.reset()
        self.grouped_rows.clear()
        self.collected_data.clear()
        self.crawl_stats = {"매매": 0, "전세": 0, "월세": 0, "new": 0, "price_up": 0, "price_down": 0}
        self.advanced_filters = None  # 필터 초기화
        
        self.crawler_thread = CrawlerThread(tgs, tts, af, pf, self.db, self.speed_slider.current_speed())
        self.crawler_thread.log_signal.connect(self._update_log)
        self.crawler_thread.progress_signal.connect(self._update_progress)
        self.crawler_thread.item_signal.connect(self._add_result)
        self.crawler_thread.stats_signal.connect(self._update_stats)
        self.crawler_thread.complex_finished_signal.connect(self._on_complex_done)
        self.crawler_thread.finished_signal.connect(self._crawling_done)
        self.crawler_thread.error_signal.connect(self._crawling_error)
        self.crawler_thread.start()
        self.status_bar.showMessage("🚀 크롤링 진행 중...")
    
    def _stop_crawling(self):
        if self.crawler_thread:
            self.crawler_thread.stop()
            self.log_browser.append("\n⏹️ 사용자에 의해 중지됨")
        self.btn_stop.setEnabled(False)
    
    def _update_log(self, msg, level=20):
        c = "#ff6b6b" if level >= 40 else "#f39c12" if level >= 30 else "#00ff00"
        self.log_browser.append(f"<span style='color: {c};'>{msg}</span>")
        self.log_browser.verticalScrollBar().setValue(self.log_browser.verticalScrollBar().maximum())
    
    def _update_progress(self, percent, current_name, remaining):
        self.progress_widget.update_progress(percent, current_name, remaining)
    
    def _add_result(self, d):
        tt = d["거래유형"]
        pv = d["매매가"] if tt == "매매" else d["보증금"] if tt == "전세" else f"{d['보증금']}/{d['월세']}"
        gk = (d["단지명"], tt, pv, d["면적(평)"])
        
        # 통계 업데이트
        self.crawl_stats[tt] = self.crawl_stats.get(tt, 0) + 1
        
        # v7.3: 신규/가격변동 체크
        is_new = False
        price_change = 0
        price_change_text = ""
        
        article_id = d.get("매물ID", "")
        complex_id = d.get("단지ID", "")
        
        # 가격을 숫자로 변환
        if tt == "매매":
            current_price = PriceConverter.to_int(d.get("매매가", "0"))
        else:
            current_price = PriceConverter.to_int(d.get("보증금", "0"))
        
        if article_id and complex_id:
            is_new, price_change, prev_price = self.db.check_article_history(
                article_id, complex_id, current_price
            )
            
            # 매물 히스토리 업데이트
            self.db.update_article_history(
                article_id, complex_id, d["단지명"], tt,
                current_price, pv, d["면적(평)"],
                d.get("층/방향", ""), d.get("타입/특징", "")
            )
            
            # 가격 변동 텍스트
            if price_change > 0:
                price_change_text = f"📈 +{PriceConverter.to_string(price_change)}"
                self.crawl_stats['price_up'] = self.crawl_stats.get('price_up', 0) + 1
            elif price_change < 0:
                price_change_text = f"📉 {PriceConverter.to_string(price_change)}"
                self.crawl_stats['price_down'] = self.crawl_stats.get('price_down', 0) + 1
            
            if is_new:
                self.crawl_stats['new'] = self.crawl_stats.get('new', 0) + 1
        
        # 데이터에 추가 정보 저장
        d['is_new'] = is_new
        d['price_change'] = price_change
        d['price_int'] = current_price
        
        if gk in self.grouped_rows:
            ri = self.grouped_rows[gk]
            cur = self.result_table.item(ri, 2).text()
            m = re.search(r'\((\d+)건\)', cur)
            cnt = int(m.group(1)) + 1 if m else 2
            self.result_table.setItem(ri, 2, SortableTableWidgetItem(f"{pv} ({cnt}건)"))
        else:
            row = self.result_table.rowCount()
            self.result_table.insertRow(row)
            self.result_table.setRowHeight(row, 32)  # 행 높이 고정
            self.grouped_rows[gk] = row
            
            is_dark = self.current_theme == "dark"
            self.result_table.setItem(row, 0, QTableWidgetItem(d["단지명"]))
            self.result_table.setItem(row, 1, ColoredTableWidgetItem(tt, tt, is_dark))
            self.result_table.setItem(row, 2, SortableTableWidgetItem(str(pv)))
            self.result_table.setItem(row, 3, SortableTableWidgetItem(f"{d['면적(평)']}평"))
            self.result_table.setItem(row, 4, QTableWidgetItem(d["층/방향"]))
            self.result_table.setItem(row, 5, QTableWidgetItem(d["타입/특징"]))
            
            # v7.3: 신규 배지
            new_item = QTableWidgetItem("🆕" if is_new else "")
            if is_new:
                new_item.setBackground(QColor("#f39c12") if is_dark else QColor("#ffeaa7"))
            self.result_table.setItem(row, 6, new_item)
            
            # v7.3: 가격 변동
            change_item = QTableWidgetItem(price_change_text)
            if price_change > 0:
                change_item.setForeground(QColor("#e74c3c"))
            elif price_change < 0:
                change_item.setForeground(QColor("#27ae60"))
            self.result_table.setItem(row, 7, change_item)
            
            # 시각
            self.result_table.setItem(row, 8, QTableWidgetItem(
                d["수집시각"].split()[1] if " " in d["수집시각"] else d["수집시각"]
            ))
            
            # 링크 버튼
            url = get_article_url(d["단지ID"], d.get("매물ID", "")) if d.get("매물ID") else get_complex_url(d["단지ID"])
            link_btn = LinkButton(url)
            self.result_table.setCellWidget(row, 9, link_btn)
            self.result_table.setItem(row, 10, QTableWidgetItem(url))
            
            # 가격 숫자 (정렬용)
            self.result_table.setItem(row, 11, SortableTableWidgetItem(str(current_price)))
        
        self.collected_data.append(d)
    
    def _update_stats(self, s):
        total = s.get('total_found', 0)
        filtered = s.get('filtered_out', 0)
        self.summary_card.update_stats(
            total, 
            self.crawl_stats.get("매매", 0),
            self.crawl_stats.get("전세", 0),
            self.crawl_stats.get("월세", 0),
            filtered,
            self.crawl_stats.get("new", 0),
            self.crawl_stats.get("price_up", 0),
            self.crawl_stats.get("price_down", 0)
        )
        self.status_bar.showMessage(f"📊 수집: {total}건 | 🆕 신규: {self.crawl_stats.get('new', 0)}건 | 필터 제외: {filtered}건")
    
    def _on_complex_done(self, n, c, t, cnt):
        self.db.add_crawl_history(n, c, t, cnt)
    
    def _crawling_done(self, data):
        # 시그널 연결 해제 (메모리 누수 방지)
        if self.crawler_thread:
            try:
                self.crawler_thread.log_signal.disconnect()
                self.crawler_thread.progress_signal.disconnect()
                self.crawler_thread.item_signal.disconnect()
                self.crawler_thread.stats_signal.disconnect()
                self.crawler_thread.complex_finished_signal.disconnect()
                self.crawler_thread.finished_signal.disconnect()
                self.crawler_thread.error_signal.disconnect()
            except (TypeError, RuntimeError):
                pass  # 이미 해제되었거나 연결되지 않은 경우
        
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.btn_save.setEnabled(len(self.collected_data) > 0)
        self.progress_widget.complete()
        self.status_bar.showMessage(f"✅ 완료! 총 {len(self.collected_data)}건 수집")
        self._load_history()
        
        # 가격 스냅샷 저장 (통계용)
        self._save_price_snapshots()
        
        # 완료 알림
        if settings.get("show_notifications") and NOTIFICATION_AVAILABLE:
            try:
                notification.notify(
                    title="크롤링 완료",
                    message=f"총 {len(self.collected_data)}건 수집 완료!",
                    timeout=5
                )
            except Exception as e:
                get_logger('RealEstateApp').debug(f"알림 표시 실패: {e}")
        
        # 완료 사운드
        if settings.get("play_sound_on_complete"):
            try:
                QApplication.beep()
            except Exception as e:
                get_logger('RealEstateApp').debug(f"완료 사운드 실패: {e}")
    
    def _save_price_snapshots(self):
        """크롤링 결과를 가격 스냅샷으로 저장"""
        if not self.collected_data:
            return
        
        # 단지별, 거래유형별, 평형별로 그룹화
        from collections import defaultdict
        grouped = defaultdict(list)
        
        for item in self.collected_data:
            cid = item.get("단지ID", "")
            ttype = item.get("거래유형", "")
            pyeong = item.get("면적(평)", 0)
            
            # 가격 추출
            if ttype == "매매":
                price = PriceConverter.to_int(item.get("매매가", "0"))
            else:
                price = PriceConverter.to_int(item.get("보증금", "0"))
            
            if cid and ttype and price > 0:
                # 평형 그룹화 (5평 단위)
                pyeong_group = round(pyeong / 5) * 5
                key = (cid, ttype, pyeong_group)
                grouped[key].append(price)
        
        # 스냅샷 저장
        saved = 0
        for (cid, ttype, pyeong), prices in grouped.items():
            if prices:
                min_price = min(prices)
                max_price = max(prices)
                avg_price = sum(prices) // len(prices)
                
                if self.db.add_price_snapshot(cid, ttype, pyeong, min_price, max_price, avg_price, len(prices)):
                    saved += 1
        
        print(f"[UI] 가격 스냅샷 저장: {saved}건")
    
    def _crawling_error(self, err):
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        QMessageBox.critical(self, "오류", f"크롤링 중 오류 발생:\n{err}")
    
    def _show_save_menu(self):
        menu = QMenu(self)
        menu.addAction("📊 Excel로 저장", self._save_excel)
        menu.addAction("📄 CSV로 저장", self._save_csv)
        menu.addAction("📋 JSON으로 저장", self._save_json)
        menu.addSeparator()
        menu.addAction("⚙️ 엑셀 템플릿 설정", self._show_excel_template_dialog)
        menu.exec(self.btn_save.mapToGlobal(self.btn_save.rect().bottomLeft()))
    
    def _save_excel(self):
        if not self.collected_data: return
        path, _ = QFileDialog.getSaveFileName(self, "Excel 저장", f"부동산_{DateTimeHelper.file_timestamp()}.xlsx", "Excel (*.xlsx)")
        if path:
            # v7.3: 템플릿 적용
            template = settings.get("excel_template")
            if DataExporter(self.collected_data).to_excel(Path(path), template):
                QMessageBox.information(self, "저장 완료", f"Excel 파일 저장 완료!\n{path}")
    
    def _save_csv(self):
        if not self.collected_data: return
        path, _ = QFileDialog.getSaveFileName(self, "CSV 저장", f"부동산_{DateTimeHelper.file_timestamp()}.csv", "CSV (*.csv)")
        if path:
            if DataExporter(self.collected_data).to_csv(Path(path)):
                QMessageBox.information(self, "저장 완료", f"CSV 파일 저장 완료!\n{path}")
    
    def _save_json(self):
        if not self.collected_data: return
        path, _ = QFileDialog.getSaveFileName(self, "JSON 저장", f"부동산_{DateTimeHelper.file_timestamp()}.json", "JSON (*.json)")
        if path:
            if DataExporter(self.collected_data).to_json(Path(path)):
                QMessageBox.information(self, "저장 완료", f"JSON 파일 저장 완료!\n{path}")
    
    # DB Tab handlers
    def _load_db_complexes(self):
        """DB에서 단지 목록 로드 - 디버깅 강화"""
        print(f"[UI] DB 단지 로드 시작...")
        self.db_table.setRowCount(0)
        try:
            complexes = self.db.get_all_complexes()
            print(f"[UI] 로드된 단지: {len(complexes)}개")
            
            for db_id, name, cid, memo in complexes:
                row = self.db_table.rowCount()
                self.db_table.insertRow(row)
                self.db_table.setItem(row, 0, QTableWidgetItem(str(db_id)))
                self.db_table.setItem(row, 1, QTableWidgetItem(str(name)))
                self.db_table.setItem(row, 2, QTableWidgetItem(str(cid)))
                self.db_table.setItem(row, 3, QTableWidgetItem(str(memo) if memo else ""))
            
            print(f"[UI] DB 테이블 갱신 완료: {self.db_table.rowCount()}행")
        except Exception as e:
            print(f"[UI ERROR] DB 단지 로드 실패: {e}")
            import traceback
            traceback.print_exc()
    
    def _delete_db_complex(self):
        row = self.db_table.currentRow()
        if row >= 0:
            db_id = int(self.db_table.item(row, 0).text())
            if self.db.delete_complex(db_id):
                self._load_db_complexes()
    
    def _delete_db_complexes_multi(self):
        rows = set(item.row() for item in self.db_table.selectedItems())
        if rows:
            ids = [int(self.db_table.item(r, 0).text()) for r in rows]
            cnt = self.db.delete_complexes_bulk(ids)
            QMessageBox.information(self, "삭제 완료", f"{cnt}개 단지 삭제됨")
            self._load_db_complexes()
    
    def _edit_memo(self):
        row = self.db_table.currentRow()
        if row >= 0:
            db_id = int(self.db_table.item(row, 0).text())
            old = self.db_table.item(row, 3).text()
            new, ok = QInputDialog.getText(self, "메모 수정", "메모:", text=old)
            if ok:
                self.db.update_complex_memo(db_id, new)
                self._load_db_complexes()
    
    # Group Tab handlers
    def _load_all_groups(self):
        self.group_list.clear()
        for gid, name, desc in self.db.get_all_groups():
            item = QListWidgetItem(f"{name} ({desc})" if desc else name)
            item.setData(Qt.ItemDataRole.UserRole, gid)
            self.group_list.addItem(item)
    
    def _create_group(self):
        name, ok = QInputDialog.getText(self, "새 그룹", "그룹 이름:")
        if ok and name:
            if self.db.create_group(name):
                self._load_all_groups()
                self._load_schedule_groups()
    
    def _delete_group(self):
        item = self.group_list.currentItem()
        if item:
            gid = item.data(Qt.ItemDataRole.UserRole)
            if self.db.delete_group(gid):
                self._load_all_groups()
                self._load_schedule_groups()
                self.group_complex_table.setRowCount(0)
    
    def _load_group_complexes(self, item):
        gid = item.data(Qt.ItemDataRole.UserRole)
        self.group_complex_table.setRowCount(0)
        for db_id, name, cid, memo in self.db.get_complexes_in_group(gid):
            row = self.group_complex_table.rowCount()
            self.group_complex_table.insertRow(row)
            self.group_complex_table.setItem(row, 0, QTableWidgetItem(str(db_id)))
            self.group_complex_table.setItem(row, 1, QTableWidgetItem(name))
            self.group_complex_table.setItem(row, 2, QTableWidgetItem(cid))
            self.group_complex_table.setItem(row, 3, QTableWidgetItem(memo or ""))
    
    def _add_to_group(self):
        group_item = self.group_list.currentItem()
        if not group_item:
            QMessageBox.warning(self, "알림", "그룹을 선택해주세요.")
            return
        gid = group_item.data(Qt.ItemDataRole.UserRole)
        complexes = self.db.get_all_complexes()
        if not complexes:
            QMessageBox.information(self, "알림", "DB에 저장된 단지가 없습니다.")
            return
        items = [(f"{name} ({cid})", db_id) for db_id, name, cid, _ in complexes]
        dlg = MultiSelectDialog("단지 추가", items, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.db.add_complexes_to_group(gid, dlg.selected_items())
            self._load_group_complexes(group_item)
    
    def _add_to_group_multi(self):
        self._add_to_group()  # 같은 기능
    
    def _remove_from_group(self):
        group_item = self.group_list.currentItem()
        if not group_item: return
        gid = group_item.data(Qt.ItemDataRole.UserRole)
        row = self.group_complex_table.currentRow()
        if row >= 0:
            db_id = int(self.group_complex_table.item(row, 0).text())
            self.db.remove_complex_from_group(gid, db_id)
            self._load_group_complexes(group_item)
    
    # Schedule handlers
    def _load_schedule_groups(self):
        self.schedule_group_combo.clear()
        for gid, name, _ in self.db.get_all_groups():
            self.schedule_group_combo.addItem(name, gid)
    
    def _check_schedule(self):
        if not self.check_schedule.isChecked(): return
        now = QTime.currentTime()
        target = self.time_edit.time()
        if abs(now.secsTo(target)) < 60 and not self.is_scheduled_run:
            self.is_scheduled_run = True
            self._run_scheduled()
    
    def _run_scheduled(self):
        gid = self.schedule_group_combo.currentData()
        if gid:
            self._clear_list()
            for _, name, cid, _ in self.db.get_complexes_in_group(gid):
                self._add_row(name, cid)
            self._start_crawling()
        self.is_scheduled_run = False
    
    # History handlers
    def _load_history(self):
        self.history_table.setRowCount(0)
        for name, cid, types, cnt, ts in self.db.get_crawl_history():
            row = self.history_table.rowCount()
            self.history_table.insertRow(row)
            self.history_table.setItem(row, 0, QTableWidgetItem(name))
            self.history_table.setItem(row, 1, QTableWidgetItem(cid))
            self.history_table.setItem(row, 2, QTableWidgetItem(types))
            self.history_table.setItem(row, 3, QTableWidgetItem(str(cnt)))
            self.history_table.setItem(row, 4, QTableWidgetItem(str(ts)))
    
    # Stats handlers
    def _load_stats_complexes(self):
        """통계 탭 단지 콤보박스 로드 - 디버깅 강화"""
        print("[UI] 통계 단지 콤보박스 로드 시작...")
        self.stats_complex_combo.clear()
        try:
            complexes = self.db.get_all_complexes()
            print(f"[UI] 통계용 단지: {len(complexes)}개")
            
            for _, name, cid, _ in complexes:
                self.stats_complex_combo.addItem(f"{name} ({cid})", cid)
            
            print(f"[UI] 통계 콤보박스 항목: {self.stats_complex_combo.count()}개")
        except Exception as e:
            print(f"[UI ERROR] 통계 단지 로드 실패: {e}")
            import traceback
            traceback.print_exc()
    
    def _load_stats(self):
        """통계 데이터 로드"""
        cid = self.stats_complex_combo.currentData()
        if not cid:
            print("[UI] 통계 조회: 선택된 단지 없음")
            return
        
        tt = self.stats_type_combo.currentText()
        pyeong = self.stats_pyeong_combo.currentText()
        print(f"[UI] 통계 조회: {cid}, 유형: {tt}, 평형: {pyeong}")
        
        self.stats_table.setRowCount(0)
        try:
            history = self.db.get_complex_price_history(cid, tt, pyeong)
            print(f"[UI] 가격 히스토리: {len(history)}건")
            
            for date, ttype, pyeong_val, minp, maxp, avgp in history:
                row = self.stats_table.rowCount()
                self.stats_table.insertRow(row)
                self.stats_table.setItem(row, 0, QTableWidgetItem(str(date)))
                self.stats_table.setItem(row, 1, QTableWidgetItem(ttype))
                self.stats_table.setItem(row, 2, QTableWidgetItem(f"{pyeong_val}평"))
                self.stats_table.setItem(row, 3, QTableWidgetItem(PriceConverter.to_string(minp) if minp else "-"))
                self.stats_table.setItem(row, 4, QTableWidgetItem(PriceConverter.to_string(maxp) if maxp else "-"))
                self.stats_table.setItem(row, 5, QTableWidgetItem(PriceConverter.to_string(avgp) if avgp else "-"))
            
            # v10.0 Update Chart
            # Extract simple (Date, AvgPrice) for chart
            chart_data = []
            for date, _, _, _, _, avgp in history:
                if avgp:
                    chart_data.append((date, avgp))
            self.chart_widget.update_chart(chart_data)

        except Exception as e:
            print(f"[UI ERROR] 통계 로드 실패: {e}")
            import traceback
            traceback.print_exc()

    def _on_stats_complex_changed(self, index):
        """단지 변경 시 평형 목록 갱신"""
        cid = self.stats_complex_combo.currentData()
        if not cid: return
        
        current_pyeong = self.stats_pyeong_combo.currentText()
        self.stats_pyeong_combo.blockSignals(True)
        self.stats_pyeong_combo.clear()
        self.stats_pyeong_combo.addItem("전체")
        
        # DB에서 해당 단지의 평형 목록 가져오기
        try:
            conn = self.db._pool.get_connection()
            rows = conn.cursor().execute(
                "SELECT DISTINCT pyeong FROM price_snapshots WHERE complex_id = ? ORDER BY pyeong", 
                (cid,)
            ).fetchall()
            self.db._pool.return_connection(conn)
            
            for r in rows:
                self.stats_pyeong_combo.addItem(f"{r[0]}평")
        except Exception as e: get_logger('RealEstateApp').debug(f"평형 목록 조회 실패: {e}")
        
        # 이전 선택 유지 시도
        idx = self.stats_pyeong_combo.findText(current_pyeong)
        if idx >= 0:
            self.stats_pyeong_combo.setCurrentIndex(idx)
            
        self.stats_pyeong_combo.blockSignals(False)

    # Settings handlers
    def _toggle_theme(self):
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        self.setStyleSheet(get_stylesheet(self.current_theme))
        self.summary_card.set_theme(self.current_theme)
        settings.set("theme", self.current_theme)
        self.status_bar.showMessage(f"🎨 테마 변경: {self.current_theme}")
    
    def _show_settings(self):
        dlg = SettingsDialog(self)
        dlg.settings_changed.connect(self._apply_settings)
        dlg.exec()
    
    def _apply_settings(self, new):
        if new.get("theme") != self.current_theme:
            self.current_theme = new["theme"]
            self.setStyleSheet(get_stylesheet(self.current_theme))
            self.summary_card.set_theme(self.current_theme)
        self.speed_slider.set_speed(new.get("crawl_speed", "보통"))
    
    def _save_preset(self):
        name, ok = QInputDialog.getText(self, "프리셋 저장", "프리셋 이름:")
        if ok and name:
            cfg = {
                "trade_types": {"매매": self.check_trade.isChecked(), "전세": self.check_jeonse.isChecked(), "월세": self.check_monthly.isChecked()},
                "area_filter": {"enabled": self.check_area_filter.isChecked(), "min": self.spin_area_min.value(), "max": self.spin_area_max.value()},
                "price_filter": {
                    "enabled": self.check_price_filter.isChecked(),
                    "매매_min": self.spin_trade_min.value(), "매매_max": self.spin_trade_max.value(),
                    "전세_min": self.spin_jeonse_min.value(), "전세_max": self.spin_jeonse_max.value(),
                    "월세_min": self.spin_monthly_min.value(), "월세_max": self.spin_monthly_max.value()
                }
            }
            self.preset_manager.add(name, cfg)
            QMessageBox.information(self, "저장 완료", f"프리셋 '{name}' 저장됨")
    
    def _load_preset(self):
        dlg = PresetDialog(self, self.preset_manager)
        if dlg.exec() == QDialog.DialogCode.Accepted and dlg.selected_preset:
            cfg = self.preset_manager.get(dlg.selected_preset)
            if cfg:
                self.check_trade.setChecked(cfg["trade_types"]["매매"])
                self.check_jeonse.setChecked(cfg["trade_types"]["전세"])
                self.check_monthly.setChecked(cfg["trade_types"]["월세"])
                self.check_area_filter.setChecked(cfg["area_filter"]["enabled"])
                self.spin_area_min.setValue(cfg["area_filter"]["min"])
                self.spin_area_max.setValue(cfg["area_filter"]["max"])
                self.check_price_filter.setChecked(cfg["price_filter"]["enabled"])
                self.spin_trade_min.setValue(cfg["price_filter"]["매매_min"])
                self.spin_trade_max.setValue(cfg["price_filter"]["매매_max"])
                self.spin_jeonse_min.setValue(cfg["price_filter"]["전세_min"])
                self.spin_jeonse_max.setValue(cfg["price_filter"]["전세_max"])
                self.spin_monthly_min.setValue(cfg["price_filter"]["월세_min"])
                self.spin_monthly_max.setValue(cfg["price_filter"]["월세_max"])
    
    def _show_alert_settings(self):
        AlertSettingDialog(self, self.db).exec()
    
    def _show_shortcuts(self):
        ShortcutsDialog(self).exec()
    
    def _show_about(self):
        AboutDialog(self).exec()
    
    # ========== v7.3 신규 메서드 ==========
    
    def _show_advanced_filter(self):
        """고급 필터 다이얼로그"""
        dlg = AdvancedFilterDialog(self)
        dlg.filter_applied.connect(self._apply_advanced_filter)
        dlg.exec()
    
    def _apply_advanced_filter(self, filters):
        """고급 필터 적용"""
        self.advanced_filters = filters
        self._filter_results_advanced()
    
    def _filter_results_advanced(self):
        """고급 필터로 결과 필터링"""
        if not hasattr(self, 'advanced_filters') or not self.advanced_filters:
            return
        
        f = self.advanced_filters
        hidden_count = 0
        
        for row in range(self.result_table.rowCount()):
            show = True
            
            # 가격 필터
            price_item = self.result_table.item(row, 11)  # 가격 숫자 컬럼
            if price_item:
                price = int(price_item.text()) if price_item.text().isdigit() else 0
                if f['price_min'] > 0 and price < f['price_min']:
                    show = False
                if f['price_max'] < 9999999 and price > f['price_max']:
                    show = False
            
            # 면적 필터
            if show:
                area_item = self.result_table.item(row, 3)
                if area_item:
                    area_text = area_item.text().replace("평", "")
                    try:
                        area = float(area_text)
                        if f['area_min'] > 0 and area < f['area_min']:
                            show = False
                        if f['area_max'] < 500 and area > f['area_max']:
                            show = False
                    except (ValueError, TypeError):
                        pass
            
            # 층수 필터
            if show:
                floor_item = self.result_table.item(row, 4)
                if floor_item:
                    floor_text = floor_item.text()
                    if "저층" in floor_text and not f['floor_low']:
                        show = False
                    elif "중층" in floor_text and not f['floor_mid']:
                        show = False
                    elif "고층" in floor_text and not f['floor_high']:
                        show = False
            
            # 특수 필터
            if show and f['only_new']:
                new_item = self.result_table.item(row, 6)
                if not new_item or "🆕" not in new_item.text():
                    show = False
            
            if show and f['only_price_down']:
                change_item = self.result_table.item(row, 7)
                if not change_item or "📉" not in change_item.text():
                    show = False
            
            if show and f['only_price_change']:
                change_item = self.result_table.item(row, 7)
                if not change_item or change_item.text() == "":
                    show = False
            
            # 키워드 필터
            if show and f['include_keywords']:
                row_text = " ".join([
                    self.result_table.item(row, c).text() if self.result_table.item(row, c) else ""
                    for c in range(6)
                ])
                if not any(kw in row_text for kw in f['include_keywords']):
                    show = False
            
            if show and f['exclude_keywords']:
                row_text = " ".join([
                    self.result_table.item(row, c).text() if self.result_table.item(row, c) else ""
                    for c in range(6)
                ])
                if any(kw in row_text for kw in f['exclude_keywords']):
                    show = False
            
            self.result_table.setRowHidden(row, not show)
            if not show:
                hidden_count += 1
        
        visible = self.result_table.rowCount() - hidden_count
        self.status_bar.showMessage(f"🔍 필터 적용: {visible}건 표시 / {hidden_count}건 숨김")
    
    def _show_url_batch_dialog(self):
        """URL 일괄 등록 다이얼로그"""
        dlg = URLBatchDialog(self)
        dlg.complexes_added.connect(self._add_complexes_from_url)
        dlg.exec()
    
    def _add_complexes_from_url(self, complexes):
        """URL에서 추출한 단지 추가"""
        for name, cid in complexes:
            self._add_row(name, cid)
        QMessageBox.information(self, "추가 완료", f"{len(complexes)}개 단지가 추가되었습니다.")
    
    def _show_excel_template_dialog(self):
        """엑셀 템플릿 설정 다이얼로그"""
        current = settings.get("excel_template")
        dlg = ExcelTemplateDialog(self, current)
        dlg.template_saved.connect(self._save_excel_template)
        dlg.exec()
    
    def _save_excel_template(self, template):
        """엑셀 템플릿 저장"""
        settings.set("excel_template", template)
        QMessageBox.information(self, "저장 완료", "엑셀 템플릿이 저장되었습니다.")
    
    def _backup_db(self):
        path, _ = QFileDialog.getSaveFileName(self, "DB 백업", f"backup_{DateTimeHelper.file_timestamp()}.db", "Database (*.db)")
        if path:
            if self.db.backup_database(Path(path)):
                QMessageBox.information(self, "백업 완료", f"DB 백업 완료!\n{path}")
    
    def _restore_db(self):
        """DB 복원 - 안전한 UI 처리"""
        path, _ = QFileDialog.getOpenFileName(self, "DB 복원", "", "Database (*.db)")
        if not path:
            return
        
        # 확인 대화상자
        reply = QMessageBox.question(
            self, "DB 복원 확인",
            f"현재 DB를 선택한 파일로 교체합니다.\n\n"
            f"복원 파일: {path}\n\n"
            f"계속하시겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # 진행 중 표시
        self.status_bar.showMessage("🔄 DB 복원 중...")
        QApplication.processEvents()
        
        try:
            print(f"[UI] DB 복원 시작: {path}")
            
            if self.db.restore_database(Path(path)):
                # 성공 시 모든 데이터 다시 로드
                print("[UI] DB 복원 성공, 데이터 다시 로드 중...")
                
                try:
                    self._load_db_complexes()
                except Exception as e:
                    print(f"[UI WARN] _load_db_complexes 실패: {e}")
                
                try:
                    self._load_all_groups()
                except Exception as e:
                    print(f"[UI WARN] _load_all_groups 실패: {e}")
                
                try:
                    self._load_history()
                except Exception as e:
                    print(f"[UI WARN] _load_history 실패: {e}")
                
                try:
                    self._load_stats_complexes()
                except Exception as e:
                    print(f"[UI WARN] _load_stats_complexes 실패: {e}")
                
                try:
                    self._load_schedule_groups()
                except Exception as e:
                    print(f"[UI WARN] _load_schedule_groups 실패: {e}")
                
                self.status_bar.showMessage("✅ DB 복원 완료!")
                QMessageBox.information(self, "복원 완료", "DB 복원이 완료되었습니다!")
                print("[UI] DB 복원 완료")
            else:
                self.status_bar.showMessage("❌ DB 복원 실패")
                QMessageBox.critical(self, "복원 실패", "DB 복원에 실패했습니다.\n콘솔 로그를 확인하세요.")
                print("[UI] DB 복원 실패")
                
        except Exception as e:
            print(f"[UI ERROR] DB 복원 중 예외: {e}")
            import traceback
            traceback.print_exc()
            self.status_bar.showMessage("❌ DB 복원 중 오류 발생")
            QMessageBox.critical(self, "오류", f"DB 복원 중 오류가 발생했습니다:\n{e}")
    
    def _refresh_tab(self):
        idx = self.tabs.currentIndex()
        if idx == 1: self._load_db_complexes()
        elif idx == 2: self._load_all_groups()
        elif idx == 4: self._load_history()
        elif idx == 5: self._load_stats_complexes()
    
    def _focus_search(self):
        idx = self.tabs.currentIndex()
        if idx == 0: self.result_search.setFocus()
        elif idx == 1: self.db_search.setFocus()
    
    def _minimize_to_tray(self):
        if self.tray_icon:
            self.hide()
            self.tray_icon.showMessage(APP_TITLE, "트레이로 최소화됨", QSystemTrayIcon.MessageIcon.Information, 2000)
    
    def _show_from_tray(self):
        self.showNormal()
        self.activateWindow()
    
    def _tray_activated(self, reason):
        if reason == QSystemTrayIcon.ActivationReason.DoubleClick:
            self._show_from_tray()
    
    def _quit_app(self):
        if settings.get("confirm_before_close"):
            if QMessageBox.question(self, "종료", "정말 종료하시겠습니까?") != QMessageBox.StandardButton.Yes:
                return
        
        # 스레드 안전 종료
        if self.crawler_thread and self.crawler_thread.isRunning():
            get_logger('RealEstateApp').info("크롤러 스레드 종료 중...")
            self.crawler_thread.stop()
            if not self.crawler_thread.wait(5000):  # 5초 대기
                get_logger('RealEstateApp').warning("크롤러 스레드 강제 종료")
        
        # 타이머 정리
        if hasattr(self, 'schedule_timer') and self.schedule_timer:
            self.schedule_timer.stop()
            get_logger('RealEstateApp').debug("예약 타이머 종료")
        
        # DB 연결 정리
        if hasattr(self, 'db') and self.db:
            try:
                self.db._pool.close_all()
            except Exception as e:
                get_logger('RealEstateApp').warning(f"DB 연결 풀 종료 중 오류: {e}")
        
        # 설정 저장
        settings.set("window_geometry", [self.x(), self.y(), self.width(), self.height()])
        get_logger('RealEstateApp').info("프로그램 종료")
        QApplication.quit()
    
    def closeEvent(self, event):
        if settings.get("minimize_to_tray") and self.tray_icon:
            event.ignore()
            self._minimize_to_tray()
        else:
            event.accept()
            self._quit_app()
    
    # v11.0: Toast \uc54c\ub9bc \uba54\uc11c\ub4dc
    def show_toast(self, message: str, toast_type: str = "info", duration: int = 3000):
        """\ube44\uce68\uc2b5\uc801 Toast \uc54c\ub9bc \ud45c\uc2dc"""
        toast = ToastWidget(message, toast_type, self)
        self.toast_widgets.append(toast)
        
        # \uc704\uce58 \uc124\uc815 (\uc6b0\uce21 \ud558\ub2e8)
        self._reposition_toasts()
        toast.show_toast(duration)
    
    def _reposition_toasts(self):
        """\ubaa8\ub4e0 Toast \uc704\uc82f \uc7ac\ubc30\uce58"""
        y_offset = self.height() - 20
        for i, toast in enumerate(reversed(self.toast_widgets)):
            if toast and not toast.isHidden():
                x = self.width() - toast.width() - 20
                y = y_offset - toast.height()
                toast.move(self.mapToGlobal(QPoint(x, y)))
                y_offset = y - 10  # \uac04\uaca9

# ============ MAIN ============
def main():
    import multiprocessing
    multiprocessing.freeze_support()
    
    # v11.0: Windows 콘솔 UTF-8 인코딩 설정 (이모지 출력용)
    import sys
    import io
    if sys.platform == 'win32':
        # IDE 환경에서는 이미 래핑되어 있을 수 있으므로 buffer 속성 확인
        try:
            if hasattr(sys.stdout, 'buffer'):
                sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
            if hasattr(sys.stderr, 'buffer'):
                sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
        except (AttributeError, OSError) as e:
            # 이미 UTF-8로 설정되어 있거나 래핑할 수 없는 환경
            pass
    
    print(f"\n{'='*60}")
    print(f"  {APP_TITLE}")
    print(f"{'='*60}")
    print(f"[INIT] 시작 시간: {DateTimeHelper.now_string()}")
    print(f"[INIT] 기본 디렉토리: {BASE_DIR}")
    print(f"[INIT] 데이터 디렉토리: {DATA_DIR}")
    print(f"[INIT] DB 파일: {DB_PATH}")
    print(f"[INIT] DB 존재 여부: {DB_PATH.exists()}")
    
    logger = setup_logger()
    logger.info(f"=== {APP_TITLE} 시작 ===")
    
    app = QApplication(sys.argv)
    
    # v11.0: PyQt6에서는 HiDPI가 기본 활성화되어 있음
    # Qt6에서 AA_EnableHighDpiScaling은 deprecated됨
    
    app.setStyle("Fusion")
    
    font = QFont("Malgun Gothic", 9)
    app.setFont(font)
    
    app.setQuitOnLastWindowClosed(False)
    
    try:
        print("[INIT] 메인 윈도우 생성 중...")
        window = RealEstateApp()
        print("[INIT] 메인 윈도우 생성 완료")
        window.show()
        print("[INIT] 윈도우 표시 완료")
    except Exception as e:
        print(f"[CRITICAL] 메인 윈도우 생성 실패: {e}")
        logger.critical(f"메인 윈도우 생성 실패: {e}")
        import traceback
        traceback.print_exc()
        logger.critical(traceback.format_exc())
        sys.exit(1)
    
    code = app.exec()
    print(f"[INIT] 종료 코드: {code}")
    logger.info(f"=== 종료 (code: {code}) ===")
    sys.exit(code)

if __name__ == "__main__":
    main()
