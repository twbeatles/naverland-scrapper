
import json
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Dict

# Import from converters.py and re-export for backward compatibility
from src.utils.converters import PriceConverter, AreaConverter

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class DateTimeHelper:
    @staticmethod
    def now_string(fmt="%Y-%m-%d %H:%M:%S"):
        return datetime.now().strftime(fmt)

    @staticmethod
    def file_timestamp():
        return datetime.now().strftime("%Y%m%d_%H%M%S")

class DataExporter:
    def __init__(self, data: List[Dict]):
        self.data = data
        
    def to_excel(self, path: Path):
        if not OPENPYXL_AVAILABLE:
            print("openpyxl not available")
            return False
            
        try:
            wb = Workbook()
            ws = wb.active
            if not self.data:
                wb.save(path)
                return True
                
            headers = list(self.data[0].keys())
            ws.append(headers)
            
            # Style headers
            for col_num, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                
            for row_data in self.data:
                row = [row_data.get(h, "") for h in headers]
                ws.append(row)
                
            # Auto-width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 50)
                
            wb.save(path)
            return True
        except Exception as e:
            print(f"Export Excel Error: {e}")
            return False

    def to_csv(self, path: Path):
        try:
            if not self.data:
                with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                    pass
                return True
                
            headers = list(self.data[0].keys())
            with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(self.data)
            return True
        except Exception as e:
            print(f"Export CSV Error: {e}")
            return False

    def to_json(self, path: Path):
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"Export JSON Error: {e}")
            return False
