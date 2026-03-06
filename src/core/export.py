import csv
import json
from src.utils.helpers import PriceConverter, DateTimeHelper
from src.utils.logger import get_logger

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
logger = get_logger("Export")

class ExcelTemplate:
    """엑셀 내보내기 템플릿 (v7.3)"""
    DEFAULT_COLUMNS = [
        ("단지명", True),
        ("자산유형", True),
        ("거래유형", True),
        ("매매가", True),
        ("보증금", True),
        ("월세", True),
        ("면적(㎡)", True),
        ("면적(평)", True),
        ("평당가_표시", True),
        ("층/방향", True),
        ("타입/특징", True),
        ("기전세금(원)", True),
        ("갭금액(원)", True),
        ("갭비율", True),
        ("부동산상호", False),
        ("중개사이름", False),
        ("전화1", False),
        ("전화2", False),
        ("수집모드", False),
        ("위도", False),
        ("경도", False),
        ("줌", False),
        ("마커ID", False),
        ("매물ID", False),
        ("단지ID", False),
        ("수집시각", True),
        ("신규여부", False),
        ("가격변동", False),
    ]
    
    @staticmethod
    def get_column_order():
        return [
            "단지명", "자산유형", "거래유형", "매매가", "보증금", "월세", 
            "면적(㎡)", "면적(평)", "평당가_표시", "층/방향", "타입/특징",
            "기전세금(원)", "갭금액(원)", "갭비율", "부동산상호", "중개사이름",
            "전화1", "전화2", "수집모드", "위도", "경도", "줌", "마커ID",
            "매물ID", "단지ID", "수집시각", "신규여부", "가격변동"
        ]
    
    @staticmethod
    def get_default_template():
        return {col: True for col in ExcelTemplate.get_column_order()}

class DataExporter:
    # v12.0: 확장된 컬럼 (평당가, 신규, 가격변동 포함)
    COLUMNS = [
        "단지명", "자산유형", "거래유형", "매매가", "보증금", "월세", 
        "면적(㎡)", "면적(평)", "평당가_표시", "층/방향", "타입/특징",
        "기전세금(원)", "갭금액(원)", "갭비율", "부동산상호", "중개사이름",
        "전화1", "전화2", "수집모드", "위도", "경도", "줌", "마커ID",
        "매물ID", "단지ID", "수집시각", "신규여부", "가격변동"
    ]
    
    def __init__(self, data): 
        self.data = data

    @staticmethod
    def _change_to_int(value):
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if value is None:
            return 0
        text = str(value).strip()
        if not text:
            return 0
        sign = -1 if text.startswith("-") else 1
        cleaned = text.lstrip("+-")
        parsed = PriceConverter.to_int(cleaned)
        return sign * parsed

    @classmethod
    def _format_price_change(cls, value):
        return PriceConverter.to_signed_string(cls._change_to_int(value), zero_text="")

    @staticmethod
    def _format_gap_ratio(value):
        try:
            ratio = float(value or 0)
        except (TypeError, ValueError):
            ratio = 0.0
        return f"{ratio:.4f}" if ratio else ""
    
    def to_excel(self, path, template=None):
        """엑셀로 내보내기 - 템플릿 지원 (v7.3)"""
        if not OPENPYXL_AVAILABLE: return None
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "매물 데이터"
            
            # 템플릿에서 컬럼 결정
            if template and 'order' in template and 'columns' in template:
                columns = [c for c in template['order'] if template['columns'].get(c, False)]
            else:
                columns = ["단지명", "거래유형", "매매가", "보증금", "월세", 
                          "면적(㎡)", "면적(평)", "층/방향", "타입/특징", "수집시각"]
            
            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # 헤더 작성
            for col, h in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # 거래유형별 색상
            trade_colors = {
                "매매": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
                "전세": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
                "월세": PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            }
            
            # v7.3: 신규/가격변동 색상
            new_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            price_up_font = Font(color="FF0000")
            price_down_font = Font(color="008000")
            
            # 데이터 작성
            for ri, item in enumerate(self.data, 2):
                tt = item.get("거래유형", "")
                fill = trade_colors.get(tt)
                
                for ci, cn in enumerate(columns, 1):
                    # 특수 컬럼 처리
                    if cn == "신규여부":
                        value = "🆕 신규" if item.get('is_new', False) else ""
                    elif cn == "가격변동":
                        value = self._format_price_change(item.get('price_change', 0))
                    elif cn == "갭비율":
                        value = self._format_gap_ratio(item.get("갭비율", 0))
                    else:
                        value = item.get(cn, "")
                    
                    cell = ws.cell(row=ri, column=ci, value=value)
                    
                    # 스타일 적용
                    if fill:
                        cell.fill = fill
                    
                    # 신규 매물 강조
                    if item.get('is_new', False) and cn == "단지명":
                        cell.fill = new_fill
                    
                    # 가격 변동 색상
                    if cn == "가격변동":
                        pc = self._change_to_int(item.get('price_change', 0))
                        if pc > 0:
                            cell.font = price_up_font
                        elif pc < 0:
                            cell.font = price_down_font
            
            # 컬럼 너비 설정
            for col in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            ws.freeze_panes = "A2"
            wb.save(path)
            return path
        except Exception as e:
            logger.error(f"Excel 저장 실패: {e}")
            return None
    
    def to_csv(self, path, template=None):
        """CSV로 내보내기 - 템플릿 지원"""
        try:
            # 템플릿에서 컬럼 결정
            if template and 'order' in template and 'columns' in template:
                columns = [c for c in template['order'] if template['columns'].get(c, False)]
            else:
                columns = ["단지명", "거래유형", "매매가", "보증금", "월세", 
                          "면적(㎡)", "면적(평)", "층/방향", "타입/특징", "수집시각"]
            
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')
                writer.writeheader()
                
                for item in self.data:
                    # 특수 컬럼 추가
                    row = dict(item)
                    row['신규여부'] = "신규" if item.get('is_new', False) else ""
                    row['가격변동'] = self._format_price_change(item.get('price_change', 0))
                    row['갭비율'] = self._format_gap_ratio(item.get('갭비율', 0))
                    writer.writerow(row)
            return path
        except Exception as e:
            logger.error(f"CSV 저장 실패: {e}")
            return None
    
    def to_json(self, path):
        """JSON으로 내보내기"""
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump({
                    "exported_at": DateTimeHelper.now_string(), 
                    "total_count": len(self.data),
                    "new_count": sum(1 for d in self.data if d.get('is_new', False)),
                    "price_change_count": sum(1 for d in self.data if d.get('price_change', 0) != 0),
                    "data": self.data
                }, f, ensure_ascii=False, indent=2)
            return path
        except Exception as e:
            logger.error(f"JSON 저장 실패: {e}")
            return None
