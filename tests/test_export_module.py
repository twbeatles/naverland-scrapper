import csv
import os
import tempfile
import unittest
from unittest.mock import patch

from src.core.export import DataExporter, ExcelTemplate, ExportResult, OPENPYXL_AVAILABLE
from src.utils.helpers import PriceConverter


class TestExportModule(unittest.TestCase):
    def test_price_converter_to_signed_string(self):
        self.assertEqual(PriceConverter.to_signed_string(1500), "+1,500만")
        self.assertEqual(PriceConverter.to_signed_string(-1500), "-1,500만")
        self.assertEqual(PriceConverter.to_signed_string(0), "")
        self.assertEqual(PriceConverter.to_signed_string(0, zero_text="-"), "-")

    def test_csv_exports_price_change_with_sign(self):
        data = [
            {"단지명": "A", "가격변동": 0, "price_change": 1500},
            {"단지명": "B", "가격변동": 0, "price_change": -2500},
            {"단지명": "C", "가격변동": 0, "price_change": 0},
        ]
        template = {
            "order": ["단지명", "가격변동"],
            "columns": {"단지명": True, "가격변동": True},
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "export.csv")
            out = DataExporter(data).to_csv(path, template=template)
            self.assertEqual(out, path)
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))
        self.assertEqual(rows[0]["가격변동"], "+1,500만")
        self.assertEqual(rows[1]["가격변동"], "-2,500만")
        self.assertEqual(rows[2]["가격변동"], "")

    def test_export_result_bool_and_last_error_on_csv_failure(self):
        exporter = DataExporter([{"단지명": "A"}])
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "missing", "export.csv")
            result = exporter.export_csv(path)

        self.assertIsInstance(result, ExportResult)
        self.assertFalse(result)
        self.assertFalse(result.ok)
        self.assertIsNone(result.path)
        self.assertIn("CSV 저장 실패", result.error)
        self.assertEqual(exporter.last_error, result.error)
        self.assertIsNone(exporter.to_csv(path))

    def test_excel_export_result_reports_missing_openpyxl(self):
        exporter = DataExporter([{"단지명": "A"}])
        with patch("src.core.export.OPENPYXL_AVAILABLE", False):
            result = exporter.export_excel("export.xlsx")

        self.assertFalse(result.ok)
        self.assertIn("openpyxl", result.error)
        self.assertEqual(exporter.last_error, result.error)
        with patch("src.core.export.OPENPYXL_AVAILABLE", False):
            self.assertIsNone(exporter.to_excel("export.xlsx"))

    def test_csv_default_columns_use_default_template(self):
        data = [
            {
                "단지명": "A",
                "자산유형": "APT",
                "거래유형": "매매",
                "갭금액(원)": 10000000,
                "갭비율": 0.25,
                "상세수집상태": "success",
                "매물ID": "A1",
                "price_change": 1500,
            }
        ]
        expected = [col for col, enabled in ExcelTemplate.DEFAULT_COLUMNS if enabled]
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "export.csv")
            out = DataExporter(data).to_csv(path)
            self.assertEqual(out, path)
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                rows = list(reader)

        self.assertEqual(reader.fieldnames, expected)
        self.assertIn("자산유형", reader.fieldnames or [])
        self.assertIn("갭금액(원)", reader.fieldnames or [])
        self.assertIn("갭비율", reader.fieldnames or [])
        self.assertNotIn("상세수집상태", reader.fieldnames or [])
        self.assertNotIn("매물ID", reader.fieldnames or [])
        self.assertNotIn("가격변동", reader.fieldnames or [])
        self.assertEqual(rows[0]["갭비율"], "0.2500")

    def test_csv_explicit_template_can_enable_optional_columns(self):
        data = [{"단지명": "A", "상세수집상태": "success", "매물ID": "A1", "price_change": -1500}]
        template = {
            "order": ["단지명", "상세수집상태", "매물ID", "가격변동"],
            "columns": {"단지명": True, "상세수집상태": True, "매물ID": True, "가격변동": True},
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "export.csv")
            out = DataExporter(data).to_csv(path, template=template)
            self.assertEqual(out, path)
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        self.assertEqual(rows[0]["상세수집상태"], "success")
        self.assertEqual(rows[0]["매물ID"], "A1")
        self.assertEqual(rows[0]["가격변동"], "-1,500만")

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not installed")
    def test_excel_exports_price_change_with_sign(self):
        from openpyxl import load_workbook

        data = [{"단지명": "A", "price_change": -3500}]
        template = {
            "order": ["단지명", "가격변동"],
            "columns": {"단지명": True, "가격변동": True},
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "export.xlsx")
            out = DataExporter(data).to_excel(path, template=template)
            self.assertEqual(out, path)
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            assert ws is not None
            value = ws.cell(row=2, column=2).value
            wb.close()

        self.assertEqual(value, "-3,500만")

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not installed")
    def test_excel_default_columns_use_default_template(self):
        from openpyxl import load_workbook

        data = [{"단지명": "A", "자산유형": "APT", "거래유형": "매매", "갭금액(원)": 10000000}]
        expected = [col for col, enabled in ExcelTemplate.DEFAULT_COLUMNS if enabled]
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "export.xlsx")
            out = DataExporter(data).to_excel(path)
            self.assertEqual(out, path)
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            assert ws is not None
            headers = [ws.cell(row=1, column=i).value for i in range(1, len(expected) + 1)]
            wb.close()

        self.assertEqual(headers, expected)


if __name__ == "__main__":
    unittest.main()
